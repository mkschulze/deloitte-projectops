"""
Deloitte TaxOps Calendar - Business Logic Services

This module contains service classes that encapsulate complex business logic,
keeping routes thin and models focused on data representation.
"""

from datetime import datetime, timedelta
from enum import Enum
from dataclasses import dataclass
from typing import Optional, List, Tuple
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from extensions import db
from models import Task, TaskReviewer, User

# Logger for email operations
email_logger = logging.getLogger('email_service')


class ApprovalResult(Enum):
    """Result of an approval action"""
    SUCCESS = "success"
    ALREADY_APPROVED = "already_approved"
    ALREADY_REJECTED = "already_rejected"
    NOT_A_REVIEWER = "not_a_reviewer"
    WRONG_STATUS = "wrong_status"
    ALL_APPROVED = "all_approved"
    TASK_REJECTED = "task_rejected"


@dataclass
class ApprovalStatus:
    """Status summary for task approval workflow"""
    total_reviewers: int
    approved_count: int
    rejected_count: int
    pending_count: int
    is_complete: bool
    is_rejected: bool
    progress_percent: int
    pending_reviewers: List[User]
    approved_reviewers: List[User]
    rejected_reviewers: List[User]


class ApprovalService:
    """
    Centralized service for multi-stage approval workflow.
    
    Workflow Rules:
    - All assigned reviewers must approve for task to be approved
    - Any single rejection immediately rejects the entire task
    - Task must be in 'in_review' status for reviewer actions
    - Team members can also review if reviewer_team is assigned
    """
    
    @staticmethod
    def get_approval_status(task: Task) -> ApprovalStatus:
        """
        Get comprehensive approval status for a task.
        
        Returns:
            ApprovalStatus dataclass with all approval metrics
        """
        reviewers = task.reviewers.all()
        total = len(reviewers)
        
        approved = [tr for tr in reviewers if tr.has_approved]
        rejected = [tr for tr in reviewers if tr.has_rejected]
        pending = [tr for tr in reviewers if not tr.has_approved and not tr.has_rejected]
        
        approved_count = len(approved)
        rejected_count = len(rejected)
        pending_count = len(pending)
        
        is_complete = approved_count == total and total > 0
        is_rejected = rejected_count > 0
        
        progress_percent = int((approved_count / total * 100)) if total > 0 else 0
        
        return ApprovalStatus(
            total_reviewers=total,
            approved_count=approved_count,
            rejected_count=rejected_count,
            pending_count=pending_count,
            is_complete=is_complete,
            is_rejected=is_rejected,
            progress_percent=progress_percent,
            pending_reviewers=[tr.user for tr in pending],
            approved_reviewers=[tr.user for tr in approved],
            rejected_reviewers=[tr.user for tr in rejected]
        )
    
    @staticmethod
    def can_user_review(task: Task, user: User) -> Tuple[bool, str]:
        """
        Check if user can perform review actions on this task.
        
        Returns:
            Tuple of (can_review: bool, reason: str)
        """
        # Check task status
        if task.status != 'in_review':
            return False, "Task must be in 'in_review' status"
        
        # Check direct reviewer assignment
        tr = TaskReviewer.query.filter_by(task_id=task.id, user_id=user.id).first()
        if tr:
            if tr.has_approved:
                return False, "You have already approved this task"
            if tr.has_rejected:
                return False, "You have already rejected this task"
            return True, "Direct reviewer"
        
        # Check team membership
        if task.reviewer_team and task.reviewer_team.is_member(user):
            return True, "Reviewer via team membership"
        
        return False, "You are not a reviewer for this task"
    
    @staticmethod
    def get_reviewer_record(task: Task, user: User) -> Optional[TaskReviewer]:
        """Get or create TaskReviewer record for team-based reviewer"""
        tr = TaskReviewer.query.filter_by(task_id=task.id, user_id=user.id).first()
        
        # If not direct reviewer but in reviewer team, create a record
        if not tr and task.reviewer_team and task.reviewer_team.is_member(user):
            # Add as reviewer dynamically
            max_order = task.reviewers.count()
            tr = TaskReviewer(
                task_id=task.id,
                user_id=user.id,
                order=max_order + 1
            )
            db.session.add(tr)
            db.session.flush()
        
        return tr
    
    @staticmethod
    def approve(task: Task, user: User, note: Optional[str] = None) -> Tuple[ApprovalResult, str]:
        """
        Process approval action by a reviewer.
        
        Returns:
            Tuple of (result: ApprovalResult, message: str)
        """
        can_review, reason = ApprovalService.can_user_review(task, user)
        if not can_review:
            if "already approved" in reason.lower():
                return ApprovalResult.ALREADY_APPROVED, reason
            if "already rejected" in reason.lower():
                return ApprovalResult.ALREADY_REJECTED, reason
            if "in_review" in reason.lower():
                return ApprovalResult.WRONG_STATUS, reason
            return ApprovalResult.NOT_A_REVIEWER, reason
        
        tr = ApprovalService.get_reviewer_record(task, user)
        if not tr:
            return ApprovalResult.NOT_A_REVIEWER, "Could not create reviewer record"
        
        # Record approval
        tr.approve(note)
        
        # Check if all reviewers have now approved
        status = ApprovalService.get_approval_status(task)
        if status.is_complete:
            # Auto-transition to approved
            task.status = 'approved'
            task.approved_by_id = user.id
            task.approved_at = datetime.utcnow()
            return ApprovalResult.ALL_APPROVED, f"All {status.total_reviewers} reviewers approved. Task is now approved."
        
        remaining = status.pending_count
        return ApprovalResult.SUCCESS, f"Your approval was recorded. {remaining} reviewer(s) still pending."
    
    @staticmethod
    def reject(task: Task, user: User, note: Optional[str] = None) -> Tuple[ApprovalResult, str]:
        """
        Process rejection action by a reviewer.
        
        Returns:
            Tuple of (result: ApprovalResult, message: str)
        """
        can_review, reason = ApprovalService.can_user_review(task, user)
        if not can_review:
            if "already approved" in reason.lower():
                return ApprovalResult.ALREADY_APPROVED, reason
            if "already rejected" in reason.lower():
                return ApprovalResult.ALREADY_REJECTED, reason
            if "in_review" in reason.lower():
                return ApprovalResult.WRONG_STATUS, reason
            return ApprovalResult.NOT_A_REVIEWER, reason
        
        tr = ApprovalService.get_reviewer_record(task, user)
        if not tr:
            return ApprovalResult.NOT_A_REVIEWER, "Could not create reviewer record"
        
        # Record rejection
        tr.reject(note)
        
        # Auto-transition to rejected (any rejection = task rejected)
        task.status = 'rejected'
        task.rejected_by_id = user.id
        task.rejected_at = datetime.utcnow()
        task.rejection_reason = note
        
        return ApprovalResult.TASK_REJECTED, "Task has been rejected and returned for revision."
    
    @staticmethod
    def reset_approvals(task: Task) -> int:
        """
        Reset all approvals for a task (e.g., when resubmitting after rejection).
        
        Returns:
            Number of approvals reset
        """
        count = 0
        for tr in task.reviewers.all():
            if tr.has_approved or tr.has_rejected:
                tr.reset()
                count += 1
        return count
    
    @staticmethod
    def get_next_action_info(task: Task) -> dict:
        """
        Get information about what action is needed next.
        
        Returns:
            Dict with next_action, action_by, and details
        """
        status_actions = {
            'draft': {
                'action': 'submit',
                'action_label_de': 'Einreichen',
                'action_label_en': 'Submit',
                'by': 'owner',
                'by_user': task.owner
            },
            'submitted': {
                'action': 'start_review',
                'action_label_de': 'PrÃ¼fung starten',
                'action_label_en': 'Start Review',
                'by': 'reviewer',
                'by_user': None  # Multiple possible
            },
            'in_review': {
                'action': 'approve_or_reject',
                'action_label_de': 'Genehmigen/Ablehnen',
                'action_label_en': 'Approve/Reject',
                'by': 'reviewer',
                'by_user': None  # Pending reviewers
            },
            'approved': {
                'action': 'complete',
                'action_label_de': 'AbschlieÃŸen',
                'action_label_en': 'Complete',
                'by': 'owner_or_approver',
                'by_user': task.owner or task.approver
            },
            'completed': {
                'action': 'none',
                'action_label_de': 'Abgeschlossen',
                'action_label_en': 'Completed',
                'by': 'none',
                'by_user': None
            },
            'rejected': {
                'action': 'revise_and_resubmit',
                'action_label_de': 'Ãœberarbeiten und erneut einreichen',
                'action_label_en': 'Revise and Resubmit',
                'by': 'owner',
                'by_user': task.owner
            }
        }
        
        result = status_actions.get(task.status, {
            'action': 'unknown',
            'action_label_de': 'Unbekannt',
            'action_label_en': 'Unknown',
            'by': 'unknown',
            'by_user': None
        })
        
        # Add pending reviewers info for in_review status
        if task.status == 'in_review':
            approval_status = ApprovalService.get_approval_status(task)
            result['pending_reviewers'] = approval_status.pending_reviewers
            result['approval_progress'] = approval_status.progress_percent
        
        return result
    
    @staticmethod
    def format_approval_summary(task: Task, lang: str = 'de') -> str:
        """
        Get a human-readable approval summary.
        
        Args:
            task: Task object
            lang: Language code ('de' or 'en')
            
        Returns:
            Formatted summary string
        """
        status = ApprovalService.get_approval_status(task)
        
        if status.total_reviewers == 0:
            return "Keine PrÃ¼fer zugewiesen" if lang == 'de' else "No reviewers assigned"
        
        if status.is_rejected:
            rejector = status.rejected_reviewers[0].name if status.rejected_reviewers else "Unknown"
            return f"Abgelehnt von {rejector}" if lang == 'de' else f"Rejected by {rejector}"
        
        if status.is_complete:
            return "Alle PrÃ¼fer haben genehmigt" if lang == 'de' else "All reviewers approved"
        
        if lang == 'de':
            return f"{status.approved_count}/{status.total_reviewers} genehmigt, {status.pending_count} ausstehend"
        return f"{status.approved_count}/{status.total_reviewers} approved, {status.pending_count} pending"


class WorkflowService:
    """
    Service for general workflow operations beyond approval.
    """
    
    @staticmethod
    def submit_for_review(task: Task, user: User) -> Tuple[bool, str]:
        """
        Submit a task for review.
        
        Returns:
            Tuple of (success: bool, message: str)
        """
        if task.status != 'draft':
            return False, "Task must be in draft status to submit"
        
        if not task.can_transition_to('submitted', user):
            return False, "You do not have permission to submit this task"
        
        # Reset any previous approvals
        ApprovalService.reset_approvals(task)
        
        task.status = 'submitted'
        task.submitted_at = datetime.utcnow()
        task.submitted_by_id = user.id
        
        return True, "Task submitted for review"
    
    @staticmethod
    def start_review(task: Task, user: User) -> Tuple[bool, str]:
        """
        Start reviewing a task.
        
        Returns:
            Tuple of (success: bool, message: str)
        """
        if task.status != 'submitted':
            return False, "Task must be submitted to start review"
        
        if not task.can_transition_to('in_review', user):
            return False, "You do not have permission to start review"
        
        task.status = 'in_review'
        task.reviewed_at = datetime.utcnow()
        task.reviewed_by_id = user.id
        
        return True, "Review started"
    
    @staticmethod
    def complete_task(task: Task, user: User, note: Optional[str] = None) -> Tuple[bool, str]:
        """
        Mark task as completed.
        
        Returns:
            Tuple of (success: bool, message: str)
        """
        if task.status != 'approved':
            return False, "Task must be approved before completion"
        
        if not task.can_transition_to('completed', user):
            return False, "You do not have permission to complete this task"
        
        task.status = 'completed'
        task.completed_at = datetime.utcnow()
        task.completed_by_id = user.id
        if note:
            task.completion_note = note
        
        return True, "Task completed successfully"
    
    @staticmethod
    def restart_task(task: Task, user: User) -> Tuple[bool, str]:
        """
        Restart a rejected task (back to draft).
        
        Returns:
            Tuple of (success: bool, message: str)
        """
        if task.status != 'rejected':
            return False, "Only rejected tasks can be restarted"
        
        if not task.can_transition_to('draft', user):
            return False, "You do not have permission to restart this task"
        
        # Reset all approvals for fresh start
        ApprovalService.reset_approvals(task)
        
        task.status = 'draft'
        task.rejection_reason = None
        
        return True, "Task restarted - ready for revision"
    
    @staticmethod
    def get_workflow_timeline(task: Task) -> List[dict]:
        """
        Get timeline of workflow events for a task.
        
        Returns:
            List of event dictionaries with timestamp, action, user
        """
        timeline = []
        
        # Created
        if task.created_at:
            timeline.append({
                'timestamp': task.created_at,
                'action': 'created',
                'action_label_de': 'Erstellt',
                'action_label_en': 'Created',
                'user': None,
                'icon': 'bi-plus-circle',
                'color': 'secondary'
            })
        
        # Submitted
        if task.submitted_at:
            timeline.append({
                'timestamp': task.submitted_at,
                'action': 'submitted',
                'action_label_de': 'Eingereicht',
                'action_label_en': 'Submitted',
                'user': task.submitted_by,
                'icon': 'bi-send',
                'color': 'primary'
            })
        
        # Review started
        if task.reviewed_at:
            timeline.append({
                'timestamp': task.reviewed_at,
                'action': 'review_started',
                'action_label_de': 'PrÃ¼fung gestartet',
                'action_label_en': 'Review Started',
                'user': task.reviewed_by,
                'icon': 'bi-eye',
                'color': 'info'
            })
        
        # Individual reviewer actions
        for tr in task.reviewers.all():
            if tr.has_approved and tr.approved_at:
                timeline.append({
                    'timestamp': tr.approved_at,
                    'action': 'reviewer_approved',
                    'action_label_de': f'Genehmigt von {tr.user.name}',
                    'action_label_en': f'Approved by {tr.user.name}',
                    'user': tr.user,
                    'note': tr.approval_note,
                    'icon': 'bi-check-circle',
                    'color': 'success'
                })
            elif tr.has_rejected and tr.rejected_at:
                timeline.append({
                    'timestamp': tr.rejected_at,
                    'action': 'reviewer_rejected',
                    'action_label_de': f'Abgelehnt von {tr.user.name}',
                    'action_label_en': f'Rejected by {tr.user.name}',
                    'user': tr.user,
                    'note': tr.rejection_note,
                    'icon': 'bi-x-circle',
                    'color': 'danger'
                })
        
        # Approved
        if task.approved_at:
            timeline.append({
                'timestamp': task.approved_at,
                'action': 'approved',
                'action_label_de': 'Genehmigt',
                'action_label_en': 'Approved',
                'user': task.approved_by,
                'icon': 'bi-check-circle-fill',
                'color': 'success'
            })
        
        # Rejected
        if task.rejected_at:
            timeline.append({
                'timestamp': task.rejected_at,
                'action': 'rejected',
                'action_label_de': 'Abgelehnt',
                'action_label_en': 'Rejected',
                'user': task.rejected_by,
                'note': task.rejection_reason,
                'icon': 'bi-x-circle-fill',
                'color': 'danger'
            })
        
        # Completed
        if task.completed_at:
            timeline.append({
                'timestamp': task.completed_at,
                'action': 'completed',
                'action_label_de': 'Abgeschlossen',
                'action_label_en': 'Completed',
                'user': task.completed_by,
                'note': task.completion_note,
                'icon': 'bi-flag-fill',
                'color': 'success'
            })
        
        # Sort by timestamp
        timeline.sort(key=lambda x: x['timestamp'] or datetime.min)
        
        return timeline


# ============================================================================
# NOTIFICATION SERVICE
# ============================================================================

class NotificationService:
    """
    Service for creating and managing in-app notifications.
    Integrates with SocketIO for real-time delivery.
    """
    
    @staticmethod
    def create(user_id: int, notification_type: str, title_de: str, title_en: str,
               message_de: str = None, message_en: str = None,
               entity_type: str = None, entity_id: int = None,
               actor_id: int = None) -> 'Notification':
        """
        Create a notification for a user.
        
        Args:
            user_id: Target user ID
            notification_type: Type from NotificationType enum
            title_de: German title
            title_en: English title
            message_de: German message (optional)
            message_en: English message (optional)
            entity_type: Related entity type ('task', 'comment', etc.)
            entity_id: Related entity ID
            actor_id: User who triggered the notification
            
        Returns:
            Created Notification object (not yet committed)
        """
        from models import Notification
        
        notification = Notification(
            user_id=user_id,
            notification_type=notification_type,
            title=title_de,  # Default fallback
            title_de=title_de,
            title_en=title_en,
            message=message_de,
            message_de=message_de,
            message_en=message_en,
            entity_type=entity_type,
            entity_id=entity_id,
            actor_id=actor_id
        )
        db.session.add(notification)
        return notification
    
    @staticmethod
    def notify_users(user_ids: List[int], notification_type: str, 
                     title_de: str, title_en: str, **kwargs) -> List['Notification']:
        """
        Create notifications for multiple users.
        
        Args:
            user_ids: List of user IDs to notify
            notification_type: Type from NotificationType enum
            title_de: German title
            title_en: English title
            **kwargs: Additional args passed to create()
            
        Returns:
            List of created Notification objects
        """
        notifications = []
        for user_id in set(user_ids):  # Deduplicate
            if user_id:  # Skip None values
                n = NotificationService.create(
                    user_id, notification_type, 
                    title_de, title_en, **kwargs
                )
                notifications.append(n)
        return notifications
    
    @staticmethod
    def get_unread_count(user_id: int) -> int:
        """Get count of unread notifications for a user."""
        from models import Notification
        return Notification.query.filter_by(user_id=user_id, is_read=False).count()
    
    @staticmethod
    def get_recent(user_id: int, limit: int = 10, include_read: bool = True) -> List['Notification']:
        """
        Get recent notifications for a user.
        
        Args:
            user_id: User ID
            limit: Maximum number to return (default 10)
            include_read: Whether to include read notifications
            
        Returns:
            List of Notification objects, newest first
        """
        from models import Notification
        query = Notification.query.filter_by(user_id=user_id)
        if not include_read:
            query = query.filter_by(is_read=False)
        return query.order_by(Notification.created_at.desc()).limit(limit).all()
    
    @staticmethod
    def get_all_paginated(user_id: int, page: int = 1, per_page: int = 20):
        """
        Get all notifications for a user with pagination.
        
        Args:
            user_id: User ID
            page: Page number (1-based)
            per_page: Items per page
            
        Returns:
            Pagination object
        """
        from models import Notification
        return Notification.query.filter_by(user_id=user_id)\
            .order_by(Notification.created_at.desc())\
            .paginate(page=page, per_page=per_page, error_out=False)
    
    @staticmethod
    def mark_as_read(notification_id: int, user_id: int) -> bool:
        """
        Mark a single notification as read.
        
        Args:
            notification_id: Notification ID
            user_id: User ID (for ownership verification)
            
        Returns:
            True if marked, False if not found
        """
        from models import Notification
        notification = Notification.query.filter_by(
            id=notification_id, user_id=user_id
        ).first()
        if notification:
            notification.mark_as_read()
            return True
        return False
    
    @staticmethod
    def mark_all_as_read(user_id: int) -> int:
        """
        Mark all notifications as read for a user.
        
        Args:
            user_id: User ID
            
        Returns:
            Number of notifications marked
        """
        from models import Notification
        count = Notification.query.filter_by(
            user_id=user_id, is_read=False
        ).update({'is_read': True, 'read_at': datetime.utcnow()})
        return count
    
    @staticmethod
    def delete_old_notifications(days: int = 30) -> int:
        """
        Delete notifications older than specified days.
        Useful for cleanup/retention.
        
        Args:
            days: Delete notifications older than this many days
            
        Returns:
            Number of notifications deleted
        """
        from models import Notification
        from datetime import timedelta
        cutoff = datetime.utcnow() - timedelta(days=days)
        count = Notification.query.filter(Notification.created_at < cutoff).delete()
        return count
    
    # =========================================================================
    # NOTIFICATION CREATION HELPERS
    # =========================================================================
    
    @staticmethod
    def notify_task_assigned(task, assignee_id: int, actor_id: int) -> 'Notification':
        """Create notification when task is assigned to a user."""
        return NotificationService.create(
            user_id=assignee_id,
            notification_type='task_assigned',
            title_de=f'Neue Aufgabe zugewiesen: {task.title}',
            title_en=f'New task assigned: {task.title}',
            message_de=f'Sie wurden als Bearbeiter fÃ¼r die Aufgabe "{task.title}" zugewiesen.',
            message_en=f'You have been assigned as owner of the task "{task.title}".',
            entity_type='task',
            entity_id=task.id,
            actor_id=actor_id
        )
    
    @staticmethod
    def notify_reviewer_added(task, reviewer_id: int, actor_id: int) -> 'Notification':
        """Create notification when user is added as reviewer."""
        return NotificationService.create(
            user_id=reviewer_id,
            notification_type='reviewer_added',
            title_de=f'Review angefordert: {task.title}',
            title_en=f'Review requested: {task.title}',
            message_de=f'Sie wurden als PrÃ¼fer fÃ¼r die Aufgabe "{task.title}" hinzugefÃ¼gt.',
            message_en=f'You have been added as a reviewer for the task "{task.title}".',
            entity_type='task',
            entity_id=task.id,
            actor_id=actor_id
        )
    
    @staticmethod
    def notify_status_changed(task, user_id: int, old_status: str, new_status: str, actor_id: int) -> 'Notification':
        """Create notification when task status changes."""
        status_labels = {
            'draft': ('Entwurf', 'Draft'),
            'submitted': ('Eingereicht', 'Submitted'),
            'in_review': ('In PrÃ¼fung', 'In Review'),
            'approved': ('Genehmigt', 'Approved'),
            'rejected': ('Abgelehnt', 'Rejected'),
            'completed': ('Abgeschlossen', 'Completed')
        }
        new_de, new_en = status_labels.get(new_status, (new_status, new_status))
        
        return NotificationService.create(
            user_id=user_id,
            notification_type='task_status_changed',
            title_de=f'Status geÃ¤ndert: {task.title}',
            title_en=f'Status changed: {task.title}',
            message_de=f'Der Status wurde auf "{new_de}" geÃ¤ndert.',
            message_en=f'The status has been changed to "{new_en}".',
            entity_type='task',
            entity_id=task.id,
            actor_id=actor_id
        )
    
    @staticmethod
    def notify_task_approved(task, user_id: int, actor_id: int, note: str = None) -> 'Notification':
        """Create notification when task is approved."""
        return NotificationService.create(
            user_id=user_id,
            notification_type='task_approved',
            title_de=f'Aufgabe genehmigt: {task.title}',
            title_en=f'Task approved: {task.title}',
            message_de=note or 'Die Aufgabe wurde genehmigt.',
            message_en=note or 'The task has been approved.',
            entity_type='task',
            entity_id=task.id,
            actor_id=actor_id
        )
    
    @staticmethod
    def notify_task_rejected(task, user_id: int, actor_id: int, reason: str = None) -> 'Notification':
        """Create notification when task is rejected."""
        return NotificationService.create(
            user_id=user_id,
            notification_type='task_rejected',
            title_de=f'Aufgabe abgelehnt: {task.title}',
            title_en=f'Task rejected: {task.title}',
            message_de=reason or 'Die Aufgabe wurde abgelehnt.',
            message_en=reason or 'The task has been rejected.',
            entity_type='task',
            entity_id=task.id,
            actor_id=actor_id
        )
    
    @staticmethod
    def notify_comment_added(task, comment, recipient_id: int, actor_id: int) -> 'Notification':
        """Create notification when comment is added to task."""
        preview = (comment.content[:50] + '...') if len(comment.content) > 50 else comment.content
        return NotificationService.create(
            user_id=recipient_id,
            notification_type='task_comment',
            title_de=f'Neuer Kommentar: {task.title}',
            title_en=f'New comment: {task.title}',
            message_de=preview,
            message_en=preview,
            entity_type='task',
            entity_id=task.id,
            actor_id=actor_id
        )


# ============================================================================
# EXPORT SERVICE
# ============================================================================

class ExportService:
    """
    Service for exporting tasks to Excel and PDF formats.
    """
    
    @staticmethod
    def export_tasks_to_excel(tasks: List[Task], lang: str = 'de') -> bytes:
        """
        Export a list of tasks to Excel format.
        Returns bytes of the Excel file.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks" if lang == 'en' else "Aufgaben"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="86BC25", end_color="86BC25", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Status colors
        status_fills = {
            'draft': PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid"),
            'submitted': PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid"),
            'in_review': PatternFill(start_color="0DCAF0", end_color="0DCAF0", fill_type="solid"),
            'approved': PatternFill(start_color="198754", end_color="198754", fill_type="solid"),
            'completed': PatternFill(start_color="86BC25", end_color="86BC25", fill_type="solid"),
            'rejected': PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
        }
        
        # Headers
        headers = [
            'ID',
            'Titel' if lang == 'de' else 'Title',
            'Status',
            'FÃ¤llig' if lang == 'de' else 'Due Date',
            'Mandant' if lang == 'de' else 'Entity',
            'Steuerart' if lang == 'de' else 'Tax Type',
            'Zeitraum' if lang == 'de' else 'Period',
            'Jahr' if lang == 'de' else 'Year',
            'Bearbeiter' if lang == 'de' else 'Owner',
            'Bearbeiter-Team' if lang == 'de' else 'Owner Team',
            'PrÃ¼fer' if lang == 'de' else 'Reviewers',
            'Genehmigt' if lang == 'de' else 'Approved',
            'Beschreibung' if lang == 'de' else 'Description',
            'Erstellt' if lang == 'de' else 'Created',
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_num, task in enumerate(tasks, 2):
            # Get approval info
            approval_info = task.get_approval_count()
            reviewers = task.reviewers.all()
            reviewer_names = ', '.join([tr.user.name for tr in reviewers]) if reviewers else '-'
            
            row_data = [
                task.id,
                task.title,
                task.status,
                task.due_date.strftime('%d.%m.%Y') if task.due_date else '',
                task.entity.get_name(lang) if task.entity else '',
                task.template.tax_type.code if task.template and task.template.tax_type else '',
                task.period or '',
                task.year,
                task.owner.name if task.owner else '',
                task.owner_team.get_name(lang) if task.owner_team else '',
                reviewer_names,
                f'{approval_info[0]}/{approval_info[1]}' if approval_info[1] > 0 else '-',
                task.description or '',
                task.created_at.strftime('%d.%m.%Y %H:%M') if task.created_at else '',
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Color status cell
                if col == 3 and value in status_fills:
                    cell.fill = status_fills[value]
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Adjust column widths
        column_widths = [6, 40, 12, 12, 25, 10, 15, 8, 20, 20, 30, 10, 50, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def export_task_to_pdf(task: Task, lang: str = 'de') -> bytes:
        """
        Export a single task to PDF format.
        Returns bytes of the PDF file.
        """
        from weasyprint import HTML, CSS
        from io import BytesIO
        
        # Build HTML content
        status_colors = {
            'draft': '#6C757D',
            'submitted': '#0D6EFD',
            'in_review': '#0DCAF0',
            'approved': '#198754',
            'completed': '#86BC25',
            'rejected': '#DC3545',
        }
        
        status_labels = {
            'draft': 'Entwurf' if lang == 'de' else 'Draft',
            'submitted': 'Eingereicht' if lang == 'de' else 'Submitted',
            'in_review': 'In PrÃ¼fung' if lang == 'de' else 'In Review',
            'approved': 'Genehmigt' if lang == 'de' else 'Approved',
            'completed': 'Abgeschlossen' if lang == 'de' else 'Completed',
            'rejected': 'Abgelehnt' if lang == 'de' else 'Rejected',
        }
        
        reviewers = task.reviewers.all()
        approval_info = task.get_approval_count()
        evidence_list = task.evidence.all()
        comments_list = task.comments.order_by(db.text('created_at desc')).all()
        
        # Build reviewer HTML
        reviewer_html = ''
        for tr in reviewers:
            status_icon = 'âœ“' if tr.has_approved else ('âœ—' if tr.has_rejected else 'â—‹')
            status_class = 'approved' if tr.has_approved else ('rejected' if tr.has_rejected else 'pending')
            date_span = ''
            if tr.approved_at:
                date_span = '<span class="date">' + tr.approved_at.strftime("%d.%m.%Y %H:%M") + '</span>'
            reviewer_html += '<div class="reviewer ' + status_class + '">'
            reviewer_html += '<span class="icon">' + status_icon + '</span>'
            reviewer_html += '<span class="name">' + tr.user.name + '</span>'
            reviewer_html += date_span
            reviewer_html += '</div>'
        
        # Build evidence HTML
        evidence_html = ''
        for ev in evidence_list:
            icon = 'ðŸ“Ž' if ev.evidence_type == 'file' else 'ðŸ”—'
            evidence_html += '<div class="evidence-item">' + icon + ' ' + (ev.filename or ev.url) + '</div>'
        
        # Build comments HTML
        comments_html = ''
        for comment in comments_list:
            comments_html += '<div class="comment">'
            comments_html += '<div class="comment-header">'
            comments_html += '<strong>' + comment.user.name + '</strong>'
            comments_html += '<span class="date">' + comment.created_at.strftime("%d.%m.%Y %H:%M") + '</span>'
            comments_html += '</div>'
            comments_html += '<div class="comment-body">' + comment.content + '</div>'
            comments_html += '</div>'
        
        # Build optional sections
        overdue_html = ''
        if task.is_overdue:
            overdue_text = 'ÃœberfÃ¤llig' if lang == 'de' else 'Overdue'
            overdue_html = '<span class="overdue" style="margin-left: 10px;">âš  ' + overdue_text + '</span>'
        
        description_section = ''
        if task.description:
            desc_title = 'Beschreibung' if lang == 'de' else 'Description'
            description_section = '''
            <div class="section">
                <div class="section-title">''' + desc_title + '''</div>
                <div class="description">''' + task.description + '''</div>
            </div>
            '''
        
        evidence_section = ''
        if evidence_list:
            ev_title = 'Belege' if lang == 'de' else 'Evidence'
            evidence_section = (
                '<div class="section">'
                '<div class="section-title">' + ev_title + ' (' + str(len(evidence_list)) + ')</div>'
                + evidence_html +
                '</div>'
            )
        
        comments_section = ''
        if comments_list:
            com_title = 'Kommentare' if lang == 'de' else 'Comments'
            comments_section = (
                '<div class="section">'
                '<div class="section-title">' + com_title + ' (' + str(len(comments_list)) + ')</div>'
                + comments_html +
                '</div>'
            )
        
        reviewer_section = ''
        if reviewer_html:
            reviewer_section = '<div style="margin-top: 10px;">' + reviewer_html + '</div>'
        
        # Labels
        lbl_details = 'Details'
        lbl_due_date = 'FÃ¤lligkeitsdatum' if lang == 'de' else 'Due Date'
        lbl_entity = 'Mandant' if lang == 'de' else 'Entity'
        lbl_tax_type = 'Steuerart' if lang == 'de' else 'Tax Type'
        lbl_period = 'Zeitraum' if lang == 'de' else 'Period'
        lbl_year = 'Jahr' if lang == 'de' else 'Year'
        lbl_assignment = 'Zuweisung' if lang == 'de' else 'Assignment'
        lbl_owner = 'Bearbeiter' if lang == 'de' else 'Owner'
        lbl_owner_team = 'Bearbeiter-Team' if lang == 'de' else 'Owner Team'
        lbl_approval = 'Genehmigung' if lang == 'de' else 'Approval'
        lbl_approved = 'genehmigt' if lang == 'de' else 'approved'
        lbl_generated = 'Erstellt' if lang == 'de' else 'Generated'
        
        # Values
        due_date_val = task.due_date.strftime("%d.%m.%Y") if task.due_date else "-"
        entity_val = task.entity.get_name(lang) if task.entity else "-"
        tax_type_val = "-"
        if task.template and task.template.tax_type:
            tax_type_val = task.template.tax_type.code + " - " + task.template.tax_type.name
        period_val = task.period or "-"
        year_val = str(task.year)
        owner_val = task.owner.name if task.owner else "-"
        owner_team_val = task.owner_team.get_name(lang) if task.owner_team else "-"
        approval_val = str(approval_info[0]) + "/" + str(approval_info[1]) + " " + lbl_approved
        status_color = status_colors.get(task.status, '#6C757D')
        status_label = status_labels.get(task.status, task.status)
        generated_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        
        html_content = '''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>''' + task.title + '''</title>
            <style>
                @page {
                    size: A4;
                    margin: 2cm;
                }
                body {
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    font-size: 11pt;
                    line-height: 1.5;
                    color: #333;
                }
                .header {
                    border-bottom: 3px solid #86BC25;
                    padding-bottom: 15px;
                    margin-bottom: 20px;
                }
                .logo {
                    color: #86BC25;
                    font-size: 24pt;
                    font-weight: bold;
                }
                .title {
                    font-size: 18pt;
                    font-weight: bold;
                    margin: 10px 0;
                }
                .status-badge {
                    display: inline-block;
                    padding: 5px 15px;
                    border-radius: 4px;
                    color: white;
                    font-weight: bold;
                    background-color: ''' + status_color + ''';
                }
                .section {
                    margin: 20px 0;
                }
                .section-title {
                    font-size: 14pt;
                    font-weight: bold;
                    color: #86BC25;
                    border-bottom: 1px solid #ddd;
                    padding-bottom: 5px;
                    margin-bottom: 10px;
                }
                table.details {
                    width: 100%;
                    border-collapse: collapse;
                }
                table.details td {
                    padding: 8px;
                    border-bottom: 1px solid #eee;
                }
                table.details td:first-child {
                    width: 30%;
                    color: #666;
                    font-weight: 500;
                }
                .reviewer {
                    padding: 8px;
                    margin: 5px 0;
                    border-radius: 4px;
                    background-color: #f8f9fa;
                }
                .reviewer.approved {
                    border-left: 4px solid #86BC25;
                }
                .reviewer.rejected {
                    border-left: 4px solid #DC3545;
                }
                .reviewer.pending {
                    border-left: 4px solid #6C757D;
                }
                .reviewer .icon {
                    font-size: 14pt;
                    margin-right: 10px;
                }
                .reviewer .date {
                    color: #666;
                    font-size: 10pt;
                    margin-left: 10px;
                }
                .evidence-item {
                    padding: 5px 0;
                }
                .comment {
                    background-color: #f8f9fa;
                    padding: 10px;
                    margin: 10px 0;
                    border-radius: 4px;
                }
                .comment-header {
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 5px;
                }
                .comment-header .date {
                    color: #666;
                    font-size: 10pt;
                }
                .description {
                    background-color: #f8f9fa;
                    padding: 15px;
                    border-radius: 4px;
                    white-space: pre-wrap;
                }
                .footer {
                    margin-top: 30px;
                    padding-top: 10px;
                    border-top: 1px solid #ddd;
                    font-size: 9pt;
                    color: #666;
                    text-align: center;
                }
                .overdue {
                    color: #DC3545;
                    font-weight: bold;
                }
            </style>
        </head>
        <body>
            <div class="header">
                <div class="logo">Deloitte TaxOps Calendar</div>
                <div class="title">''' + task.title + '''</div>
                <span class="status-badge">''' + status_label + '''</span>
                ''' + overdue_html + '''
            </div>
            
            <div class="section">
                <div class="section-title">''' + lbl_details + '''</div>
                <table class="details">
                    <tr>
                        <td>''' + lbl_due_date + '''</td>
                        <td><strong>''' + due_date_val + '''</strong></td>
                    </tr>
                    <tr>
                        <td>''' + lbl_entity + '''</td>
                        <td>''' + entity_val + '''</td>
                    </tr>
                    <tr>
                        <td>''' + lbl_tax_type + '''</td>
                        <td>''' + tax_type_val + '''</td>
                    </tr>
                    <tr>
                        <td>''' + lbl_period + '''</td>
                        <td>''' + period_val + '''</td>
                    </tr>
                    <tr>
                        <td>''' + lbl_year + '''</td>
                        <td>''' + year_val + '''</td>
                    </tr>
                </table>
            </div>
            
            <div class="section">
                <div class="section-title">''' + lbl_assignment + '''</div>
                <table class="details">
                    <tr>
                        <td>''' + lbl_owner + '''</td>
                        <td>''' + owner_val + '''</td>
                    </tr>
                    <tr>
                        <td>''' + lbl_owner_team + '''</td>
                        <td>''' + owner_team_val + '''</td>
                    </tr>
                    <tr>
                        <td>''' + lbl_approval + '''</td>
                        <td>''' + approval_val + '''</td>
                    </tr>
                </table>
                ''' + reviewer_section + '''
            </div>
            
            ''' + description_section + '''
            
            ''' + evidence_section + '''
            
            ''' + comments_section + '''
            
            <div class="footer">
                ''' + lbl_generated + ''': ''' + generated_date + ''' | Deloitte TaxOps Calendar
            </div>
        </body>
        </html>
        '''
        
        # Generate PDF
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf()
        
        return pdf_bytes
    
    @staticmethod
    def export_summary_report(tasks: List[Task], lang: str = 'de') -> bytes:
        """
        Export a summary report with statistics to Excel.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import PieChart, Reference, BarChart
        from openpyxl.chart.label import DataLabelList
        from io import BytesIO
        from collections import Counter
        
        wb = Workbook()
        
        # --- Sheet 1: Summary ---
        ws_summary = wb.active
        ws_summary.title = "Summary" if lang == 'en' else "Zusammenfassung"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="86BC25", end_color="86BC25", fill_type="solid")
        
        # Status counts
        status_counts = Counter(task.status for task in tasks)
        overdue_count = sum(1 for task in tasks if task.is_overdue)
        due_soon_count = sum(1 for task in tasks if task.is_due_soon)
        
        ws_summary['A1'] = "TaxOps Calendar - " + ("Report" if lang == 'en' else "Bericht")
        ws_summary['A1'].font = Font(bold=True, size=16, color="86BC25")
        generated_label = "Generated" if lang == "en" else "Erstellt"
        ws_summary['A2'] = generated_label + ": " + datetime.now().strftime('%d.%m.%Y %H:%M')
        
        # KPIs
        ws_summary['A4'] = "Status" if lang == 'en' else "Status"
        ws_summary['B4'] = "Count" if lang == 'en' else "Anzahl"
        ws_summary['A4'].font = header_font
        ws_summary['A4'].fill = header_fill
        ws_summary['B4'].font = header_font
        ws_summary['B4'].fill = header_fill
        
        statuses = ['draft', 'submitted', 'in_review', 'approved', 'completed', 'rejected']
        status_labels = {
            'draft': 'Entwurf' if lang == 'de' else 'Draft',
            'submitted': 'Eingereicht' if lang == 'de' else 'Submitted',
            'in_review': 'In PrÃ¼fung' if lang == 'de' else 'In Review',
            'approved': 'Genehmigt' if lang == 'de' else 'Approved',
            'completed': 'Abgeschlossen' if lang == 'de' else 'Completed',
            'rejected': 'Abgelehnt' if lang == 'de' else 'Rejected',
        }
        
        for i, status in enumerate(statuses, 5):
            ws_summary[f'A{i}'] = status_labels.get(status, status)
            ws_summary[f'B{i}'] = status_counts.get(status, 0)
        
        ws_summary['A11'] = "ÃœberfÃ¤llig" if lang == 'de' else "Overdue"
        ws_summary['B11'] = overdue_count
        ws_summary['A11'].font = Font(color="DC3545", bold=True)
        
        ws_summary['A12'] = "Bald fÃ¤llig" if lang == 'de' else "Due Soon"
        ws_summary['B12'] = due_soon_count
        ws_summary['A12'].font = Font(color="FFC107", bold=True)
        
        ws_summary['A14'] = "Gesamt" if lang == 'de' else "Total"
        ws_summary['B14'] = len(tasks)
        ws_summary['A14'].font = Font(bold=True)
        ws_summary['B14'].font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 12
        
        # --- Sheet 2: By Entity ---
        ws_entity = wb.create_sheet("By Entity" if lang == 'en' else "Nach Mandant")
        entity_counts = Counter(task.entity.get_name(lang) for task in tasks if task.entity)
        
        ws_entity['A1'] = "Entity" if lang == 'en' else "Mandant"
        ws_entity['B1'] = "Tasks" if lang == 'en' else "Aufgaben"
        ws_entity['A1'].font = header_font
        ws_entity['A1'].fill = header_fill
        ws_entity['B1'].font = header_font
        ws_entity['B1'].fill = header_fill
        
        for i, (entity, count) in enumerate(entity_counts.most_common(), 2):
            ws_entity[f'A{i}'] = entity
            ws_entity[f'B{i}'] = count
        
        ws_entity.column_dimensions['A'].width = 40
        ws_entity.column_dimensions['B'].width = 12
        
        # --- Sheet 3: By Tax Type ---
        ws_taxtype = wb.create_sheet("By Tax Type" if lang == 'en' else "Nach Steuerart")
        taxtype_counts = Counter(
            task.template.tax_type.code for task in tasks 
            if task.template and task.template.tax_type
        )
        
        ws_taxtype['A1'] = "Tax Type" if lang == 'en' else "Steuerart"
        ws_taxtype['B1'] = "Tasks" if lang == 'en' else "Aufgaben"
        ws_taxtype['A1'].font = header_font
        ws_taxtype['A1'].fill = header_fill
        ws_taxtype['B1'].font = header_font
        ws_taxtype['B1'].fill = header_fill
        
        for i, (taxtype, count) in enumerate(taxtype_counts.most_common(), 2):
            ws_taxtype[f'A{i}'] = taxtype
            ws_taxtype[f'B{i}'] = count
        
        ws_taxtype.column_dimensions['A'].width = 20
        ws_taxtype.column_dimensions['B'].width = 12
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class CalendarService:
    """
    Service for generating iCal feeds for calendar synchronization.
    """
    
    @staticmethod
    def generate_user_token(user_id: int) -> str:
        """
        Generate a unique, secure token for a user's calendar feed.
        """
        import hashlib
        import secrets
        # Create a unique token based on user_id and random bytes
        random_part = secrets.token_hex(16)
        token_base = f"{user_id}-{random_part}"
        return hashlib.sha256(token_base.encode()).hexdigest()[:32]
    
    @staticmethod
    def generate_ical_feed(tasks: List[Task], user_name: str = "TaxOps User", lang: str = 'de') -> bytes:
        """
        Generate an iCal feed from a list of tasks.
        Returns bytes of the .ics file.
        """
        from icalendar import Calendar, Event, Alarm
        from datetime import timedelta
        import pytz
        
        # Create calendar
        cal = Calendar()
        cal.add('prodid', '-//Deloitte TaxOps Calendar//taxops.deloitte.com//')
        cal.add('version', '2.0')
        cal.add('calscale', 'GREGORIAN')
        cal.add('method', 'PUBLISH')
        cal.add('x-wr-calname', 'TaxOps Calendar - ' + user_name)
        cal.add('x-wr-timezone', 'Europe/Berlin')
        
        # Status labels for event descriptions
        status_labels = {
            'draft': 'Entwurf' if lang == 'de' else 'Draft',
            'submitted': 'Eingereicht' if lang == 'de' else 'Submitted',
            'in_review': 'In PrÃ¼fung' if lang == 'de' else 'In Review',
            'approved': 'Genehmigt' if lang == 'de' else 'Approved',
            'completed': 'Abgeschlossen' if lang == 'de' else 'Completed',
            'rejected': 'Abgelehnt' if lang == 'de' else 'Rejected',
        }
        
        for task in tasks:
            if not task.due_date:
                continue
                
            event = Event()
            
            # Basic event properties
            event.add('uid', f'task-{task.id}@taxops.deloitte.com')
            event.add('summary', task.title)
            
            # All-day event on the due date
            event.add('dtstart', task.due_date.date())
            event.add('dtend', task.due_date.date() + timedelta(days=1))
            
            # Description with task details
            description_parts = []
            if task.entity:
                entity_label = 'Mandant' if lang == 'de' else 'Entity'
                description_parts.append(f"{entity_label}: {task.entity.get_name(lang)}")
            
            status_label = 'Status' if lang == 'en' else 'Status'
            description_parts.append(f"{status_label}: {status_labels.get(task.status, task.status)}")
            
            if task.template and task.template.tax_type:
                taxtype_label = 'Steuerart' if lang == 'de' else 'Tax Type'
                description_parts.append(f"{taxtype_label}: {task.template.tax_type.code} - {task.template.tax_type.name}")
            
            if task.period:
                period_label = 'Zeitraum' if lang == 'de' else 'Period'
                description_parts.append(f"{period_label}: {task.period}")
            
            if task.owner:
                owner_label = 'Bearbeiter' if lang == 'de' else 'Owner'
                description_parts.append(f"{owner_label}: {task.owner.name}")
            
            if task.description:
                description_parts.append("")
                description_parts.append(task.description[:500])
            
            event.add('description', '\n'.join(description_parts))
            
            # Location (entity name)
            if task.entity:
                event.add('location', task.entity.get_name(lang))
            
            # Categories based on status
            categories = [task.status.upper()]
            if task.template and task.template.tax_type:
                categories.append(task.template.tax_type.code)
            event.add('categories', categories)
            
            # Priority based on urgency
            if task.is_overdue:
                event.add('priority', 1)  # Highest priority
            elif task.is_due_soon:
                event.add('priority', 3)  # High priority
            else:
                event.add('priority', 5)  # Normal priority
            
            # Status mapping
            if task.status == 'completed':
                event.add('status', 'COMPLETED')
            elif task.status == 'rejected':
                event.add('status', 'CANCELLED')
            else:
                event.add('status', 'CONFIRMED')
            
            # Color coding via extended property (for some calendar apps)
            if task.is_overdue:
                event.add('x-apple-calendar-color', '#DC3545')  # Red
            elif task.status == 'completed':
                event.add('x-apple-calendar-color', '#86BC25')  # Deloitte Green
            elif task.is_due_soon:
                event.add('x-apple-calendar-color', '#FFA500')  # Orange
            
            # Add alarm for upcoming tasks (1 day before)
            if task.status not in ['completed', 'rejected']:
                alarm = Alarm()
                alarm.add('action', 'DISPLAY')
                reminder_text = 'FÃ¤llig morgen' if lang == 'de' else 'Due tomorrow'
                alarm.add('description', f'{task.title} - {reminder_text}')
                alarm.add('trigger', timedelta(days=-1))
                event.add_component(alarm)
            
            # Timestamps
            event.add('dtstamp', datetime.now())
            if hasattr(task, 'created_at') and task.created_at:
                event.add('created', task.created_at)
            if hasattr(task, 'updated_at') and task.updated_at:
                event.add('last-modified', task.updated_at)
            
            cal.add_component(event)
        
        return cal.to_ical()


# ============================================================================
# EMAIL SERVICE
# ============================================================================

class EmailService:
    """
    Email notification service with multiple provider support.
    
    Providers:
    - smtp: Standard SMTP server
    - sendgrid: SendGrid API
    - ses: AWS Simple Email Service
    """
    
    def __init__(self, app=None):
        self.app = app
        self._initialized = False
    
    def init_app(self, app):
        """Initialize with Flask app"""
        self.app = app
        self._initialized = True
    
    @property
    def is_enabled(self):
        """Check if email is enabled"""
        if not self.app:
            return False
        return self.app.config.get('MAIL_ENABLED', False)
    
    @property
    def provider(self):
        """Get configured provider"""
        if not self.app:
            return 'smtp'
        return self.app.config.get('MAIL_PROVIDER', 'smtp')
    
    def send_email(self, to_email: str, subject: str, html_content: str, 
                   text_content: str = None, from_email: str = None) -> bool:
        """
        Send an email using the configured provider.
        
        Args:
            to_email: Recipient email address
            subject: Email subject
            html_content: HTML body content
            text_content: Plain text body (optional, generated from HTML if not provided)
            from_email: Sender email (uses default if not provided)
        
        Returns:
            True if sent successfully, False otherwise
        """
        if not self.is_enabled:
            email_logger.info(f"Email disabled - would send to {to_email}: {subject}")
            return True  # Return True so app logic continues
        
        if not from_email:
            sender_name = self.app.config.get('MAIL_DEFAULT_SENDER_NAME', 'TaxOps Calendar')
            sender_email = self.app.config.get('MAIL_DEFAULT_SENDER', 'noreply@deloitte.com')
            from_email = f"{sender_name} <{sender_email}>"
        
        # Generate plain text from HTML if not provided
        if not text_content:
            import re
            text_content = re.sub(r'<[^>]+>', '', html_content)
            text_content = re.sub(r'\s+', ' ', text_content).strip()
        
        try:
            if self.provider == 'sendgrid':
                return self._send_via_sendgrid(to_email, subject, html_content, text_content, from_email)
            elif self.provider == 'ses':
                return self._send_via_ses(to_email, subject, html_content, text_content, from_email)
            else:
                return self._send_via_smtp(to_email, subject, html_content, text_content, from_email)
        except Exception as e:
            email_logger.error(f"Failed to send email to {to_email}: {str(e)}")
            return False
    
    def _send_via_smtp(self, to_email, subject, html_content, text_content, from_email) -> bool:
        """Send email via SMTP"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = from_email
        msg['To'] = to_email
        
        # Attach both plain text and HTML
        part1 = MIMEText(text_content, 'plain', 'utf-8')
        part2 = MIMEText(html_content, 'html', 'utf-8')
        msg.attach(part1)
        msg.attach(part2)
        
        server = self.app.config.get('MAIL_SERVER', 'localhost')
        port = self.app.config.get('MAIL_PORT', 587)
        use_tls = self.app.config.get('MAIL_USE_TLS', True)
        use_ssl = self.app.config.get('MAIL_USE_SSL', False)
        username = self.app.config.get('MAIL_USERNAME', '')
        password = self.app.config.get('MAIL_PASSWORD', '')
        
        if use_ssl:
            smtp = smtplib.SMTP_SSL(server, port)
        else:
            smtp = smtplib.SMTP(server, port)
            if use_tls:
                smtp.starttls()
        
        if username and password:
            smtp.login(username, password)
        
        smtp.sendmail(from_email, [to_email], msg.as_string())
        smtp.quit()
        
        email_logger.info(f"Email sent via SMTP to {to_email}: {subject}")
        return True
    
    def _send_via_sendgrid(self, to_email, subject, html_content, text_content, from_email) -> bool:
        """Send email via SendGrid API"""
        try:
            from sendgrid import SendGridAPIClient
            from sendgrid.helpers.mail import Mail, Email, To, Content
        except ImportError:
            email_logger.error("SendGrid package not installed. Run: pip install sendgrid")
            return False
        
        api_key = self.app.config.get('SENDGRID_API_KEY')
        if not api_key:
            email_logger.error("SENDGRID_API_KEY not configured")
            return False
        
        message = Mail(
            from_email=from_email,
            to_emails=to_email,
            subject=subject,
            html_content=html_content,
            plain_text_content=text_content
        )
        
        sg = SendGridAPIClient(api_key)
        response = sg.send(message)
        
        email_logger.info(f"Email sent via SendGrid to {to_email}: {subject} (status: {response.status_code})")
        return response.status_code in [200, 202]
    
    def _send_via_ses(self, to_email, subject, html_content, text_content, from_email) -> bool:
        """Send email via AWS SES"""
        try:
            import boto3
            from botocore.exceptions import ClientError
        except ImportError:
            email_logger.error("boto3 package not installed. Run: pip install boto3")
            return False
        
        aws_access_key = self.app.config.get('AWS_ACCESS_KEY_ID')
        aws_secret_key = self.app.config.get('AWS_SECRET_ACCESS_KEY')
        aws_region = self.app.config.get('AWS_REGION', 'eu-central-1')
        
        client = boto3.client(
            'ses',
            region_name=aws_region,
            aws_access_key_id=aws_access_key,
            aws_secret_access_key=aws_secret_key
        )
        
        # Parse from_email to extract just the email address
        import re
        match = re.search(r'<([^>]+)>', from_email)
        source_email = match.group(1) if match else from_email
        
        response = client.send_email(
            Source=from_email,
            Destination={'ToAddresses': [to_email]},
            Message={
                'Subject': {'Data': subject, 'Charset': 'UTF-8'},
                'Body': {
                    'Text': {'Data': text_content, 'Charset': 'UTF-8'},
                    'Html': {'Data': html_content, 'Charset': 'UTF-8'}
                }
            }
        )
        
        email_logger.info(f"Email sent via SES to {to_email}: {subject} (MessageId: {response['MessageId']})")
        return True
    
    # =========================================================================
    # HIGH-LEVEL EMAIL METHODS
    # =========================================================================
    
    def send_task_assigned(self, task: 'Task', assignee: 'User', assigned_by: 'User', lang: str = 'de') -> bool:
        """Send task assignment notification"""
        if not assignee.email_notifications or not assignee.email_on_assignment:
            email_logger.debug(f"User {assignee.email} has assignment emails disabled")
            return True
        
        app_url = self.app.config.get('APP_URL', 'http://localhost:5000')
        task_url = f"{app_url}/tasks/{task.id}"
        
        if lang == 'de':
            subject = f"Neue Aufgabe zugewiesen: {task.title}"
        else:
            subject = f"New Task Assigned: {task.title}"
        
        html_content = self._render_email_template('task_assigned', {
            'task': task,
            'assignee': assignee,
            'assigned_by': assigned_by,
            'task_url': task_url,
            'lang': lang
        })
        
        return self.send_email(assignee.email, subject, html_content)
    
    def send_status_changed(self, task: 'Task', user: 'User', old_status: str, new_status: str, lang: str = 'de') -> bool:
        """Send status change notification to task owner and reviewers"""
        if not user.email_notifications or not user.email_on_status_change:
            email_logger.debug(f"User {user.email} has status change emails disabled")
            return True
        
        app_url = self.app.config.get('APP_URL', 'http://localhost:5000')
        task_url = f"{app_url}/tasks/{task.id}"
        
        if lang == 'de':
            subject = f"StatusÃ¤nderung: {task.title} - {new_status}"
        else:
            subject = f"Status Changed: {task.title} - {new_status}"
        
        html_content = self._render_email_template('status_changed', {
            'task': task,
            'user': user,
            'old_status': old_status,
            'new_status': new_status,
            'task_url': task_url,
            'lang': lang
        })
        
        return self.send_email(user.email, subject, html_content)
    
    def send_due_reminder(self, task: 'Task', user: 'User', days_until_due: int, lang: str = 'de') -> bool:
        """Send due date reminder"""
        if not user.email_notifications or not user.email_on_due_reminder:
            email_logger.debug(f"User {user.email} has due reminder emails disabled")
            return True
        
        app_url = self.app.config.get('APP_URL', 'http://localhost:5000')
        task_url = f"{app_url}/tasks/{task.id}"
        
        if lang == 'de':
            if days_until_due == 0:
                subject = f"FÃ¤llig heute: {task.title}"
            elif days_until_due < 0:
                subject = f"ÃœberfÃ¤llig: {task.title}"
            else:
                subject = f"FÃ¤llig in {days_until_due} Tagen: {task.title}"
        else:
            if days_until_due == 0:
                subject = f"Due Today: {task.title}"
            elif days_until_due < 0:
                subject = f"Overdue: {task.title}"
            else:
                subject = f"Due in {days_until_due} days: {task.title}"
        
        html_content = self._render_email_template('due_reminder', {
            'task': task,
            'user': user,
            'days_until_due': days_until_due,
            'task_url': task_url,
            'lang': lang
        })
        
        return self.send_email(user.email, subject, html_content)
    
    def send_comment_notification(self, task: 'Task', commenter: 'User', recipient: 'User', comment_text: str, lang: str = 'de') -> bool:
        """Send new comment notification"""
        if not recipient.email_notifications or not recipient.email_on_comment:
            email_logger.debug(f"User {recipient.email} has comment emails disabled")
            return True
        
        app_url = self.app.config.get('APP_URL', 'http://localhost:5000')
        task_url = f"{app_url}/tasks/{task.id}"
        
        if lang == 'de':
            subject = f"Neuer Kommentar: {task.title}"
        else:
            subject = f"New Comment: {task.title}"
        
        html_content = self._render_email_template('new_comment', {
            'task': task,
            'commenter': commenter,
            'recipient': recipient,
            'comment_text': comment_text,
            'task_url': task_url,
            'lang': lang
        })
        
        return self.send_email(recipient.email, subject, html_content)
    
    def _render_email_template(self, template_name: str, context: dict) -> str:
        """Render an email template with Deloitte branding"""
        # Base styles
        base_style = """
            body { font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }
            .container { max-width: 600px; margin: 0 auto; padding: 20px; }
            .header { background: #000000; color: #FFFFFF; padding: 20px; text-align: center; }
            .header img { height: 30px; }
            .header h1 { margin: 10px 0 0 0; font-size: 18px; font-weight: normal; }
            .content { padding: 30px 20px; background: #FFFFFF; }
            .task-card { background: #F5F5F5; border-left: 4px solid #86BC25; padding: 15px; margin: 20px 0; border-radius: 4px; }
            .task-title { font-size: 18px; font-weight: bold; color: #000; margin-bottom: 10px; }
            .task-meta { color: #666; font-size: 14px; }
            .btn { display: inline-block; padding: 12px 24px; background: #86BC25; color: #FFFFFF; text-decoration: none; border-radius: 4px; font-weight: bold; }
            .btn:hover { background: #6B9B1E; }
            .btn-danger { background: #DC3545; }
            .btn-warning { background: #FFA500; color: #000; }
            .footer { padding: 20px; text-align: center; color: #999; font-size: 12px; background: #F5F5F5; }
            .status-badge { display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: bold; }
            .status-completed { background: #E8F5E9; color: #2E7D32; }
            .status-in_review { background: #E3F2FD; color: #1565C0; }
            .status-submitted { background: #FFF3E0; color: #E65100; }
            .status-draft { background: #EEEEEE; color: #616161; }
            .status-rejected { background: #FFEBEE; color: #C62828; }
        """
        
        lang = context.get('lang', 'de')
        task = context.get('task')
        task_url = context.get('task_url', '#')
        
        # Template-specific content
        if template_name == 'task_assigned':
            assignee = context.get('assignee')
            assigned_by = context.get('assigned_by')
            
            if lang == 'de':
                greeting = f"Hallo {assignee.name},"
                intro = f"<strong>{assigned_by.name}</strong> hat Ihnen eine neue Aufgabe zugewiesen:"
                due_label = "FÃ¤lligkeitsdatum"
                entity_label = "Mandant"
                btn_text = "Aufgabe Ã¶ffnen"
            else:
                greeting = f"Hello {assignee.name},"
                intro = f"<strong>{assigned_by.name}</strong> has assigned you a new task:"
                due_label = "Due Date"
                entity_label = "Entity"
                btn_text = "Open Task"
            
            content = f"""
                <p>{greeting}</p>
                <p>{intro}</p>
                <div class="task-card">
                    <div class="task-title">{task.title}</div>
                    <div class="task-meta">
                        <p><strong>{due_label}:</strong> {task.due_date.strftime('%d.%m.%Y') if task.due_date else 'N/A'}</p>
                        <p><strong>{entity_label}:</strong> {task.entity.name if task.entity else 'N/A'}</p>
                    </div>
                </div>
                <p style="text-align: center;">
                    <a href="{task_url}" class="btn">{btn_text}</a>
                </p>
            """
        
        elif template_name == 'status_changed':
            user = context.get('user')
            old_status = context.get('old_status', '')
            new_status = context.get('new_status', '')
            
            if lang == 'de':
                greeting = f"Hallo {user.name},"
                intro = "Der Status einer Ihrer Aufgaben wurde geÃ¤ndert:"
                from_label = "Von"
                to_label = "Zu"
                btn_text = "Details anzeigen"
            else:
                greeting = f"Hello {user.name},"
                intro = "The status of one of your tasks has been changed:"
                from_label = "From"
                to_label = "To"
                btn_text = "View Details"
            
            content = f"""
                <p>{greeting}</p>
                <p>{intro}</p>
                <div class="task-card">
                    <div class="task-title">{task.title}</div>
                    <div class="task-meta">
                        <p>{from_label}: <span class="status-badge status-{old_status}">{old_status.upper()}</span></p>
                        <p>{to_label}: <span class="status-badge status-{new_status}">{new_status.upper()}</span></p>
                    </div>
                </div>
                <p style="text-align: center;">
                    <a href="{task_url}" class="btn">{btn_text}</a>
                </p>
            """
        
        elif template_name == 'due_reminder':
            user = context.get('user')
            days_until_due = context.get('days_until_due', 0)
            
            if lang == 'de':
                greeting = f"Hallo {user.name},"
                if days_until_due == 0:
                    intro = "Die folgende Aufgabe ist <strong>heute fÃ¤llig</strong>:"
                    btn_class = "btn btn-warning"
                elif days_until_due < 0:
                    intro = f"Die folgende Aufgabe ist <strong>{abs(days_until_due)} Tage Ã¼berfÃ¤llig</strong>:"
                    btn_class = "btn btn-danger"
                else:
                    intro = f"Die folgende Aufgabe ist in <strong>{days_until_due} Tagen fÃ¤llig</strong>:"
                    btn_class = "btn"
                due_label = "FÃ¤lligkeitsdatum"
                btn_text = "Aufgabe bearbeiten"
            else:
                greeting = f"Hello {user.name},"
                if days_until_due == 0:
                    intro = "The following task is <strong>due today</strong>:"
                    btn_class = "btn btn-warning"
                elif days_until_due < 0:
                    intro = f"The following task is <strong>{abs(days_until_due)} days overdue</strong>:"
                    btn_class = "btn btn-danger"
                else:
                    intro = f"The following task is due in <strong>{days_until_due} days</strong>:"
                    btn_class = "btn"
                due_label = "Due Date"
                btn_text = "Edit Task"
            
            content = f"""
                <p>{greeting}</p>
                <p>{intro}</p>
                <div class="task-card">
                    <div class="task-title">{task.title}</div>
                    <div class="task-meta">
                        <p><strong>{due_label}:</strong> {task.due_date.strftime('%d.%m.%Y') if task.due_date else 'N/A'}</p>
                    </div>
                </div>
                <p style="text-align: center;">
                    <a href="{task_url}" class="{btn_class}">{btn_text}</a>
                </p>
            """
        
        elif template_name == 'new_comment':
            recipient = context.get('recipient')
            commenter = context.get('commenter')
            comment_text = context.get('comment_text', '')
            
            if lang == 'de':
                greeting = f"Hallo {recipient.name},"
                intro = f"<strong>{commenter.name}</strong> hat einen Kommentar zu einer Aufgabe hinzugefÃ¼gt:"
                comment_label = "Kommentar"
                btn_text = "Zur Diskussion"
            else:
                greeting = f"Hello {recipient.name},"
                intro = f"<strong>{commenter.name}</strong> added a comment to a task:"
                comment_label = "Comment"
                btn_text = "View Discussion"
            
            content = f"""
                <p>{greeting}</p>
                <p>{intro}</p>
                <div class="task-card">
                    <div class="task-title">{task.title}</div>
                    <div class="task-meta">
                        <p><strong>{comment_label}:</strong></p>
                        <blockquote style="border-left: 3px solid #86BC25; padding-left: 15px; margin: 10px 0; color: #555;">
                            {comment_text[:500]}{'...' if len(comment_text) > 500 else ''}
                        </blockquote>
                    </div>
                </div>
                <p style="text-align: center;">
                    <a href="{task_url}" class="btn">{btn_text}</a>
                </p>
            """
        else:
            content = "<p>Email content</p>"
        
        # Footer
        if lang == 'de':
            footer_text = "Diese E-Mail wurde automatisch vom TaxOps Calendar generiert."
            unsubscribe_text = "E-Mail-Einstellungen verwalten"
        else:
            footer_text = "This email was automatically generated by TaxOps Calendar."
            unsubscribe_text = "Manage email settings"
        
        app_url = self.app.config.get('APP_URL', 'http://localhost:5000')
        
        # Full HTML template
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>{base_style}</style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>TaxOps Calendar</h1>
                </div>
                <div class="content">
                    {content}
                </div>
                <div class="footer">
                    <p>{footer_text}</p>
                    <p><a href="{app_url}/profile/notifications">{unsubscribe_text}</a></p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html


# Global email service instance
email_service = EmailService()


# ============================================================================
# RECURRENCE SERVICE
# ============================================================================

class RecurrenceService:
    """
    Service for managing recurring task generation from TaskPresets.
    
    Supports multiple recurrence patterns:
    - monthly: Every month on a specific day
    - quarterly: Every quarter (Q1, Q2, Q3, Q4)
    - semi_annual: Twice a year (H1, H2)
    - annual: Once a year
    - custom: Using RRULE strings for complex patterns
    """
    
    @staticmethod
    def get_period_dates(frequency: str, year: int, day_offset: int = 10) -> List[Tuple[str, 'date']]:
        """
        Generate period labels and due dates for a given frequency and year.
        
        Args:
            frequency: Recurrence frequency (monthly, quarterly, semi_annual, annual)
            year: Year to generate dates for
            day_offset: Day of period when task is due (default: 10th)
        
        Returns:
            List of tuples (period_label, due_date)
        """
        from datetime import date
        
        periods = []
        
        if frequency == 'monthly':
            for month in range(1, 13):
                # Due date is day_offset of the following month (or current if offset < 0)
                if day_offset > 0:
                    # Due on day_offset of the following month
                    due_month = month + 1
                    due_year = year
                    if due_month > 12:
                        due_month = 1
                        due_year = year + 1
                    try:
                        due_date = date(due_year, due_month, min(day_offset, 28))
                    except ValueError:
                        due_date = date(due_year, due_month, 28)
                else:
                    # Due within the same month
                    try:
                        due_date = date(year, month, abs(day_offset) if day_offset != 0 else 15)
                    except ValueError:
                        due_date = date(year, month, 28)
                
                periods.append((f'M{month:02d}', due_date))
        
        elif frequency == 'quarterly':
            quarter_months = [(1, 'Q1'), (4, 'Q2'), (7, 'Q3'), (10, 'Q4')]
            for start_month, quarter in quarter_months:
                # Due date is day_offset after quarter end
                end_month = start_month + 2  # Last month of quarter
                due_month = end_month + 1
                due_year = year
                if due_month > 12:
                    due_month = 1
                    due_year = year + 1
                try:
                    due_date = date(due_year, due_month, min(day_offset if day_offset > 0 else 15, 28))
                except ValueError:
                    due_date = date(due_year, due_month, 28)
                
                periods.append((quarter, due_date))
        
        elif frequency == 'semi_annual':
            # H1 (Jan-Jun) and H2 (Jul-Dec)
            # H1 due in July, H2 due in January next year
            periods.append(('H1', date(year, 7, min(day_offset if day_offset > 0 else 15, 28))))
            periods.append(('H2', date(year + 1, 1, min(day_offset if day_offset > 0 else 15, 28))))
        
        elif frequency == 'annual':
            # Annual task, typically due in following year
            try:
                due_date = date(year + 1, 3, min(day_offset if day_offset > 0 else 31, 28))  # Default March 31
            except ValueError:
                due_date = date(year + 1, 3, 28)
            periods.append(('', due_date))
        
        return periods
    
    @staticmethod
    def generate_tasks_from_preset(preset, year: int, entities: List = None, 
                                   owner_id: int = None, force: bool = False) -> List['Task']:
        """
        Generate task instances from a recurring preset.
        
        Args:
            preset: TaskPreset with recurrence settings
            year: Year to generate tasks for
            entities: List of entities to create tasks for (defaults to preset.default_entity)
            owner_id: Owner for generated tasks (defaults to preset.default_owner_id)
            force: If True, generate even if tasks already exist for this period
        
        Returns:
            List of created Task instances
        """
        from models import Task, Entity
        from datetime import date
        
        if not preset.is_recurring or preset.recurrence_frequency == 'none':
            return []
        
        # Determine entities
        if not entities:
            if preset.default_entity_id:
                entities = [preset.default_entity]
            else:
                # Generate for all active entities
                entities = Entity.query.filter_by(is_active=True).all()
        
        # Determine owner
        owner = owner_id or preset.default_owner_id
        
        # Get periods for this frequency
        periods = RecurrenceService.get_period_dates(
            preset.recurrence_frequency, 
            year, 
            preset.recurrence_day_offset or 10
        )
        
        # Check recurrence end date
        if preset.recurrence_end_date:
            periods = [(p, d) for p, d in periods if d <= preset.recurrence_end_date]
        
        created_tasks = []
        
        for entity in entities:
            for period, due_date in periods:
                # Check if task already exists for this preset/entity/period/year
                existing = Task.query.filter_by(
                    preset_id=preset.id,
                    entity_id=entity.id,
                    year=year,
                    period=period
                ).first()
                
                if existing and not force:
                    continue
                
                # Create new task
                task = Task(
                    title=preset.get_title('de'),
                    description=preset.get_description('de'),
                    entity_id=entity.id,
                    year=year,
                    period=period,
                    due_date=due_date,
                    status='draft',
                    owner_id=owner,
                    preset_id=preset.id,
                    is_recurring_instance=True
                )
                
                db.session.add(task)
                created_tasks.append(task)
        
        # Update last_generated_date
        if created_tasks:
            preset.last_generated_date = date.today()
        
        return created_tasks
    
    @staticmethod
    def generate_all_recurring_tasks(year: int, dry_run: bool = False) -> dict:
        """
        Generate tasks from all active recurring presets.
        
        Args:
            year: Year to generate tasks for
            dry_run: If True, don't actually create tasks, just report what would be created
        
        Returns:
            Dictionary with generation statistics
        """
        from models import TaskPreset
        
        presets = TaskPreset.query.filter_by(is_recurring=True, is_active=True).all()
        
        stats = {
            'presets_processed': 0,
            'tasks_created': 0,
            'tasks_skipped': 0,
            'errors': []
        }
        
        for preset in presets:
            try:
                tasks = RecurrenceService.generate_tasks_from_preset(preset, year)
                stats['presets_processed'] += 1
                stats['tasks_created'] += len(tasks)
            except Exception as e:
                stats['errors'].append(f"Preset {preset.id}: {str(e)}")
        
        if not dry_run and stats['tasks_created'] > 0:
            db.session.commit()
        
        return stats
    
    @staticmethod
    def parse_rrule(rrule_string: str, start_date: 'date', count: int = 12) -> List['date']:
        """
        Parse an RRULE string and return the next N occurrences.
        
        Args:
            rrule_string: RRULE string (e.g., "FREQ=MONTHLY;BYMONTHDAY=15")
            start_date: Start date for recurrence
            count: Number of occurrences to return
        
        Returns:
            List of dates
        """
        try:
            from dateutil.rrule import rrulestr
            from datetime import datetime
            
            # Parse the RRULE
            rule = rrulestr(rrule_string, dtstart=datetime.combine(start_date, datetime.min.time()))
            
            # Get next occurrences
            occurrences = list(rule[:count])
            
            return [occ.date() for occ in occurrences]
        except ImportError:
            # python-dateutil not installed
            return []
        except Exception:
            return []

