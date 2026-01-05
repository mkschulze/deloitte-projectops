"""
Presets Routes Blueprint

Handles task preset management:
- Admin preset CRUD (list, create, edit, delete)
- API preset operations (PATCH, bulk toggle, bulk delete)
- Custom field CRUD
- Import/Export functionality
- Seed from JSON files
"""

import json as json_module
import os as os_module
from datetime import datetime
from io import BytesIO
from functools import wraps
from flask import Blueprint, render_template, redirect, url_for, flash, request, session, jsonify, make_response, send_file
from flask_login import login_required, current_user

from extensions import db
from models import TaskPreset, TaskCategory, Entity, User, PresetCustomField, AuditLog
from translations import get_translation as t

presets_bp = Blueprint('presets', __name__)


def admin_required(f):
    """Decorator that requires admin role"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.role != 'admin':
            flash('Zugriff verweigert.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function


def log_action(action, entity_type, entity_id, entity_name, old_value=None, new_value=None):
    """Helper to log actions to audit log"""
    log = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_name=entity_name[:200] if entity_name else None,
        old_value=str(old_value)[:500] if old_value else None,
        new_value=str(new_value)[:500] if new_value else None
    )
    db.session.add(log)
    db.session.commit()


# ============================================================================
# ADMIN PRESET VIEWS
# ============================================================================

@presets_bp.route('/admin/presets')
@admin_required
def preset_list():
    """Task preset management"""
    category_filter = request.args.get('category', '')
    tax_type_filter = request.args.get('tax_type', '')
    search = request.args.get('search', '').strip()
    
    query = TaskPreset.query
    if category_filter:
        query = query.filter(TaskPreset.category == category_filter)
    if tax_type_filter:
        query = query.filter(TaskPreset.tax_type == tax_type_filter)
    if search:
        query = query.filter(TaskPreset.title.ilike(f'%{search}%') | 
                            TaskPreset.tax_type.ilike(f'%{search}%') |
                            TaskPreset.law_reference.ilike(f'%{search}%'))
    
    presets = query.order_by(TaskPreset.category, TaskPreset.tax_type, TaskPreset.title).all()
    
    # Get unique tax types for filter
    tax_types_used = db.session.query(TaskPreset.tax_type).filter(TaskPreset.tax_type.isnot(None)).distinct().all()
    tax_types_used = sorted([t[0] for t in tax_types_used if t[0]])
    
    return render_template('admin/presets_enhanced.html', 
                           presets=presets, 
                           category_filter=category_filter,
                           tax_type_filter=tax_type_filter,
                           search=search,
                           tax_types_used=tax_types_used)


@presets_bp.route('/admin/presets/new', methods=['GET', 'POST'])
@admin_required
def preset_new():
    """Create new task preset"""
    if request.method == 'POST':
        title_de = request.form.get('title_de', '').strip()
        title_en = request.form.get('title_en', '').strip()
        category = request.form.get('category', 'aufgabe')
        tax_type = request.form.get('tax_type', '').strip() or None
        law_reference = request.form.get('law_reference', '').strip() or None
        description_de = request.form.get('description_de', '').strip() or None
        description_en = request.form.get('description_en', '').strip() or None
        
        # Recurrence fields
        is_recurring = request.form.get('is_recurring') == 'on'
        recurrence_frequency = request.form.get('recurrence_frequency', 'monthly')
        recurrence_day_offset = int(request.form.get('recurrence_day_offset', 10) or 10)
        recurrence_rrule = request.form.get('recurrence_rrule', '').strip() or None
        recurrence_end_date_str = request.form.get('recurrence_end_date', '').strip()
        recurrence_end_date = datetime.strptime(recurrence_end_date_str, '%Y-%m-%d').date() if recurrence_end_date_str else None
        default_entity_id = int(request.form.get('default_entity_id')) if request.form.get('default_entity_id') else None
        default_owner_id = int(request.form.get('default_owner_id')) if request.form.get('default_owner_id') else None
        
        if not title_de or not title_en:
            flash('Titel (DE/EN) ist erforderlich.', 'warning')
        else:
            preset = TaskPreset(
                title=title_de,
                title_de=title_de,
                title_en=title_en,
                category=category,
                tax_type=tax_type,
                law_reference=law_reference,
                description=description_de,
                description_de=description_de,
                description_en=description_en,
                source='manual',
                is_active=True,
                is_recurring=is_recurring,
                recurrence_frequency=recurrence_frequency if is_recurring else None,
                recurrence_day_offset=recurrence_day_offset if is_recurring else None,
                recurrence_rrule=recurrence_rrule if is_recurring and recurrence_frequency == 'custom' else None,
                recurrence_end_date=recurrence_end_date if is_recurring else None,
                default_entity_id=default_entity_id if is_recurring else None,
                default_owner_id=default_owner_id if is_recurring else None
            )
            db.session.add(preset)
            db.session.commit()
            log_action('CREATE', 'TaskPreset', preset.id, preset.title[:50])
            flash('Aufgabenvorlage wurde erstellt.', 'success')
            return redirect(url_for('presets.preset_list'))
    
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    categories = TaskCategory.query.filter_by(is_active=True).order_by(TaskCategory.code).all()
    return render_template('admin/preset_form_enhanced.html', preset=None, entities=entities, users=users, categories=categories, tax_types=categories)


@presets_bp.route('/admin/presets/<int:preset_id>', methods=['GET', 'POST'])
@admin_required
def preset_edit(preset_id):
    """Edit task preset"""
    preset = TaskPreset.query.get_or_404(preset_id)
    
    if request.method == 'POST':
        preset.title_de = request.form.get('title_de', '').strip()
        preset.title_en = request.form.get('title_en', '').strip()
        preset.title = preset.title_de
        preset.category = request.form.get('category', 'aufgabe')
        preset.tax_type = request.form.get('tax_type', '').strip() or None
        preset.law_reference = request.form.get('law_reference', '').strip() or None
        preset.description_de = request.form.get('description_de', '').strip() or None
        preset.description_en = request.form.get('description_en', '').strip() or None
        preset.description = preset.description_de
        preset.is_active = request.form.get('is_active') == 'on'
        
        # Recurrence fields
        preset.is_recurring = request.form.get('is_recurring') == 'on'
        if preset.is_recurring:
            preset.recurrence_frequency = request.form.get('recurrence_frequency', 'monthly')
            preset.recurrence_day_offset = int(request.form.get('recurrence_day_offset', 10) or 10)
            recurrence_rrule = request.form.get('recurrence_rrule', '').strip()
            preset.recurrence_rrule = recurrence_rrule if preset.recurrence_frequency == 'custom' else None
            recurrence_end_date_str = request.form.get('recurrence_end_date', '').strip()
            preset.recurrence_end_date = datetime.strptime(recurrence_end_date_str, '%Y-%m-%d').date() if recurrence_end_date_str else None
            preset.default_entity_id = int(request.form.get('default_entity_id')) if request.form.get('default_entity_id') else None
            preset.default_owner_id = int(request.form.get('default_owner_id')) if request.form.get('default_owner_id') else None
        else:
            preset.recurrence_frequency = None
            preset.recurrence_day_offset = None
            preset.recurrence_rrule = None
            preset.recurrence_end_date = None
            preset.default_entity_id = None
            preset.default_owner_id = None
        
        db.session.commit()
        log_action('UPDATE', 'TaskPreset', preset.id, preset.title[:50])
        flash('Aufgabenvorlage wurde aktualisiert.', 'success')
        return redirect(url_for('presets.preset_list'))
    
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    categories = TaskCategory.query.filter_by(is_active=True).order_by(TaskCategory.code).all()
    return render_template('admin/preset_form_enhanced.html', preset=preset, entities=entities, users=users, categories=categories, tax_types=categories)


@presets_bp.route('/admin/presets/<int:preset_id>/delete', methods=['POST'])
@admin_required
def preset_delete(preset_id):
    """Delete task preset"""
    preset = TaskPreset.query.get_or_404(preset_id)
    title = preset.title
    db.session.delete(preset)
    db.session.commit()
    log_action('DELETE', 'TaskPreset', preset_id, title[:50])
    flash('Aufgabenvorlage wurde gelöscht.', 'success')
    return redirect(url_for('presets.preset_list'))


# ============================================================================
# API ENDPOINTS FOR PRESET MANAGEMENT
# ============================================================================

@presets_bp.route('/api/presets/<int:preset_id>', methods=['PATCH'])
@admin_required
def api_preset_update(preset_id):
    """Quick update preset"""
    preset = TaskPreset.query.get_or_404(preset_id)
    data = request.get_json()
    
    if 'title_de' in data:
        preset.title_de = data['title_de']
        preset.title = data['title_de']
    if 'title_en' in data:
        preset.title_en = data['title_en']
    if 'tax_type' in data:
        preset.tax_type = data['tax_type'] or None
    if 'law_reference' in data:
        preset.law_reference = data['law_reference'] or None
    if 'is_active' in data:
        preset.is_active = data['is_active']
    
    db.session.commit()
    log_action('UPDATE', 'TaskPreset', preset.id, f"Quick edit: {preset.title[:30]}")
    return jsonify({'success': True})


@presets_bp.route('/api/presets/bulk-toggle-active', methods=['POST'])
@admin_required
def api_presets_bulk_toggle():
    """Bulk activate/deactivate presets"""
    data = request.get_json()
    ids = data.get('ids', [])
    active = data.get('active', True)
    
    count = TaskPreset.query.filter(TaskPreset.id.in_(ids)).update(
        {'is_active': active}, synchronize_session=False
    )
    db.session.commit()
    log_action('BULK_UPDATE', 'TaskPreset', None, f"{'Activated' if active else 'Deactivated'} {count} presets")
    return jsonify({'success': True, 'count': count})


@presets_bp.route('/api/presets/bulk-delete', methods=['POST'])
@admin_required
def api_presets_bulk_delete():
    """Bulk delete presets"""
    data = request.get_json()
    ids = data.get('ids', [])
    
    count = TaskPreset.query.filter(TaskPreset.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    log_action('BULK_DELETE', 'TaskPreset', None, f"Deleted {count} presets")
    return jsonify({'success': True, 'count': count})


# ============================================================================
# CUSTOM FIELD CRUD API
# ============================================================================

@presets_bp.route('/api/preset-fields', methods=['POST'])
@admin_required
def api_field_create():
    """Create a new custom field for a preset"""
    data = request.get_json()
    
    preset_id = data.get('preset_id')
    if not preset_id:
        return jsonify({'error': 'preset_id required'}), 400
    
    # Get next sort order
    max_order = db.session.query(db.func.max(PresetCustomField.sort_order)).filter(
        PresetCustomField.preset_id == preset_id
    ).scalar() or 0
    
    field = PresetCustomField(
        preset_id=preset_id,
        name=data.get('name', '').lower().replace(' ', '_'),
        label_de=data.get('label_de', ''),
        label_en=data.get('label_en', ''),
        field_type=data.get('field_type', 'text'),
        is_required=data.get('is_required', False),
        placeholder_de=data.get('placeholder_de', ''),
        placeholder_en=data.get('placeholder_en', ''),
        default_value=data.get('default_value', ''),
        options=data.get('options', ''),
        help_text_de=data.get('help_text_de', ''),
        help_text_en=data.get('help_text_en', ''),
        condition_field=data.get('condition_field', ''),
        condition_operator=data.get('condition_operator', ''),
        condition_value=data.get('condition_value', ''),
        sort_order=max_order + 1
    )
    
    db.session.add(field)
    db.session.commit()
    log_action('CREATE', 'PresetCustomField', field.id, f"Created custom field {field.name}")
    
    return jsonify({'success': True, 'id': field.id})


@presets_bp.route('/api/preset-fields/<int:field_id>', methods=['GET'])
@admin_required
def api_field_get(field_id):
    """Get a custom field by ID"""
    field = PresetCustomField.query.get_or_404(field_id)
    
    return jsonify({
        'id': field.id,
        'preset_id': field.preset_id,
        'name': field.name,
        'label_de': field.label_de,
        'label_en': field.label_en,
        'field_type': field.field_type,
        'is_required': field.is_required,
        'placeholder_de': field.placeholder_de,
        'placeholder_en': field.placeholder_en,
        'default_value': field.default_value,
        'options': field.options,
        'help_text_de': field.help_text_de,
        'help_text_en': field.help_text_en,
        'condition_field': field.condition_field,
        'condition_operator': field.condition_operator,
        'condition_value': field.condition_value,
        'sort_order': field.sort_order
    })


@presets_bp.route('/api/preset-fields/<int:field_id>', methods=['PUT'])
@admin_required
def api_field_update(field_id):
    """Update a custom field"""
    field = PresetCustomField.query.get_or_404(field_id)
    data = request.get_json()
    
    field.name = data.get('name', field.name).lower().replace(' ', '_')
    field.label_de = data.get('label_de', field.label_de)
    field.label_en = data.get('label_en', field.label_en)
    field.field_type = data.get('field_type', field.field_type)
    field.is_required = data.get('is_required', field.is_required)
    field.placeholder_de = data.get('placeholder_de', field.placeholder_de)
    field.placeholder_en = data.get('placeholder_en', field.placeholder_en)
    field.default_value = data.get('default_value', field.default_value)
    field.options = data.get('options', field.options)
    field.help_text_de = data.get('help_text_de', field.help_text_de)
    field.help_text_en = data.get('help_text_en', field.help_text_en)
    field.condition_field = data.get('condition_field', field.condition_field)
    field.condition_operator = data.get('condition_operator', field.condition_operator)
    field.condition_value = data.get('condition_value', field.condition_value)
    
    db.session.commit()
    log_action('UPDATE', 'PresetCustomField', field.id, f"Updated custom field {field.name}")
    
    return jsonify({'success': True})


@presets_bp.route('/api/preset-fields/<int:field_id>', methods=['DELETE'])
@admin_required
def api_field_delete(field_id):
    """Delete a custom field"""
    field = PresetCustomField.query.get_or_404(field_id)
    
    db.session.delete(field)
    db.session.commit()
    log_action('DELETE', 'PresetCustomField', field_id, f"Deleted custom field {field.name}")
    
    return jsonify({'success': True})


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

@presets_bp.route('/admin/presets/export')
@admin_required
def preset_export():
    """Export all presets as JSON with custom fields"""
    presets = TaskPreset.query.order_by(TaskPreset.category, TaskPreset.tax_type).all()
    
    data = []
    for p in presets:
        preset_data = {
            'category': p.category,
            'tax_type': p.tax_type,
            'title_de': p.title_de,
            'title_en': p.title_en,
            'law_reference': p.law_reference,
            'description_de': p.description_de,
            'description_en': p.description_en,
            'is_recurring': p.is_recurring,
            'recurrence_frequency': p.recurrence_frequency,
            'recurrence_day_offset': p.recurrence_day_offset,
            'recurrence_rrule': p.recurrence_rrule,
            'is_active': p.is_active,
            'custom_fields': []
        }
        
        # Include custom fields
        for field in p.custom_fields.all():
            preset_data['custom_fields'].append({
                'name': field.name,
                'label_de': field.label_de,
                'label_en': field.label_en,
                'field_type': field.field_type,
                'is_required': field.is_required,
                'placeholder_de': field.placeholder_de,
                'placeholder_en': field.placeholder_en,
                'default_value': field.default_value,
                'options': field.options,
                'help_text_de': field.help_text_de,
                'help_text_en': field.help_text_en,
                'condition_field': field.condition_field,
                'condition_operator': field.condition_operator,
                'condition_value': field.condition_value,
                'sort_order': field.sort_order
            })
        
        data.append(preset_data)
    
    response = make_response(json_module.dumps(data, ensure_ascii=False, indent=2))
    response.headers['Content-Type'] = 'application/json'
    response.headers['Content-Disposition'] = 'attachment; filename=presets_export.json'
    return response


@presets_bp.route('/admin/presets/template')
@admin_required
def preset_template():
    """Download Excel template for import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aufgabenvorlagen"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['Kategorie', 'Titel', 'Steuerart', '§ Paragraph', 'Beschreibung']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    examples = [
        ['aufgabe', 'USt-Voranmeldung einreichen', 'Umsatzsteuer', '', 'Monatliche USt-Voranmeldung'],
        ['aufgabe', 'Körperschaftsteuererklärung erstellen', 'Körperschaftsteuer', '', 'Jährliche KSt-Erklärung'],
        ['antrag', 'Unbeschränkte Einkommensteuerpflicht', 'EStG', '§1 (3) EStG', 'Antrag für ausländische Mitarbeiter'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Aufgabenvorlagen_Vorlage.xlsx')


@presets_bp.route('/admin/presets/import', methods=['POST'])
@admin_required
def preset_import():
    """Import presets from JSON or Excel file"""
    if 'file' not in request.files:
        flash('Keine Datei hochgeladen.', 'error')
        return redirect(url_for('presets.preset_list'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Keine Datei ausgewählt.', 'error')
        return redirect(url_for('presets.preset_list'))
    
    filename = file.filename.lower()
    imported = 0
    skipped = 0
    errors = []
    
    try:
        if filename.endswith('.json'):
            # Import JSON
            data = json_module.load(file)
            if not isinstance(data, list):
                data = [data]
            
            for preset_data in data:
                title = preset_data.get('title_de') or preset_data.get('title')
                if not title:
                    skipped += 1
                    continue
                
                # Check for duplicates
                existing = TaskPreset.query.filter_by(
                    title_de=title,
                    category=preset_data.get('category', 'aufgabe')
                ).first()
                
                if existing:
                    skipped += 1
                    continue
                
                preset = TaskPreset(
                    category=preset_data.get('category', 'aufgabe'),
                    tax_type=preset_data.get('tax_type'),
                    title_de=title,
                    title_en=preset_data.get('title_en'),
                    law_reference=preset_data.get('law_reference'),
                    description_de=preset_data.get('description_de'),
                    description_en=preset_data.get('description_en'),
                    is_recurring=preset_data.get('is_recurring', False),
                    recurrence_frequency=preset_data.get('recurrence_frequency'),
                    recurrence_day_offset=preset_data.get('recurrence_day_offset'),
                    recurrence_rrule=preset_data.get('recurrence_rrule'),
                    is_active=preset_data.get('is_active', True),
                    source='import'
                )
                db.session.add(preset)
                db.session.flush()
                
                # Import custom fields if present
                for field_data in preset_data.get('custom_fields', []):
                    field = PresetCustomField(
                        preset_id=preset.id,
                        name=field_data.get('name', ''),
                        label_de=field_data.get('label_de'),
                        label_en=field_data.get('label_en'),
                        field_type=field_data.get('field_type', 'text'),
                        is_required=field_data.get('is_required', False),
                        placeholder_de=field_data.get('placeholder_de'),
                        placeholder_en=field_data.get('placeholder_en'),
                        default_value=field_data.get('default_value'),
                        options=field_data.get('options'),
                        help_text_de=field_data.get('help_text_de'),
                        help_text_en=field_data.get('help_text_en'),
                        condition_field=field_data.get('condition_field'),
                        condition_operator=field_data.get('condition_operator'),
                        condition_value=field_data.get('condition_value'),
                        sort_order=field_data.get('sort_order', 0)
                    )
                    db.session.add(field)
                
                imported += 1
        
        elif filename.endswith(('.xlsx', '.xls')):
            # Import Excel
            import openpyxl
            
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Get header row
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                title = row_data.get('Titel (DE)') or row_data.get('title_de') or row_data.get('Titel')
                
                if not title:
                    skipped += 1
                    continue
                
                category = row_data.get('Kategorie') or row_data.get('category') or 'aufgabe'
                existing = TaskPreset.query.filter_by(title_de=title, category=category).first()
                
                if existing:
                    skipped += 1
                    continue
                
                preset = TaskPreset(
                    category=category,
                    tax_type=row_data.get('Steuerart') or row_data.get('tax_type'),
                    title_de=title,
                    title_en=row_data.get('Titel (EN)') or row_data.get('title_en'),
                    law_reference=row_data.get('Gesetzesreferenz') or row_data.get('law_reference'),
                    description_de=row_data.get('Beschreibung (DE)') or row_data.get('description_de'),
                    description_en=row_data.get('Beschreibung (EN)') or row_data.get('description_en'),
                    is_recurring=str(row_data.get('Wiederkehrend', '')).lower() in ('ja', 'yes', 'true', '1'),
                    recurrence_frequency=row_data.get('Häufigkeit') or row_data.get('recurrence_frequency'),
                    is_active=str(row_data.get('Aktiv', 'ja')).lower() in ('ja', 'yes', 'true', '1'),
                    source='import'
                )
                db.session.add(preset)
                imported += 1
        
        else:
            flash('Nicht unterstütztes Dateiformat. Bitte JSON oder Excel verwenden.', 'error')
            return redirect(url_for('presets.preset_list'))
        
        db.session.commit()
        log_action('IMPORT', 'TaskPreset', None, f'{imported} imported, {skipped} skipped')
        flash(f'{imported} Vorlagen importiert, {skipped} übersprungen.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Import: {str(e)}', 'error')
    
    return redirect(url_for('presets.preset_list'))


@presets_bp.route('/admin/presets/seed', methods=['POST'])
@admin_required
def preset_seed():
    """Seed presets from JSON files in data folder"""
    data_dir = os_module.path.join(os_module.path.dirname(os_module.path.dirname(__file__)), 'data')
    imported = 0
    skipped = 0
    
    antraege_path = os_module.path.join(data_dir, 'Antraege.json')
    if os_module.path.exists(antraege_path):
        with open(antraege_path, 'r', encoding='utf-8') as f:
            data = json_module.load(f)
            if 'sheets' in data:
                for sheet_name, records in data['sheets'].items():
                    for record in records:
                        title = record.get('Zweck des Antrags')
                        if not title:
                            continue
                        existing = TaskPreset.query.filter_by(title=title, category='antrag').first()
                        if existing:
                            skipped += 1
                            continue
                        preset = TaskPreset(category='antrag', title=title, law_reference=record.get('§ Paragraph'),
                                          tax_type=record.get('Gesetz'), description=record.get('Erläuterung'),
                                          source='json', is_active=True)
                        db.session.add(preset)
                        imported += 1
    
    aufgaben_path = os_module.path.join(data_dir, 'steuerarten_aufgaben.json')
    if os_module.path.exists(aufgaben_path):
        with open(aufgaben_path, 'r', encoding='utf-8') as f:
            data = json_module.load(f)
            if 'records' in data:
                for record in data['records']:
                    title = record.get('aufgabe')
                    if not title:
                        continue
                    existing = TaskPreset.query.filter_by(title=title, category='aufgabe').first()
                    if existing:
                        skipped += 1
                        continue
                    preset = TaskPreset(category='aufgabe', title=title, tax_type=record.get('steuerart'),
                                      source='json', is_active=True)
                    db.session.add(preset)
                    imported += 1
    
    db.session.commit()
    log_action('SEED', 'TaskPreset', None, f'{imported} seeded, {skipped} skipped')
    flash(f'{imported} Vorlagen aus JSON-Dateien importiert, {skipped} übersprungen.', 'success')
    return redirect(url_for('presets.preset_list'))
