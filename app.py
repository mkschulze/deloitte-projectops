"""
Deloitte TaxOps Calendar
Tax Compliance Calendar & Deadline Tracking for enterprises.
"""
import json
import os
from datetime import datetime
from functools import wraps
from io import BytesIO

import click
from flask import Flask, render_template, redirect, url_for, flash, request, session, send_file, jsonify, make_response
from flask_login import login_user, logout_user, login_required, current_user
from flask_socketio import emit, join_room, leave_room

from config import config
from extensions import db, migrate, socketio, login_manager
from models import User, AuditLog, Entity, TaxType, TaskTemplate, Task, TaskEvidence, Comment, ReferenceApplication, UserRole, TaskPreset, TaskReviewer, Team, Notification, NotificationType, UserEntity, EntityAccessLevel, Module, UserModule
from translations import get_translation as t, TRANSLATIONS
from services import ApprovalService, ApprovalResult, WorkflowService, NotificationService, ExportService, CalendarService, email_service, RecurrenceService
from modules import ModuleRegistry

# Import modules to register them
import modules.core
import modules.taxops
import modules.projects

# Import project models for migrations
from modules.projects.models import Project, ProjectMember, ProjectRole


# ============================================================================
# APP INITIALIZATION
# ============================================================================

def create_app(config_name='default'):
    """Application factory"""
    app = Flask(__name__)
    app.config.from_object(config[config_name])
    
    # Initialize extensions
    db.init_app(app)
    migrate.init_app(app, db)
    
    # Initialize SocketIO with threading mode (works everywhere, no special dependencies)
    # For production with high concurrency, consider using gevent or an ASGI server
    socketio.init_app(app, cors_allowed_origins="*", async_mode='threading')
    
    # Initialize Login Manager
    login_manager.init_app(app)
    
    @login_manager.user_loader
    def load_user(user_id):
        return db.session.get(User, int(user_id))
    
    # Initialize module system
    ModuleRegistry.init_app(app)
    
    return app


app = create_app()

# Initialize email service
email_service.init_app(app)


# ============================================================================
# CONTEXT PROCESSORS
# ============================================================================

@app.context_processor
def inject_globals():
    """Inject global variables into all templates"""
    lang = session.get('lang', app.config.get('DEFAULT_LANGUAGE', 'de'))
    
    # Get user's accessible modules for navigation
    user_modules = []
    if current_user.is_authenticated:
        user_modules = ModuleRegistry.get_user_modules(current_user)
    
    return {
        'app_name': app.config.get('APP_NAME', 'MyApp'),
        'app_version': app.config.get('APP_VERSION', '1.0.0'),
        'current_year': datetime.now().year,
        'lang': lang,
        't': lambda key: t(key, lang),
        'get_file_icon': get_file_icon,
        'ApprovalService': ApprovalService,
        'WorkflowService': WorkflowService,
        'user_modules': user_modules,
        'ModuleRegistry': ModuleRegistry
    }


def get_file_icon(filename):
    """Return Bootstrap icon class based on file extension"""
    if not filename or '.' not in filename:
        return 'bi-file-earmark'
    ext = filename.rsplit('.', 1)[1].lower()
    icons = {
        'pdf': 'bi-file-earmark-pdf text-danger',
        'doc': 'bi-file-earmark-word text-primary',
        'docx': 'bi-file-earmark-word text-primary',
        'xls': 'bi-file-earmark-excel text-success',
        'xlsx': 'bi-file-earmark-excel text-success',
        'csv': 'bi-file-earmark-spreadsheet text-success',
        'txt': 'bi-file-earmark-text',
        'png': 'bi-file-earmark-image text-info',
        'jpg': 'bi-file-earmark-image text-info',
        'jpeg': 'bi-file-earmark-image text-info',
        'gif': 'bi-file-earmark-image text-info',
        'zip': 'bi-file-earmark-zip text-warning',
    }
    return icons.get(ext, 'bi-file-earmark')


# ============================================================================
# WEBSOCKET EVENTS
# ============================================================================

@socketio.on('connect')
def handle_connect():
    """Handle WebSocket connection - join user's personal room"""
    if current_user.is_authenticated:
        join_room(f'user_{current_user.id}')
        emit('connected', {'user_id': current_user.id})


@socketio.on('disconnect')
def handle_disconnect():
    """Handle WebSocket disconnection"""
    if current_user.is_authenticated:
        leave_room(f'user_{current_user.id}')


def emit_notification(user_id: int, notification, lang: str = 'de'):
    """
    Emit real-time notification to user via WebSocket.
    
    Args:
        user_id: Target user ID
        notification: Notification object
        lang: Language for localized content
    """
    socketio.emit('notification', notification.to_dict(lang), room=f'user_{user_id}')


def emit_notifications_to_users(notifications: list, lang: str = 'de'):
    """
    Emit notifications to multiple users.
    
    Args:
        notifications: List of Notification objects
        lang: Language for localized content
    """
    for notification in notifications:
        emit_notification(notification.user_id, notification, lang)


# ============================================================================
# DECORATORS
# ============================================================================

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin():
            flash('Keine Berechtigung für diese Aktion.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def log_action(action, entity_type=None, entity_id=None, entity_name=None, old_value=None, new_value=None):
    """Log an action to the audit log"""
    log = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_name=entity_name,
        old_value=old_value,
        new_value=new_value,
        ip_address=request.remote_addr,
        user_agent=str(request.user_agent)[:500] if request.user_agent else None
    )
    db.session.add(log)
    db.session.commit()


# ============================================================================
# AUTH ROUTES
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = request.form.get('remember', False)
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            if not user.is_active:
                flash('Ihr Konto ist deaktiviert.', 'danger')
                return render_template('login.html')
            
            login_user(user, remember=remember)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            log_action('LOGIN', 'User', user.id, user.email)
            
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        
        flash('Ungültige Anmeldedaten.', 'danger')
    
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Logout user"""
    log_action('LOGOUT', 'User', current_user.id, current_user.email)
    logout_user()
    flash('Sie wurden abgemeldet.', 'success')
    return redirect(url_for('index'))


# ============================================================================
# MAIN ROUTES
# ============================================================================

@app.route('/')
def index():
    """Home page - redirect to dashboard if logged in"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')


@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard with task overview"""
    from datetime import date, timedelta
    
    # Get user's tasks based on role and entity permissions (exclude archived)
    if current_user.is_admin() or current_user.is_manager():
        # Admins and managers see all tasks
        base_query = Task.query.filter((Task.is_archived == False) | (Task.is_archived == None))
    else:
        # Get accessible entity IDs for this user
        accessible_entity_ids = current_user.get_accessible_entity_ids('view')
        
        # Others see their tasks + tasks for accessible entities
        if accessible_entity_ids:
            base_query = Task.query.filter(
                ((Task.is_archived == False) | (Task.is_archived == None)),
                (Task.owner_id == current_user.id) | 
                (Task.reviewer_id == current_user.id) |
                (Task.entity_id.in_(accessible_entity_ids))
            )
        else:
            base_query = Task.query.filter(
                ((Task.is_archived == False) | (Task.is_archived == None)),
                (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
            )
    
    today = date.today()
    soon = today + timedelta(days=7)
    
    stats = {
        'total': base_query.count(),
        'overdue': base_query.filter(Task.due_date < today, Task.status != 'completed').count(),
        'due_soon': base_query.filter(Task.due_date >= today, Task.due_date <= soon, Task.status != 'completed').count(),
        'in_review': base_query.filter_by(status='in_review').count(),
        'completed': base_query.filter_by(status='completed').count(),
    }
    
    # My upcoming tasks
    my_tasks = base_query.filter(
        Task.status != 'completed'
    ).order_by(Task.due_date).limit(10).all()
    
    # Overdue tasks
    overdue_tasks = base_query.filter(
        Task.due_date < today,
        Task.status != 'completed'
    ).order_by(Task.due_date).limit(5).all()
    
    return render_template('dashboard.html', 
                         stats=stats, 
                         my_tasks=my_tasks, 
                         overdue_tasks=overdue_tasks,
                         today=today)


@app.route('/tasks')
@login_required
def task_list():
    """Task list with filters"""
    from datetime import date
    
    # Get filter parameters
    status_filter = request.args.get('status', '')
    entity_filter = request.args.get('entity', type=int)
    tax_type_filter = request.args.get('tax_type', type=int)
    year_filter = request.args.get('year', type=int, default=date.today().year)
    show_archived = request.args.get('show_archived', 'false') == 'true'
    
    # Base query - exclude archived tasks by default
    query = Task.query
    if not show_archived:
        query = query.filter((Task.is_archived == False) | (Task.is_archived == None))
    
    # Apply role-based and entity scoping filtering
    if not (current_user.is_admin() or current_user.is_manager()):
        # Get accessible entity IDs for this user
        accessible_entity_ids = current_user.get_accessible_entity_ids('view')
        
        # Filter by owned/reviewed tasks OR tasks for accessible entities
        if accessible_entity_ids:
            query = query.filter(
                (Task.owner_id == current_user.id) | 
                (Task.reviewer_id == current_user.id) |
                (Task.entity_id.in_(accessible_entity_ids))
            )
        else:
            query = query.filter(
                (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
            )
    
    # Apply filters
    if status_filter:
        if status_filter == 'overdue':
            query = query.filter(Task.due_date < date.today(), Task.status != 'completed')
        elif status_filter == 'due_soon':
            from datetime import timedelta
            soon = date.today() + timedelta(days=7)
            query = query.filter(Task.due_date >= date.today(), Task.due_date <= soon, Task.status != 'completed')
        else:
            query = query.filter_by(status=status_filter)
    
    if entity_filter:
        query = query.filter_by(entity_id=entity_filter)
    
    if tax_type_filter:
        query = query.join(TaskTemplate).filter(TaskTemplate.tax_type_id == tax_type_filter)
    
    if year_filter:
        query = query.filter_by(year=year_filter)
    
    # Order by due date
    tasks = query.order_by(Task.due_date).all()
    
    # Get filter options - show only accessible entities for non-admins
    if current_user.is_admin() or current_user.is_manager():
        entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    else:
        entities = current_user.get_accessible_entities('view')
    
    tax_types = TaxType.query.filter_by(is_active=True).order_by(TaxType.code).all()
    years = db.session.query(Task.year).distinct().order_by(Task.year.desc()).all()
    years = [y[0] for y in years]
    
    # Get users for bulk assign modal
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    
    return render_template('tasks/list.html', 
                         tasks=tasks, 
                         entities=entities,
                         tax_types=tax_types,
                         years=years,
                         users=users,
                         current_filters={
                             'status': status_filter,
                             'entity': entity_filter,
                             'tax_type': tax_type_filter,
                             'year': year_filter
                         })


@app.route('/tasks/<int:task_id>')
@login_required
def task_detail(task_id):
    """Task detail view"""
    task = Task.query.get_or_404(task_id)
    
    # Check access (admin, manager, owner, or any assigned reviewer)
    if not (current_user.is_admin() or current_user.is_manager() or 
            task.owner_id == current_user.id or task.is_reviewer(current_user)):
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('task_list'))
    
    # Get audit log for this task
    audit_logs = AuditLog.query.filter_by(
        entity_type='Task', 
        entity_id=task.id
    ).order_by(AuditLog.timestamp.desc()).all()
    
    return render_template('tasks/detail.html', task=task, audit_logs=audit_logs)


@app.route('/tasks/new', methods=['GET', 'POST'])
@login_required
def task_create():
    """Create a new task"""
    from datetime import date
    
    if request.method == 'POST':
        # Get form data
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        entity_id = request.form.get('entity_id', type=int)
        template_id = request.form.get('template_id', type=int) or None
        due_date_str = request.form.get('due_date', '')
        year = request.form.get('year', type=int, default=date.today().year)
        period = request.form.get('period', '').strip()
        owner_id = request.form.get('owner_id', type=int) or None
        owner_team_id = request.form.get('owner_team_id', type=int) or None
        reviewer_ids = request.form.getlist('reviewer_ids', type=int)
        reviewer_team_id = request.form.get('reviewer_team_id', type=int) or None
        
        # Validate required fields
        errors = []
        if not title:
            errors.append('Titel ist erforderlich.')
        if not entity_id:
            errors.append('Gesellschaft ist erforderlich.')
        if not due_date_str:
            errors.append('Fälligkeitsdatum ist erforderlich.')
        
        if errors:
            for error in errors:
                flash(error, 'danger')
        else:
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                
                task = Task(
                    title=title,
                    description=description,
                    entity_id=entity_id,
                    template_id=template_id,
                    due_date=due_date,
                    year=year,
                    period=period,
                    owner_id=owner_id,
                    owner_team_id=owner_team_id,
                    reviewer_team_id=reviewer_team_id,
                    status='draft'
                )
                db.session.add(task)
                db.session.flush()  # Get task ID for reviewers
                
                # Add multiple reviewers
                if reviewer_ids:
                    task.set_reviewers(reviewer_ids)
                
                db.session.commit()
                
                log_action('CREATE', 'Task', task.id, task.title)
                
                # Send notifications
                lang = session.get('lang', 'de')
                notifications = []
                
                # Notify owner if assigned
                if owner_id and owner_id != current_user.id:
                    notifications.append(
                        NotificationService.notify_task_assigned(task, owner_id, current_user.id)
                    )
                
                # Notify reviewers
                for reviewer_id in reviewer_ids:
                    if reviewer_id != current_user.id:
                        notifications.append(
                            NotificationService.notify_reviewer_added(task, reviewer_id, current_user.id)
                        )
                
                if notifications:
                    db.session.commit()
                    emit_notifications_to_users(notifications, lang)
                
                # Send email notifications (async-safe, won't block if email disabled)
                if owner_id and owner_id != current_user.id:
                    owner = User.query.get(owner_id)
                    if owner:
                        email_service.send_task_assigned(task, owner, current_user, lang)
                
                for reviewer_id in reviewer_ids:
                    if reviewer_id != current_user.id:
                        reviewer = User.query.get(reviewer_id)
                        if reviewer:
                            email_service.send_task_assigned(task, reviewer, current_user, lang)
                
                flash('Aufgabe erfolgreich erstellt.', 'success')
                return redirect(url_for('task_detail', task_id=task.id))
                
            except ValueError:
                flash('Ungültiges Datumsformat.', 'danger')
    
    # GET request - show form
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    templates = TaskTemplate.query.filter_by(is_active=True).order_by(TaskTemplate.keyword).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    teams = Team.query.filter_by(is_active=True).order_by(Team.name).all()
    presets = TaskPreset.query.filter_by(is_active=True).order_by(TaskPreset.category, TaskPreset.title).all()
    
    return render_template('tasks/form.html',
                         task=None,
                         entities=entities,
                         templates=templates,
                         users=users,
                         teams=teams,
                         presets=presets,
                         current_year=date.today().year)


@app.route('/tasks/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required
def task_edit(task_id):
    """Edit an existing task"""
    from datetime import date
    
    task = Task.query.get_or_404(task_id)
    
    # Check permission
    if not (current_user.is_admin() or current_user.is_manager() or task.owner_id == current_user.id):
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if request.method == 'POST':
        old_values = f"title={task.title}, due_date={task.due_date}"
        
        task.title = request.form.get('title', '').strip()
        task.description = request.form.get('description', '').strip()
        task.entity_id = request.form.get('entity_id', type=int)
        task.template_id = request.form.get('template_id', type=int) or None
        task.year = request.form.get('year', type=int, default=date.today().year)
        task.period = request.form.get('period', '').strip()
        task.owner_id = request.form.get('owner_id', type=int) or None
        task.owner_team_id = request.form.get('owner_team_id', type=int) or None
        task.reviewer_team_id = request.form.get('reviewer_team_id', type=int) or None
        
        # Handle multiple reviewers
        old_reviewer_ids = set(tr.user_id for tr in task.reviewers)
        reviewer_ids = request.form.getlist('reviewer_ids', type=int)
        new_reviewer_ids = set(reviewer_ids)
        task.set_reviewers(reviewer_ids)
        
        due_date_str = request.form.get('due_date', '')
        if due_date_str:
            try:
                task.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Ungültiges Datumsformat.', 'danger')
                return redirect(url_for('task_edit', task_id=task_id))
        
        db.session.commit()
        
        new_values = f"title={task.title}, due_date={task.due_date}"
        log_action('UPDATE', 'Task', task.id, task.title, old_values, new_values)
        
        # Notify newly added reviewers
        lang = session.get('lang', 'de')
        notifications = []
        added_reviewer_ids = new_reviewer_ids - old_reviewer_ids
        for reviewer_id in added_reviewer_ids:
            if reviewer_id != current_user.id:
                notifications.append(
                    NotificationService.notify_reviewer_added(task, reviewer_id, current_user.id)
                )
        
        if notifications:
            db.session.commit()
            emit_notifications_to_users(notifications, lang)
        
        flash('Aufgabe erfolgreich aktualisiert.', 'success')
        return redirect(url_for('task_detail', task_id=task_id))
    
    # GET request
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    templates = TaskTemplate.query.filter_by(is_active=True).order_by(TaskTemplate.keyword).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    teams = Team.query.filter_by(is_active=True).order_by(Team.name).all()
    
    return render_template('tasks/form.html',
                         task=task,
                         entities=entities,
                         templates=templates,
                         users=users,
                         teams=teams,
                         current_year=date.today().year)


@app.route('/tasks/<int:task_id>/status', methods=['POST'])
@login_required
def task_change_status(task_id):
    """Change task status with multi-stage approval workflow"""
    task = Task.query.get_or_404(task_id)
    new_status = request.form.get('status')
    note = request.form.get('note', '').strip()
    
    # Check if transition is allowed
    if not task.can_transition_to(new_status, current_user):
        flash('Diese Statusänderung ist nicht erlaubt.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    try:
        old_status = task.transition_to(new_status, current_user, note)
        db.session.commit()
        log_action('STATUS_CHANGE', 'Task', task.id, task.title, old_status, new_status)
        
        # Send notifications for status change
        lang = session.get('lang', 'de')
        notifications = []
        
        # Notify owner about status changes (if not the one who made the change)
        if task.owner_id and task.owner_id != current_user.id:
            notifications.append(
                NotificationService.notify_status_changed(task, task.owner_id, old_status, new_status, current_user.id)
            )
        
        # Notify reviewers when task is submitted for review
        if new_status == 'submitted':
            for tr in task.reviewers:
                if tr.user_id != current_user.id:
                    notifications.append(
                        NotificationService.create(
                            user_id=tr.user_id,
                            notification_type='review_requested',
                            title_de=f'Review angefordert: {task.title}',
                            title_en=f'Review requested: {task.title}',
                            message_de='Die Aufgabe wurde zur Prüfung eingereicht.',
                            message_en='The task has been submitted for review.',
                            entity_type='task',
                            entity_id=task.id,
                            actor_id=current_user.id
                        )
                    )
        
        if notifications:
            db.session.commit()
            emit_notifications_to_users(notifications, lang)
        
        # Send email notifications for status change
        if task.owner_id and task.owner_id != current_user.id:
            owner = User.query.get(task.owner_id)
            if owner:
                email_service.send_status_changed(task, owner, old_status, new_status, lang)
        
        # Email reviewers when task is submitted for review
        if new_status == 'submitted':
            for tr in task.reviewers:
                if tr.user_id != current_user.id:
                    email_service.send_status_changed(task, tr.user, old_status, new_status, lang)
        
        # Status-specific messages
        status_messages = {
            'submitted': 'Aufgabe eingereicht. Wartet auf Prüfung.',
            'in_review': 'Aufgabe wird geprüft.',
            'approved': 'Aufgabe genehmigt. Kann abgeschlossen werden.',
            'completed': 'Aufgabe erfolgreich abgeschlossen.',
            'rejected': 'Aufgabe zur Überarbeitung zurückgewiesen.',
            'draft': 'Aufgabe zurück in Bearbeitung.',
        }
        flash(status_messages.get(new_status, f'Status geändert: {old_status} → {new_status}'), 'success')
        
    except ValueError as e:
        flash(str(e), 'danger')
    
    return redirect(url_for('task_detail', task_id=task_id))


@app.route('/tasks/<int:task_id>/reviewer-action', methods=['POST'])
@login_required
def task_reviewer_action(task_id):
    """Handle individual reviewer approval/rejection using ApprovalService"""
    task = Task.query.get_or_404(task_id)
    action = request.form.get('action')
    note = request.form.get('note', '').strip()
    
    if action == 'approve':
        result, message = ApprovalService.approve(task, current_user, note)
        db.session.commit()
        
        lang = session.get('lang', 'de')
        
        if result == ApprovalResult.ALL_APPROVED:
            log_action('REVIEWER_APPROVE', 'Task', task.id, task.title, 
                       f'Reviewer: {current_user.name}', 'Final approval')
            log_action('STATUS_CHANGE', 'Task', task.id, task.title, 'in_review', 'approved')
            
            # Notify owner that task is fully approved
            if task.owner_id and task.owner_id != current_user.id:
                notification = NotificationService.notify_task_approved(task, task.owner_id, current_user.id, note)
                db.session.commit()
                emit_notification(task.owner_id, notification, lang)
            
            flash('Alle Prüfer haben genehmigt. Aufgabe ist nun genehmigt.', 'success')
        elif result == ApprovalResult.SUCCESS:
            log_action('REVIEWER_APPROVE', 'Task', task.id, task.title, 
                       f'Reviewer: {current_user.name}', note or 'No note')
            flash(message, 'success')
        else:
            flash(message, 'danger')
    
    elif action == 'reject':
        result, message = ApprovalService.reject(task, current_user, note)
        db.session.commit()
        
        lang = session.get('lang', 'de')
        
        if result == ApprovalResult.TASK_REJECTED:
            log_action('REVIEWER_REJECT', 'Task', task.id, task.title, 
                       f'Reviewer: {current_user.name}', note or 'No note')
            log_action('STATUS_CHANGE', 'Task', task.id, task.title, 'in_review', 'rejected')
            
            # Notify owner that task was rejected
            if task.owner_id and task.owner_id != current_user.id:
                notification = NotificationService.notify_task_rejected(task, task.owner_id, current_user.id, note)
                db.session.commit()
                emit_notification(task.owner_id, notification, lang)
            
            flash('Aufgabe wurde von Ihnen abgelehnt und zur Überarbeitung zurückgewiesen.', 'warning')
        else:
            flash(message, 'danger')
    
    return redirect(url_for('task_detail', task_id=task_id))


# ============================================================================
# TASK ARCHIVE & RESTORE
# ============================================================================

@app.route('/tasks/<int:task_id>/archive', methods=['POST'])
@login_required
def task_archive(task_id):
    """Archive a task (soft-delete)"""
    task = Task.query.get_or_404(task_id)
    
    # Only admin, manager, or owner can archive
    if not (current_user.is_admin() or current_user.is_manager() or current_user.id == task.owner_id):
        flash('Keine Berechtigung zum Archivieren.' if session.get('lang', 'de') == 'de' else 'No permission to archive.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if task.is_archived:
        flash('Aufgabe ist bereits archiviert.' if session.get('lang', 'de') == 'de' else 'Task is already archived.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    reason = request.form.get('reason', '')
    task.archive(current_user, reason)
    db.session.commit()
    
    log_action('ARCHIVE', 'Task', task.id, task.title, 'active', 'archived')
    
    flash('Aufgabe wurde archiviert.' if session.get('lang', 'de') == 'de' else 'Task has been archived.', 'success')
    return redirect(url_for('task_list'))


@app.route('/tasks/<int:task_id>/restore', methods=['POST'])
@login_required
def task_restore(task_id):
    """Restore a task from archive"""
    task = Task.query.get_or_404(task_id)
    
    # Only admin or manager can restore
    if not (current_user.is_admin() or current_user.is_manager()):
        flash('Keine Berechtigung zum Wiederherstellen.' if session.get('lang', 'de') == 'de' else 'No permission to restore.', 'danger')
        return redirect(url_for('task_archive_list'))
    
    if not task.is_archived:
        flash('Aufgabe ist nicht archiviert.' if session.get('lang', 'de') == 'de' else 'Task is not archived.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    task.restore()
    db.session.commit()
    
    log_action('RESTORE', 'Task', task.id, task.title, 'archived', 'active')
    
    flash('Aufgabe wurde wiederhergestellt.' if session.get('lang', 'de') == 'de' else 'Task has been restored.', 'success')
    return redirect(url_for('task_detail', task_id=task_id))


@app.route('/tasks/archive')
@login_required
def task_archive_list():
    """View archived tasks"""
    lang = session.get('lang', 'de')
    
    # Build query for archived tasks
    query = Task.query.filter_by(is_archived=True)
    
    # Non-admin/manager users only see their own archived tasks
    if not (current_user.is_admin() or current_user.is_manager()):
        accessible_entity_ids = current_user.get_accessible_entity_ids()
        query = query.filter(Task.entity_id.in_(accessible_entity_ids))
    
    # Filters
    entity_id = request.args.get('entity_id', type=int)
    year = request.args.get('year', type=int)
    status = request.args.get('status')
    
    if entity_id:
        query = query.filter_by(entity_id=entity_id)
    if year:
        query = query.filter_by(year=year)
    if status:
        query = query.filter_by(status=status)
    
    # Order by archived date, most recent first
    query = query.order_by(Task.archived_at.desc())
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    tasks = pagination.items
    
    # Get filter options
    entities = Entity.query.order_by(Entity.name).all()
    years = db.session.query(Task.year).filter_by(is_archived=True).distinct().order_by(Task.year.desc()).all()
    years = [y[0] for y in years]
    
    return render_template('tasks/archive.html',
                           tasks=tasks,
                           pagination=pagination,
                           entities=entities,
                           years=years,
                           current_filters={
                               'entity_id': entity_id,
                               'year': year,
                               'status': status
                           },
                           t=lambda key: TRANSLATIONS.get(key, {}).get(lang, key),
                           lang=lang)


@app.route('/api/tasks/bulk-archive', methods=['POST'])
@login_required
def api_bulk_archive():
    """Bulk archive multiple tasks"""
    if not (current_user.is_admin() or current_user.is_manager()):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    reason = data.get('reason', '')
    
    if not task_ids:
        return jsonify({'success': False, 'error': 'No tasks selected'}), 400
    
    archived_count = 0
    for task_id in task_ids:
        task = Task.query.get(task_id)
        if task and not task.is_archived:
            task.archive(current_user, reason)
            log_action('ARCHIVE', 'Task', task.id, task.title, 'active', 'archived')
            archived_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'archived_count': archived_count,
        'message': f'{archived_count} Aufgabe(n) archiviert.' if session.get('lang', 'de') == 'de' else f'{archived_count} task(s) archived.'
    })


@app.route('/api/tasks/bulk-restore', methods=['POST'])
@login_required
def api_bulk_restore():
    """Bulk restore multiple tasks from archive"""
    if not (current_user.is_admin() or current_user.is_manager()):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    
    if not task_ids:
        return jsonify({'success': False, 'error': 'No tasks selected'}), 400
    
    restored_count = 0
    for task_id in task_ids:
        task = Task.query.get(task_id)
        if task and task.is_archived:
            task.restore()
            log_action('RESTORE', 'Task', task.id, task.title, 'archived', 'active')
            restored_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'restored_count': restored_count,
        'message': f'{restored_count} Aufgabe(n) wiederhergestellt.' if session.get('lang', 'de') == 'de' else f'{restored_count} task(s) restored.'
    })


@app.route('/tasks/<int:task_id>/delete', methods=['POST'])
@login_required
def task_permanent_delete(task_id):
    """Permanently delete an archived task (admin only)"""
    if not current_user.is_admin():
        flash('Nur Administratoren können Aufgaben endgültig löschen.' if session.get('lang', 'de') == 'de' else 'Only administrators can permanently delete tasks.', 'danger')
        return redirect(url_for('task_archive_list'))
    
    task = Task.query.get_or_404(task_id)
    
    if not task.is_archived:
        flash('Nur archivierte Aufgaben können endgültig gelöscht werden.' if session.get('lang', 'de') == 'de' else 'Only archived tasks can be permanently deleted.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    task_title = task.title
    task_id_log = task.id
    
    # Delete related records first
    TaskEvidence.query.filter_by(task_id=task_id).delete()
    Comment.query.filter_by(task_id=task_id).delete()
    TaskReviewer.query.filter_by(task_id=task_id).delete()
    Notification.query.filter(Notification.entity_type == 'task', Notification.entity_id == task_id).delete()
    
    db.session.delete(task)
    db.session.commit()
    
    log_action('DELETE', 'Task', task_id_log, task_title, 'archived', 'deleted')
    
    flash(f'Aufgabe "{task_title}" wurde endgültig gelöscht.' if session.get('lang', 'de') == 'de' else f'Task "{task_title}" has been permanently deleted.', 'success')
    return redirect(url_for('task_archive_list'))


@app.route('/api/tasks/archive/bulk-delete', methods=['POST'])
@login_required
def api_bulk_permanent_delete():
    """Bulk permanently delete multiple archived tasks (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'Permission denied - admin only'}), 403
    
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    
    if not task_ids:
        return jsonify({'success': False, 'error': 'No tasks selected'}), 400
    
    deleted_count = 0
    for task_id in task_ids:
        task = Task.query.get(task_id)
        if task and task.is_archived:
            task_title = task.title
            
            # Delete related records first
            TaskEvidence.query.filter_by(task_id=task_id).delete()
            Comment.query.filter_by(task_id=task_id).delete()
            TaskReviewer.query.filter_by(task_id=task_id).delete()
            Notification.query.filter(Notification.entity_type == 'task', Notification.entity_id == task_id).delete()
            
            db.session.delete(task)
            log_action('DELETE', 'Task', task_id, task_title, 'archived', 'deleted')
            deleted_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'deleted_count': deleted_count,
        'message': f'{deleted_count} Aufgabe(n) endgültig gelöscht.' if session.get('lang', 'de') == 'de' else f'{deleted_count} task(s) permanently deleted.'
    })


# ============================================================================
# TASK EVIDENCE (File Upload & Links)
# ============================================================================

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config.get('ALLOWED_EXTENSIONS', set())


@app.route('/tasks/<int:task_id>/evidence/upload', methods=['POST'])
@login_required
def task_upload_evidence(task_id):
    """Upload file evidence to a task"""
    import os
    import uuid
    from werkzeug.utils import secure_filename
    
    task = Task.query.get_or_404(task_id)
    
    # Check permission (admin, manager, owner, or any assigned reviewer)
    if not (current_user.is_admin() or current_user.is_manager() or 
            current_user.id == task.owner_id or task.is_reviewer(current_user)):
        flash('Keine Berechtigung zum Hochladen.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if 'file' not in request.files:
        flash('Keine Datei ausgewählt.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    file = request.files['file']
    if file.filename == '':
        flash('Keine Datei ausgewählt.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if file and allowed_file(file.filename):
        # Secure the filename and make it unique
        original_filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{original_filename}"
        
        # Create task-specific folder
        upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], f'task_{task_id}')
        os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        # Create evidence record
        evidence = TaskEvidence(
            task_id=task_id,
            evidence_type='file',
            filename=original_filename,
            file_path=file_path,
            file_size=file_size,
            mime_type=file.content_type,
            uploaded_by_id=current_user.id
        )
        db.session.add(evidence)
        db.session.commit()
        
        log_action('UPLOAD', 'Task', task_id, task.title, None, f'Datei: {original_filename}')
        flash(f'Datei "{original_filename}" erfolgreich hochgeladen.', 'success')
    else:
        allowed = ', '.join(app.config.get('ALLOWED_EXTENSIONS', []))
        flash(f'Dateityp nicht erlaubt. Erlaubte Formate: {allowed}', 'danger')
    
    return redirect(url_for('task_detail', task_id=task_id) + '#evidence')


@app.route('/tasks/<int:task_id>/evidence/link', methods=['POST'])
@login_required
def task_add_link(task_id):
    """Add link evidence to a task"""
    task = Task.query.get_or_404(task_id)
    
    # Check permission (admin, manager, owner, or any assigned reviewer)
    if not (current_user.is_admin() or current_user.is_manager() or 
            current_user.id == task.owner_id or task.is_reviewer(current_user)):
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    url = request.form.get('url', '').strip()
    link_title = request.form.get('link_title', '').strip()
    
    if not url:
        flash('URL ist erforderlich.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    # Basic URL validation
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    evidence = TaskEvidence(
        task_id=task_id,
        evidence_type='link',
        url=url,
        link_title=link_title or url[:100],
        uploaded_by_id=current_user.id
    )
    db.session.add(evidence)
    db.session.commit()
    
    log_action('LINK_ADD', 'Task', task_id, task.title, None, f'Link: {link_title or url[:50]}')
    flash('Link erfolgreich hinzugefügt.', 'success')
    
    return redirect(url_for('task_detail', task_id=task_id) + '#evidence')


@app.route('/tasks/<int:task_id>/evidence/<int:evidence_id>/download')
@login_required
def task_download_evidence(task_id, evidence_id):
    """Download file evidence"""
    import os
    
    evidence = TaskEvidence.query.filter_by(id=evidence_id, task_id=task_id).first_or_404()
    
    if evidence.evidence_type != 'file':
        flash('Nur Dateien können heruntergeladen werden.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if not os.path.exists(evidence.file_path):
        flash('Datei nicht gefunden.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    return send_file(evidence.file_path, as_attachment=True, download_name=evidence.filename)


@app.route('/tasks/<int:task_id>/evidence/<int:evidence_id>/preview')
@login_required
def task_preview_evidence(task_id, evidence_id):
    """Preview file evidence inline (for images, PDFs, text files)"""
    import os
    
    evidence = TaskEvidence.query.filter_by(id=evidence_id, task_id=task_id).first_or_404()
    
    if evidence.evidence_type != 'file':
        flash('Nur Dateien können angezeigt werden.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if not os.path.exists(evidence.file_path):
        flash('Datei nicht gefunden.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    # Determine mimetype for inline display
    mimetype = evidence.mime_type or 'application/octet-stream'
    
    return send_file(evidence.file_path, mimetype=mimetype, as_attachment=False, download_name=evidence.filename)


@app.route('/tasks/<int:task_id>/evidence/<int:evidence_id>/delete', methods=['POST'])
@login_required
def task_delete_evidence(task_id, evidence_id):
    """Delete evidence from a task"""
    import os
    
    evidence = TaskEvidence.query.filter_by(id=evidence_id, task_id=task_id).first_or_404()
    task = Task.query.get_or_404(task_id)
    
    # Check permission - only uploader, owner, or admin can delete
    if not (current_user.is_admin() or current_user.id == evidence.uploaded_by_id or 
            current_user.id == task.owner_id):
        flash('Keine Berechtigung zum Löschen.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    # Delete file if exists
    if evidence.evidence_type == 'file' and evidence.file_path:
        if os.path.exists(evidence.file_path):
            os.remove(evidence.file_path)
    
    description = evidence.filename or evidence.url
    db.session.delete(evidence)
    db.session.commit()
    
    log_action('EVIDENCE_DELETE', 'Task', task_id, task.title, f'Nachweis: {description[:50]}', None)
    flash('Nachweis gelöscht.', 'success')
    
    return redirect(url_for('task_detail', task_id=task_id) + '#evidence')


# ============================================================================
# TASK COMMENTS
# ============================================================================

@app.route('/tasks/<int:task_id>/comments', methods=['POST'])
@login_required
def task_add_comment(task_id):
    """Add a comment to a task"""
    task = Task.query.get_or_404(task_id)
    
    text = request.form.get('text', '').strip()
    
    if not text:
        flash('Kommentar darf nicht leer sein.', 'warning')
        return redirect(url_for('task_detail', task_id=task_id) + '#comments')
    
    comment = Comment(
        task_id=task_id,
        text=text,
        created_by_id=current_user.id
    )
    db.session.add(comment)
    db.session.commit()
    
    log_action('COMMENT', 'Task', task_id, task.title, None, text[:100] + ('...' if len(text) > 100 else ''))
    
    # Notify relevant users about the comment
    lang = session.get('lang', 'de')
    notifications = []
    notified_users = set()
    
    # Notify task owner
    if task.owner_id and task.owner_id != current_user.id:
        notifications.append(
            NotificationService.notify_comment_added(task, comment, task.owner_id, current_user.id)
        )
        notified_users.add(task.owner_id)
    
    # Notify reviewers
    for tr in task.reviewers:
        if tr.user_id != current_user.id and tr.user_id not in notified_users:
            notifications.append(
                NotificationService.notify_comment_added(task, comment, tr.user_id, current_user.id)
            )
            notified_users.add(tr.user_id)
    
    if notifications:
        db.session.commit()
        emit_notifications_to_users(notifications, lang)
    
    flash('Kommentar hinzugefügt.', 'success')
    
    return redirect(url_for('task_detail', task_id=task_id) + '#comments')


@app.route('/tasks/<int:task_id>/comments/<int:comment_id>/delete', methods=['POST'])
@login_required
def task_delete_comment(task_id, comment_id):
    """Delete a comment from a task"""
    comment = Comment.query.filter_by(id=comment_id, task_id=task_id).first_or_404()
    task = Task.query.get_or_404(task_id)
    
    # Check permission - only comment author, task owner, or admin can delete
    if not (current_user.is_admin() or current_user.id == comment.created_by_id or 
            current_user.id == task.owner_id):
        flash('Keine Berechtigung zum Löschen.', 'danger')
        return redirect(url_for('task_detail', task_id=task_id))
    
    # Save text before deletion for audit log
    comment_text = comment.text[:50] + ('...' if len(comment.text) > 50 else '')
    
    db.session.delete(comment)
    db.session.commit()
    
    log_action('COMMENT_DELETE', 'Task', task_id, task.title, comment_text, None)
    flash('Kommentar gelöscht.', 'success')
    
    return redirect(url_for('task_detail', task_id=task_id) + '#comments')


@app.route('/set-language/<lang>')
def set_language(lang):
    """Change language"""
    if lang in app.config.get('SUPPORTED_LANGUAGES', ['de', 'en']):
        session['lang'] = lang
    return redirect(request.referrer or url_for('index'))


@app.route('/calendar')
@login_required
def calendar_view():
    """Calendar month view of tasks"""
    from datetime import date, timedelta
    import calendar
    
    # Get year and month from query params, default to current
    year = request.args.get('year', type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    
    # Validate month/year
    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1
    
    # Get first and last day of month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Build calendar structure
    cal = calendar.Calendar(firstweekday=0)  # Monday first
    month_days = cal.monthdayscalendar(year, month)
    
    # Get tasks for this month (exclude archived)
    query = Task.query.filter(
        Task.due_date >= first_day,
        Task.due_date <= last_day,
        (Task.is_archived == False) | (Task.is_archived == None)
    )
    
    # Apply role-based filtering
    if not (current_user.is_admin() or current_user.is_manager()):
        query = query.filter(
            (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
        )
    
    tasks = query.order_by(Task.due_date).all()
    
    # Group tasks by day
    tasks_by_day = {}
    for task in tasks:
        day = task.due_date.day
        if day not in tasks_by_day:
            tasks_by_day[day] = []
        tasks_by_day[day].append(task)
    
    # Calculate prev/next month
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    return render_template('calendar.html',
                         year=year,
                         month=month,
                         month_name=calendar.month_name[month],
                         month_days=month_days,
                         tasks_by_day=tasks_by_day,
                         today=date.today(),
                         prev_month=prev_month,
                         prev_year=prev_year,
                         next_month=next_month,
                         next_year=next_year)


@app.route('/calendar/year')
@login_required
def calendar_year_view():
    """Calendar year view of tasks"""
    from datetime import date, timedelta
    import calendar
    
    year = request.args.get('year', type=int, default=date.today().year)
    
    # Get all tasks for the year (exclude archived)
    first_day = date(year, 1, 1)
    last_day = date(year, 12, 31)
    
    query = Task.query.filter(
        Task.due_date >= first_day,
        Task.due_date <= last_day,
        (Task.is_archived == False) | (Task.is_archived == None)
    )
    
    if not (current_user.is_admin() or current_user.is_manager()):
        query = query.filter(
            (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
        )
    
    tasks = query.order_by(Task.due_date).all()
    
    # Group tasks by month
    tasks_by_month = {m: [] for m in range(1, 13)}
    for task in tasks:
        tasks_by_month[task.due_date.month].append(task)
    
    # Count by status per month
    month_stats = {}
    for m in range(1, 13):
        month_tasks = tasks_by_month[m]
        month_stats[m] = {
            'total': len(month_tasks),
            'completed': sum(1 for t in month_tasks if t.status == 'completed'),
            'overdue': sum(1 for t in month_tasks if t.is_overdue),
            'in_review': sum(1 for t in month_tasks if t.status == 'in_review'),
        }
    
    # German month names
    month_names = ['', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 
                   'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']
    
    return render_template('calendar_year.html',
                         year=year,
                         month_names=month_names,
                         tasks_by_month=tasks_by_month,
                         month_stats=month_stats,
                         today=date.today())


@app.route('/calendar/week')
@login_required
def calendar_week_view():
    """Calendar week view of tasks"""
    from datetime import date, timedelta
    import calendar
    
    today = date.today()
    year = request.args.get('year', type=int, default=today.year)
    week = request.args.get('week', type=int, default=today.isocalendar()[1])
    
    # Validate week number
    if week < 1:
        week = 52
        year -= 1
    elif week > 52:
        week = 1
        year += 1
    
    # Get first day of the week (Monday)
    # ISO week: week 1 is the week containing Jan 4th
    jan4 = date(year, 1, 4)
    week_start = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=week - 1)
    week_end = week_start + timedelta(days=6)
    
    # Get tasks for this week (exclude archived)
    query = Task.query.filter(
        Task.due_date >= week_start,
        Task.due_date <= week_end,
        (Task.is_archived == False) | (Task.is_archived == None)
    )
    
    if not (current_user.is_admin() or current_user.is_manager()):
        query = query.filter(
            (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
        )
    
    tasks = query.order_by(Task.due_date, Task.id).all()
    
    # Build days of the week
    week_days = []
    for i in range(7):
        day_date = week_start + timedelta(days=i)
        day_tasks = [t for t in tasks if t.due_date == day_date]
        week_days.append({
            'date': day_date,
            'weekday': ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag'][i],
            'weekday_short': ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So'][i],
            'tasks': day_tasks,
            'is_today': day_date == today,
            'is_weekend': i >= 5
        })
    
    # Prev/next week
    prev_week = week - 1 if week > 1 else 52
    prev_year = year if week > 1 else year - 1
    next_week = week + 1 if week < 52 else 1
    next_year = year if week < 52 else year + 1
    
    return render_template('calendar_week.html',
                         year=year,
                         week=week,
                         week_start=week_start,
                         week_end=week_end,
                         week_days=week_days,
                         today=today,
                         prev_week=prev_week,
                         prev_year=prev_year,
                         next_week=next_week,
                         next_year=next_year)


# ============================================================================
# ADMIN ROUTES
# ============================================================================

@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard"""
    from datetime import date
    stats = {
        'users': User.query.count(),
        'active_users': User.query.filter_by(is_active=True).count(),
        'entities': Entity.query.filter_by(is_active=True).count(),
        'tax_types': TaxType.query.filter_by(is_active=True).count(),
        'tasks_total': Task.query.count(),
        'tasks_overdue': Task.query.filter(Task.due_date < date.today(), Task.status != 'completed').count(),
        'tasks_completed': Task.query.filter_by(status='completed').count(),
        'presets': TaskPreset.query.filter_by(is_active=True).count(),
        'modules': Module.query.filter_by(is_active=True).count(),
    }
    return render_template('admin/dashboard.html', stats=stats)


# ============================================================================
# ADMIN: MODULE MANAGEMENT
# ============================================================================

@app.route('/admin/modules')
@admin_required
def admin_modules():
    """Module management"""
    modules = Module.query.order_by(Module.nav_order, Module.code).all()
    lang = session.get('lang', 'de')
    return render_template('admin/modules.html', modules=modules, lang=lang)


@app.route('/admin/modules/<int:module_id>/toggle', methods=['POST'])
@admin_required
def admin_module_toggle(module_id):
    """Toggle module active status"""
    module = Module.query.get_or_404(module_id)
    lang = session.get('lang', 'de')
    
    if module.is_core:
        flash('Kernmodule können nicht deaktiviert werden.' if lang == 'de' else 'Core modules cannot be disabled.', 'warning')
        return redirect(url_for('admin_modules'))
    
    module.is_active = not module.is_active
    db.session.commit()
    
    status = 'aktiviert' if module.is_active else 'deaktiviert'
    status_en = 'enabled' if module.is_active else 'disabled'
    flash(f'Modul {module.get_name(lang)} {status if lang == "de" else status_en}.', 'success')
    return redirect(url_for('admin_modules'))


@app.route('/admin/users/<int:user_id>/modules')
@admin_required
def admin_user_modules(user_id):
    """Manage user module assignments"""
    user = User.query.get_or_404(user_id)
    modules = Module.query.filter_by(is_active=True).order_by(Module.nav_order).all()
    lang = session.get('lang', 'de')
    
    # Get user's assigned module IDs
    assigned_module_ids = {um.module_id for um in user.user_modules}
    
    return render_template('admin/user_modules.html', 
                          user=user, 
                          modules=modules, 
                          assigned_module_ids=assigned_module_ids,
                          lang=lang)


@app.route('/admin/users/<int:user_id>/modules', methods=['POST'])
@admin_required
def admin_user_modules_save(user_id):
    """Save user module assignments"""
    user = User.query.get_or_404(user_id)
    lang = session.get('lang', 'de')
    
    # Get selected module IDs from form
    selected_module_ids = set(map(int, request.form.getlist('modules')))
    
    # Get current assignments
    current_module_ids = {um.module_id for um in user.user_modules}
    
    # Remove unselected modules
    for um in list(user.user_modules):
        if um.module_id not in selected_module_ids and not um.module.is_core:
            db.session.delete(um)
    
    # Add new modules
    for module_id in selected_module_ids:
        if module_id not in current_module_ids:
            module = Module.query.get(module_id)
            if module and module.is_active:
                um = UserModule(
                    user_id=user_id,
                    module_id=module_id,
                    granted_by_id=current_user.id
                )
                db.session.add(um)
    
    db.session.commit()
    
    flash('Modulzuweisungen gespeichert.' if lang == 'de' else 'Module assignments saved.', 'success')
    return redirect(url_for('admin_user_modules', user_id=user_id))


@app.route('/admin/users')
@admin_required
def admin_users():
    """User management"""
    users = User.query.order_by(User.name).all()
    return render_template('admin/users.html', users=users)


@app.route('/admin/users/new', methods=['GET', 'POST'])
@admin_required
def admin_user_new():
    """Create new user"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        name = request.form.get('name', '').strip()
        role = request.form.get('role', 'preparer')
        password = request.form.get('password', '')
        
        if User.query.filter_by(email=email).first():
            flash('E-Mail-Adresse bereits vorhanden.', 'danger')
        elif not email or not name or not password:
            flash('Bitte füllen Sie alle Pflichtfelder aus.', 'warning')
        else:
            user = User(email=email, name=name, role=role, is_active=True)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            log_action('CREATE', 'User', user.id, user.email)
            flash(f'Benutzer {name} wurde erstellt.', 'success')
            return redirect(url_for('admin_users'))
    
    return render_template('admin/user_form.html', user=None, roles=UserRole)


@app.route('/admin/users/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def admin_user_edit(user_id):
    """Edit user"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        user.name = request.form.get('name', '').strip()
        user.role = request.form.get('role', 'preparer')
        user.is_active = request.form.get('is_active') == 'on'
        
        new_password = request.form.get('password', '')
        if new_password:
            user.set_password(new_password)
        
        db.session.commit()
        log_action('UPDATE', 'User', user.id, user.email)
        flash(f'Benutzer {user.name} wurde aktualisiert.', 'success')
        return redirect(url_for('admin_users'))
    
    return render_template('admin/user_form.html', user=user, roles=UserRole)


# ============================================================================
# ENTITY MANAGEMENT
# ============================================================================

@app.route('/admin/entities')
@admin_required
def admin_entities():
    """Entity management"""
    entities = Entity.query.order_by(Entity.name).all()
    return render_template('admin/entities.html', entities=entities)


@app.route('/admin/entities/new', methods=['GET', 'POST'])
@admin_required
def admin_entity_new():
    """Create new entity"""
    if request.method == 'POST':
        name_de = request.form.get('name_de', '').strip()
        name_en = request.form.get('name_en', '').strip()
        short_name = request.form.get('short_name', '').strip()
        country = request.form.get('country', 'DE').strip().upper()
        group_id = request.form.get('group_id', type=int)
        
        if not name_de or not name_en:
            flash('Name (DE/EN) ist erforderlich.', 'warning')
        else:
            entity = Entity(
                name=name_de,  # Legacy field gets German name
                name_de=name_de,
                name_en=name_en,
                short_name=short_name or None,
                country=country, 
                group_id=group_id if group_id else None,
                is_active=True
            )
            db.session.add(entity)
            db.session.commit()
            log_action('CREATE', 'Entity', entity.id, entity.name)
            flash(f'Gesellschaft {name_de} wurde erstellt.', 'success')
            return redirect(url_for('admin_entities'))
    
    parent_entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    return render_template('admin/entity_form.html', entity=None, parent_entities=parent_entities)


@app.route('/admin/entities/<int:entity_id>', methods=['GET', 'POST'])
@admin_required
def admin_entity_edit(entity_id):
    """Edit entity"""
    entity = Entity.query.get_or_404(entity_id)
    
    if request.method == 'POST':
        entity.name_de = request.form.get('name_de', '').strip()
        entity.name_en = request.form.get('name_en', '').strip()
        entity.name = entity.name_de  # Keep legacy field in sync
        entity.short_name = request.form.get('short_name', '').strip() or None
        entity.country = request.form.get('country', 'DE').strip().upper()
        group_id = request.form.get('group_id', type=int)
        entity.group_id = group_id if group_id and group_id != entity.id else None
        entity.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        log_action('UPDATE', 'Entity', entity.id, entity.name)
        flash(f'Gesellschaft {entity.name} wurde aktualisiert.', 'success')
        return redirect(url_for('admin_entities'))
    
    parent_entities = Entity.query.filter(Entity.id != entity_id, Entity.is_active == True).order_by(Entity.name).all()
    return render_template('admin/entity_form.html', entity=entity, parent_entities=parent_entities)


@app.route('/admin/entities/<int:entity_id>/delete', methods=['POST'])
@admin_required
def admin_entity_delete(entity_id):
    """Delete entity (soft delete)"""
    entity = Entity.query.get_or_404(entity_id)
    entity.is_active = False
    db.session.commit()
    log_action('DELETE', 'Entity', entity.id, entity.name)
    flash(f'Gesellschaft {entity.name} wurde deaktiviert.', 'success')
    return redirect(url_for('admin_entities'))


# ============================================================================
# USER ENTITY PERMISSIONS
# ============================================================================

@app.route('/admin/users/<int:user_id>/entities')
@admin_required
def admin_user_entities(user_id):
    """Manage entity access for a user"""
    user = User.query.get_or_404(user_id)
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    lang = session.get('lang', 'de')
    
    # Get current permissions as dict for easy lookup
    current_perms = {p.entity_id: p for p in user.entity_permissions}
    
    return render_template('admin/user_entities.html', 
                           user=user, 
                           entities=entities,
                           current_perms=current_perms,
                           access_levels=EntityAccessLevel,
                           lang=lang,
                           t=t)


@app.route('/admin/users/<int:user_id>/entities', methods=['POST'])
@admin_required
def admin_user_entities_save(user_id):
    """Save entity permissions for a user"""
    user = User.query.get_or_404(user_id)
    
    # Get all entity IDs that were submitted
    entity_ids = request.form.getlist('entity_ids', type=int)
    
    # Get existing permissions
    existing_perms = {p.entity_id: p for p in user.entity_permissions}
    
    # Track which entities we've processed
    processed_ids = set()
    
    for entity_id in entity_ids:
        access_level = request.form.get(f'access_level_{entity_id}', 'view')
        inherit = request.form.get(f'inherit_{entity_id}') == 'on'
        
        if entity_id in existing_perms:
            # Update existing permission
            perm = existing_perms[entity_id]
            perm.access_level = access_level
            perm.inherit_to_children = inherit
        else:
            # Create new permission
            perm = UserEntity(
                user_id=user.id,
                entity_id=entity_id,
                access_level=access_level,
                inherit_to_children=inherit,
                granted_by_id=current_user.id
            )
            db.session.add(perm)
        
        processed_ids.add(entity_id)
    
    # Remove permissions that weren't in the form (unchecked)
    for entity_id, perm in existing_perms.items():
        if entity_id not in processed_ids:
            db.session.delete(perm)
    
    db.session.commit()
    log_action('UPDATE', 'UserEntity', user.id, f'Entity permissions for {user.email}')
    
    lang = session.get('lang', 'de')
    if lang == 'de':
        flash(f'Berechtigungen für {user.name} wurden gespeichert.', 'success')
    else:
        flash(f'Permissions for {user.name} saved.', 'success')
    
    return redirect(url_for('admin_user_entities', user_id=user_id))


@app.route('/admin/entities/<int:entity_id>/users')
@admin_required
def admin_entity_users(entity_id):
    """View users with access to an entity"""
    entity = Entity.query.get_or_404(entity_id)
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    lang = session.get('lang', 'de')
    
    # Get current permissions as dict for easy lookup
    current_perms = {p.user_id: p for p in entity.user_permissions}
    
    return render_template('admin/entity_users.html',
                           entity=entity,
                           users=users,
                           current_perms=current_perms,
                           access_levels=EntityAccessLevel,
                           lang=lang,
                           t=t)


@app.route('/admin/entities/<int:entity_id>/users', methods=['POST'])
@admin_required
def admin_entity_users_save(entity_id):
    """Save user permissions for an entity"""
    entity = Entity.query.get_or_404(entity_id)
    
    # Get all user IDs that were submitted
    user_ids = request.form.getlist('user_ids', type=int)
    
    # Get existing permissions
    existing_perms = {p.user_id: p for p in entity.user_permissions}
    
    # Track which users we've processed
    processed_ids = set()
    
    for user_id in user_ids:
        access_level = request.form.get(f'access_level_{user_id}', 'view')
        inherit = request.form.get(f'inherit_{user_id}') == 'on'
        
        if user_id in existing_perms:
            # Update existing permission
            perm = existing_perms[user_id]
            perm.access_level = access_level
            perm.inherit_to_children = inherit
        else:
            # Create new permission
            perm = UserEntity(
                user_id=user_id,
                entity_id=entity.id,
                access_level=access_level,
                inherit_to_children=inherit,
                granted_by_id=current_user.id
            )
            db.session.add(perm)
        
        processed_ids.add(user_id)
    
    # Remove permissions that weren't in the form (unchecked)
    for user_id, perm in existing_perms.items():
        if user_id not in processed_ids:
            db.session.delete(perm)
    
    db.session.commit()
    log_action('UPDATE', 'UserEntity', entity.id, f'User permissions for {entity.name}')
    
    lang = session.get('lang', 'de')
    if lang == 'de':
        flash(f'Berechtigungen für {entity.name} wurden gespeichert.', 'success')
    else:
        flash(f'Permissions for {entity.name} saved.', 'success')
    
    return redirect(url_for('admin_entity_users', entity_id=entity_id))


# ============================================================================
# TAX TYPE MANAGEMENT
# ============================================================================

@app.route('/admin/tax-types')
@admin_required
def admin_tax_types():
    """Tax type management"""
    tax_types = TaxType.query.order_by(TaxType.code).all()
    return render_template('admin/tax_types.html', tax_types=tax_types)


@app.route('/admin/tax-types/new', methods=['GET', 'POST'])
@admin_required
def admin_tax_type_new():
    """Create new tax type"""
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        name_de = request.form.get('name_de', '').strip()
        name_en = request.form.get('name_en', '').strip()
        description_de = request.form.get('description_de', '').strip()
        description_en = request.form.get('description_en', '').strip()
        
        if TaxType.query.filter_by(code=code).first():
            flash('Code bereits vorhanden.', 'danger')
        elif not code or not name_de or not name_en:
            flash('Code und Name (DE/EN) sind erforderlich.', 'warning')
        else:
            tax_type = TaxType(
                code=code, 
                name=name_de,  # Legacy field gets German name
                name_de=name_de,
                name_en=name_en,
                description=description_de or None,  # Legacy field
                description_de=description_de or None,
                description_en=description_en or None,
                is_active=True
            )
            db.session.add(tax_type)
            db.session.commit()
            log_action('CREATE', 'TaxType', tax_type.id, tax_type.code)
            flash(f'Steuerart {code} wurde erstellt.', 'success')
            return redirect(url_for('admin_tax_types'))
    
    return render_template('admin/tax_type_form.html', tax_type=None)


@app.route('/admin/tax-types/<int:tax_type_id>', methods=['GET', 'POST'])
@admin_required
def admin_tax_type_edit(tax_type_id):
    """Edit tax type"""
    tax_type = TaxType.query.get_or_404(tax_type_id)
    
    if request.method == 'POST':
        tax_type.name_de = request.form.get('name_de', '').strip()
        tax_type.name_en = request.form.get('name_en', '').strip()
        tax_type.name = tax_type.name_de  # Keep legacy field in sync
        tax_type.description_de = request.form.get('description_de', '').strip() or None
        tax_type.description_en = request.form.get('description_en', '').strip() or None
        tax_type.description = tax_type.description_de  # Keep legacy field in sync
        tax_type.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        log_action('UPDATE', 'TaxType', tax_type.id, tax_type.code)
        flash(f'Steuerart {tax_type.code} wurde aktualisiert.', 'success')
        return redirect(url_for('admin_tax_types'))
    
    return render_template('admin/tax_type_form.html', tax_type=tax_type)


# ============================================================================
# TEAM MANAGEMENT
# ============================================================================

@app.route('/admin/teams')
@admin_required
def admin_teams():
    """Team management - list all teams"""
    teams = Team.query.order_by(Team.name).all()
    return render_template('admin/teams.html', teams=teams)


@app.route('/admin/teams/new', methods=['GET', 'POST'])
@admin_required
def admin_team_new():
    """Create new team"""
    if request.method == 'POST':
        name_de = request.form.get('name_de', '').strip()
        name_en = request.form.get('name_en', '').strip()
        description_de = request.form.get('description_de', '').strip()
        description_en = request.form.get('description_en', '').strip()
        color = request.form.get('color', '#86BC25').strip()
        manager_id = request.form.get('manager_id', type=int)
        member_ids = request.form.getlist('members', type=int)
        
        if Team.query.filter_by(name=name_de).first():
            flash('Teamname bereits vorhanden.', 'danger')
        elif not name_de or not name_en:
            flash('Teamname (DE/EN) ist erforderlich.', 'warning')
        else:
            team = Team(
                name=name_de,  # Legacy field gets German name
                name_de=name_de,
                name_en=name_en,
                description=description_de or None,  # Legacy field
                description_de=description_de or None,
                description_en=description_en or None,
                color=color,
                manager_id=manager_id if manager_id else None,
                is_active=True
            )
            db.session.add(team)
            db.session.flush()  # Get team.id
            
            # Add members
            for user_id in member_ids:
                user = User.query.get(user_id)
                if user:
                    team.add_member(user)
            
            db.session.commit()
            log_action('CREATE', 'Team', team.id, team.name)
            flash(f'Team "{name_de}" wurde erstellt.', 'success')
            return redirect(url_for('admin_teams'))
    
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('admin/team_form.html', team=None, users=users)


@app.route('/admin/teams/<int:team_id>', methods=['GET', 'POST'])
@admin_required
def admin_team_edit(team_id):
    """Edit team"""
    team = Team.query.get_or_404(team_id)
    
    if request.method == 'POST':
        name_de = request.form.get('name_de', '').strip()
        name_en = request.form.get('name_en', '').strip()
        
        # Check for duplicate name (excluding current team)
        existing = Team.query.filter(Team.name == name_de, Team.id != team_id).first()
        if existing:
            flash('Teamname bereits vorhanden.', 'danger')
        elif not name_de or not name_en:
            flash('Teamname (DE/EN) ist erforderlich.', 'warning')
        else:
            team.name = name_de  # Keep legacy field in sync
            team.name_de = name_de
            team.name_en = name_en
            team.description = request.form.get('description_de', '').strip() or None  # Legacy field
            team.description_de = request.form.get('description_de', '').strip() or None
            team.description_en = request.form.get('description_en', '').strip() or None
            team.color = request.form.get('color', '#86BC25').strip()
            team.manager_id = request.form.get('manager_id', type=int) or None
            team.is_active = request.form.get('is_active') == 'on'
            
            # Update members - get current and new member lists
            new_member_ids = set(request.form.getlist('members', type=int))
            current_member_ids = set(m.id for m in team.members.all())
            
            # Remove members no longer selected
            for user_id in current_member_ids - new_member_ids:
                user = User.query.get(user_id)
                if user:
                    team.remove_member(user)
            
            # Add new members
            for user_id in new_member_ids - current_member_ids:
                user = User.query.get(user_id)
                if user:
                    team.add_member(user)
            
            db.session.commit()
            log_action('UPDATE', 'Team', team.id, team.name)
            flash(f'Team "{team.name}" wurde aktualisiert.', 'success')
            return redirect(url_for('admin_teams'))
    
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('admin/team_form.html', team=team, users=users)


@app.route('/admin/teams/<int:team_id>/delete', methods=['POST'])
@admin_required
def admin_team_delete(team_id):
    """Delete team (soft delete)"""
    team = Team.query.get_or_404(team_id)
    team.is_active = False
    db.session.commit()
    log_action('DELETE', 'Team', team.id, team.name)
    flash(f'Team "{team.name}" wurde deaktiviert.', 'success')
    return redirect(url_for('admin_teams'))


# ============================================================================
# ADMIN: TASK PRESETS (Aufgabenvorlagen)
# ============================================================================

@app.route('/admin/presets')
@admin_required
def admin_presets():
    """Task preset management"""
    category_filter = request.args.get('category', '')
    tax_type_filter = request.args.get('tax_type', '')
    search = request.args.get('search', '').strip()
    
    query = TaskPreset.query
    if category_filter:
        query = query.filter(TaskPreset.category == category_filter)
    if tax_type_filter:
        query = query.filter(TaskPreset.tax_type == tax_type_filter)
    if search:
        query = query.filter(TaskPreset.title.ilike(f'%{search}%') | 
                            TaskPreset.tax_type.ilike(f'%{search}%') |
                            TaskPreset.law_reference.ilike(f'%{search}%'))
    
    presets = query.order_by(TaskPreset.category, TaskPreset.tax_type, TaskPreset.title).all()
    
    # Get unique tax types for filter
    tax_types_used = db.session.query(TaskPreset.tax_type).filter(TaskPreset.tax_type.isnot(None)).distinct().all()
    tax_types_used = sorted([t[0] for t in tax_types_used if t[0]])
    
    return render_template('admin/presets_enhanced.html', 
                           presets=presets, 
                           category_filter=category_filter,
                           tax_type_filter=tax_type_filter,
                           search=search,
                           tax_types_used=tax_types_used)


@app.route('/admin/presets/new', methods=['GET', 'POST'])
@admin_required
def admin_preset_new():
    """Create new task preset"""
    if request.method == 'POST':
        title_de = request.form.get('title_de', '').strip()
        title_en = request.form.get('title_en', '').strip()
        category = request.form.get('category', 'aufgabe')
        tax_type = request.form.get('tax_type', '').strip() or None
        law_reference = request.form.get('law_reference', '').strip() or None
        description_de = request.form.get('description_de', '').strip() or None
        description_en = request.form.get('description_en', '').strip() or None
        
        # Recurrence fields
        is_recurring = request.form.get('is_recurring') == 'on'
        recurrence_frequency = request.form.get('recurrence_frequency', 'monthly')
        recurrence_day_offset = int(request.form.get('recurrence_day_offset', 10) or 10)
        recurrence_rrule = request.form.get('recurrence_rrule', '').strip() or None
        recurrence_end_date_str = request.form.get('recurrence_end_date', '').strip()
        recurrence_end_date = datetime.strptime(recurrence_end_date_str, '%Y-%m-%d').date() if recurrence_end_date_str else None
        default_entity_id = int(request.form.get('default_entity_id')) if request.form.get('default_entity_id') else None
        default_owner_id = int(request.form.get('default_owner_id')) if request.form.get('default_owner_id') else None
        
        if not title_de or not title_en:
            flash('Titel (DE/EN) ist erforderlich.', 'warning')
        else:
            preset = TaskPreset(
                title=title_de,  # Legacy field gets German title
                title_de=title_de,
                title_en=title_en,
                category=category,
                tax_type=tax_type,
                law_reference=law_reference,
                description=description_de,  # Legacy field
                description_de=description_de,
                description_en=description_en,
                source='manual',
                is_active=True,
                is_recurring=is_recurring,
                recurrence_frequency=recurrence_frequency if is_recurring else None,
                recurrence_day_offset=recurrence_day_offset if is_recurring else None,
                recurrence_rrule=recurrence_rrule if is_recurring and recurrence_frequency == 'custom' else None,
                recurrence_end_date=recurrence_end_date if is_recurring else None,
                default_entity_id=default_entity_id if is_recurring else None,
                default_owner_id=default_owner_id if is_recurring else None
            )
            db.session.add(preset)
            db.session.commit()
            log_action('CREATE', 'TaskPreset', preset.id, preset.title[:50])
            flash('Aufgabenvorlage wurde erstellt.', 'success')
            return redirect(url_for('admin_presets'))
    
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    tax_types = TaxType.query.filter_by(is_active=True).order_by(TaxType.code).all()
    return render_template('admin/preset_form_enhanced.html', preset=None, entities=entities, users=users, tax_types=tax_types)


@app.route('/admin/presets/<int:preset_id>', methods=['GET', 'POST'])
@admin_required
def admin_preset_edit(preset_id):
    """Edit task preset"""
    preset = TaskPreset.query.get_or_404(preset_id)
    
    if request.method == 'POST':
        preset.title_de = request.form.get('title_de', '').strip()
        preset.title_en = request.form.get('title_en', '').strip()
        preset.title = preset.title_de  # Keep legacy field in sync
        preset.category = request.form.get('category', 'aufgabe')
        preset.tax_type = request.form.get('tax_type', '').strip() or None
        preset.law_reference = request.form.get('law_reference', '').strip() or None
        preset.description_de = request.form.get('description_de', '').strip() or None
        preset.description_en = request.form.get('description_en', '').strip() or None
        preset.description = preset.description_de  # Keep legacy field in sync
        preset.is_active = request.form.get('is_active') == 'on'
        
        # Recurrence fields
        preset.is_recurring = request.form.get('is_recurring') == 'on'
        if preset.is_recurring:
            preset.recurrence_frequency = request.form.get('recurrence_frequency', 'monthly')
            preset.recurrence_day_offset = int(request.form.get('recurrence_day_offset', 10) or 10)
            recurrence_rrule = request.form.get('recurrence_rrule', '').strip()
            preset.recurrence_rrule = recurrence_rrule if preset.recurrence_frequency == 'custom' else None
            recurrence_end_date_str = request.form.get('recurrence_end_date', '').strip()
            preset.recurrence_end_date = datetime.strptime(recurrence_end_date_str, '%Y-%m-%d').date() if recurrence_end_date_str else None
            preset.default_entity_id = int(request.form.get('default_entity_id')) if request.form.get('default_entity_id') else None
            preset.default_owner_id = int(request.form.get('default_owner_id')) if request.form.get('default_owner_id') else None
        else:
            preset.recurrence_frequency = None
            preset.recurrence_day_offset = None
            preset.recurrence_rrule = None
            preset.recurrence_end_date = None
            preset.default_entity_id = None
            preset.default_owner_id = None
        
        db.session.commit()
        log_action('UPDATE', 'TaskPreset', preset.id, preset.title[:50])
        flash('Aufgabenvorlage wurde aktualisiert.', 'success')
        return redirect(url_for('admin_presets'))
    
    entities = Entity.query.filter_by(is_active=True).order_by(Entity.name).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    tax_types = TaxType.query.filter_by(is_active=True).order_by(TaxType.code).all()
    return render_template('admin/preset_form_enhanced.html', preset=preset, entities=entities, users=users, tax_types=tax_types)


@app.route('/admin/presets/<int:preset_id>/delete', methods=['POST'])
@admin_required
def admin_preset_delete(preset_id):
    """Delete task preset"""
    preset = TaskPreset.query.get_or_404(preset_id)
    title = preset.title
    db.session.delete(preset)
    db.session.commit()
    log_action('DELETE', 'TaskPreset', preset_id, title[:50])
    flash('Aufgabenvorlage wurde gelöscht.', 'success')
    return redirect(url_for('admin_presets'))


# API endpoints for preset management
@app.route('/api/presets/<int:preset_id>', methods=['GET'])
@login_required
def api_preset_get(preset_id):
    """Get preset data for quick edit"""
    preset = TaskPreset.query.get_or_404(preset_id)
    return jsonify({
        'id': preset.id,
        'title_de': preset.title_de,
        'title_en': preset.title_en,
        'category': preset.category,
        'tax_type': preset.tax_type,
        'law_reference': preset.law_reference,
        'description_de': preset.description_de,
        'description_en': preset.description_en,
        'is_active': preset.is_active,
        'is_recurring': preset.is_recurring
    })


@app.route('/api/presets/<int:preset_id>', methods=['PATCH'])
@admin_required
def api_preset_update(preset_id):
    """Quick update preset"""
    preset = TaskPreset.query.get_or_404(preset_id)
    data = request.get_json()
    
    if 'title_de' in data:
        preset.title_de = data['title_de']
        preset.title = data['title_de']  # Keep legacy field in sync
    if 'title_en' in data:
        preset.title_en = data['title_en']
    if 'tax_type' in data:
        preset.tax_type = data['tax_type'] or None
    if 'law_reference' in data:
        preset.law_reference = data['law_reference'] or None
    if 'is_active' in data:
        preset.is_active = data['is_active']
    
    db.session.commit()
    log_action('UPDATE', 'TaskPreset', preset.id, f"Quick edit: {preset.title[:30]}")
    return jsonify({'success': True})


@app.route('/api/presets/bulk-toggle-active', methods=['POST'])
@admin_required
def api_presets_bulk_toggle_active():
    """Bulk activate/deactivate presets"""
    data = request.get_json()
    ids = data.get('ids', [])
    active = data.get('active', True)
    
    count = TaskPreset.query.filter(TaskPreset.id.in_(ids)).update(
        {'is_active': active}, synchronize_session=False
    )
    db.session.commit()
    log_action('BULK_UPDATE', 'TaskPreset', None, f"{'Activated' if active else 'Deactivated'} {count} presets")
    return jsonify({'success': True, 'count': count})


@app.route('/api/presets/bulk-delete', methods=['POST'])
@admin_required
def api_presets_bulk_delete():
    """Bulk delete presets"""
    data = request.get_json()
    ids = data.get('ids', [])
    
    count = TaskPreset.query.filter(TaskPreset.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    log_action('BULK_DELETE', 'TaskPreset', None, f"Deleted {count} presets")
    return jsonify({'success': True, 'count': count})


# Custom Field CRUD API Endpoints
@app.route('/api/preset-fields', methods=['POST'])
@admin_required
def api_preset_field_create():
    """Create a new custom field for a preset"""
    from models import PresetCustomField
    data = request.get_json()
    
    preset_id = data.get('preset_id')
    if not preset_id:
        return jsonify({'error': 'preset_id required'}), 400
    
    # Get next sort order
    max_order = db.session.query(db.func.max(PresetCustomField.sort_order)).filter(
        PresetCustomField.preset_id == preset_id
    ).scalar() or 0
    
    field = PresetCustomField(
        preset_id=preset_id,
        name=data.get('name', '').lower().replace(' ', '_'),
        label_de=data.get('label_de', ''),
        label_en=data.get('label_en', ''),
        field_type=data.get('field_type', 'text'),
        is_required=data.get('is_required', False),
        placeholder_de=data.get('placeholder_de', ''),
        placeholder_en=data.get('placeholder_en', ''),
        default_value=data.get('default_value', ''),
        options=data.get('options', ''),
        help_text_de=data.get('help_text_de', ''),
        help_text_en=data.get('help_text_en', ''),
        condition_field=data.get('condition_field', ''),
        condition_operator=data.get('condition_operator', ''),
        condition_value=data.get('condition_value', ''),
        sort_order=max_order + 1
    )
    
    db.session.add(field)
    db.session.commit()
    log_action('CREATE', 'PresetCustomField', field.id, f"Created custom field {field.name}")
    
    return jsonify({'success': True, 'id': field.id})


@app.route('/api/preset-fields/<int:field_id>', methods=['GET'])
@admin_required
def api_preset_field_get(field_id):
    """Get a custom field by ID"""
    from models import PresetCustomField
    field = PresetCustomField.query.get_or_404(field_id)
    
    return jsonify({
        'id': field.id,
        'preset_id': field.preset_id,
        'name': field.name,
        'label_de': field.label_de,
        'label_en': field.label_en,
        'field_type': field.field_type,
        'is_required': field.is_required,
        'placeholder_de': field.placeholder_de,
        'placeholder_en': field.placeholder_en,
        'default_value': field.default_value,
        'options': field.options,
        'help_text_de': field.help_text_de,
        'help_text_en': field.help_text_en,
        'condition_field': field.condition_field,
        'condition_operator': field.condition_operator,
        'condition_value': field.condition_value,
        'sort_order': field.sort_order
    })


@app.route('/api/preset-fields/<int:field_id>', methods=['PUT'])
@admin_required
def api_preset_field_update(field_id):
    """Update a custom field"""
    from models import PresetCustomField
    field = PresetCustomField.query.get_or_404(field_id)
    data = request.get_json()
    
    field.name = data.get('name', field.name).lower().replace(' ', '_')
    field.label_de = data.get('label_de', field.label_de)
    field.label_en = data.get('label_en', field.label_en)
    field.field_type = data.get('field_type', field.field_type)
    field.is_required = data.get('is_required', field.is_required)
    field.placeholder_de = data.get('placeholder_de', field.placeholder_de)
    field.placeholder_en = data.get('placeholder_en', field.placeholder_en)
    field.default_value = data.get('default_value', field.default_value)
    field.options = data.get('options', field.options)
    field.help_text_de = data.get('help_text_de', field.help_text_de)
    field.help_text_en = data.get('help_text_en', field.help_text_en)
    field.condition_field = data.get('condition_field', field.condition_field)
    field.condition_operator = data.get('condition_operator', field.condition_operator)
    field.condition_value = data.get('condition_value', field.condition_value)
    
    db.session.commit()
    log_action('UPDATE', 'PresetCustomField', field.id, f"Updated custom field {field.name}")
    
    return jsonify({'success': True})


@app.route('/api/preset-fields/<int:field_id>', methods=['DELETE'])
@admin_required
def api_preset_field_delete(field_id):
    """Delete a custom field"""
    from models import PresetCustomField
    field = PresetCustomField.query.get_or_404(field_id)
    
    db.session.delete(field)
    db.session.commit()
    log_action('DELETE', 'PresetCustomField', field_id, f"Deleted custom field {field.name}")
    
    return jsonify({'success': True})


@app.route('/admin/presets/export')
@admin_required
def admin_preset_export():
    """Export all presets as JSON with custom fields"""
    from models import PresetCustomField
    presets = TaskPreset.query.order_by(TaskPreset.category, TaskPreset.tax_type).all()
    
    data = []
    for p in presets:
        preset_data = {
            'category': p.category,
            'tax_type': p.tax_type,
            'title_de': p.title_de,
            'title_en': p.title_en,
            'law_reference': p.law_reference,
            'description_de': p.description_de,
            'description_en': p.description_en,
            'is_recurring': p.is_recurring,
            'recurrence_frequency': p.recurrence_frequency,
            'recurrence_day_offset': p.recurrence_day_offset,
            'recurrence_rrule': p.recurrence_rrule,
            'is_active': p.is_active,
            'custom_fields': []
        }
        
        # Include custom fields
        for field in p.custom_fields.all():
            preset_data['custom_fields'].append({
                'name': field.name,
                'label_de': field.label_de,
                'label_en': field.label_en,
                'field_type': field.field_type,
                'is_required': field.is_required,
                'placeholder_de': field.placeholder_de,
                'placeholder_en': field.placeholder_en,
                'default_value': field.default_value,
                'options': field.options,
                'help_text_de': field.help_text_de,
                'help_text_en': field.help_text_en,
                'condition_field': field.condition_field,
                'condition_operator': field.condition_operator,
                'condition_value': field.condition_value,
                'sort_order': field.sort_order
            })
        
        data.append(preset_data)
    
    import json as json_module
    response = make_response(json_module.dumps(data, ensure_ascii=False, indent=2))
    response.headers['Content-Type'] = 'application/json'
    response.headers['Content-Disposition'] = 'attachment; filename=presets_export.json'
    return response


@app.route('/admin/presets/import', methods=['POST'])
@admin_required
def admin_preset_import():
    """Import task presets from JSON or Excel file"""
    import json as json_module
    
    if 'file' not in request.files:
        flash('Keine Datei ausgewählt.', 'warning')
        return redirect(url_for('admin_presets'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Keine Datei ausgewählt.', 'warning')
        return redirect(url_for('admin_presets'))
    
    filename = file.filename.lower()
    imported = 0
    skipped = 0
    fields_imported = 0
    
    try:
        if filename.endswith('.json'):
            from models import PresetCustomField
            data = json_module.load(file)
            
            # Handle enhanced export format (list of presets with custom_fields)
            if isinstance(data, list) and len(data) > 0 and 'title_de' in data[0]:
                for record in data:
                    title = record.get('title_de') or record.get('title')
                    if not title:
                        skipped += 1
                        continue
                    category = record.get('category', 'aufgabe')
                    existing = TaskPreset.query.filter_by(title_de=title, category=category).first()
                    if existing:
                        skipped += 1
                        continue
                    preset = TaskPreset(
                        category=category,
                        title=title,
                        title_de=title,
                        title_en=record.get('title_en'),
                        tax_type=record.get('tax_type'),
                        law_reference=record.get('law_reference'),
                        description=record.get('description_de'),
                        description_de=record.get('description_de'),
                        description_en=record.get('description_en'),
                        is_recurring=record.get('is_recurring', False),
                        recurrence_frequency=record.get('recurrence_frequency'),
                        recurrence_day_offset=record.get('recurrence_day_offset'),
                        recurrence_rrule=record.get('recurrence_rrule'),
                        is_active=record.get('is_active', True),
                        source='json'
                    )
                    db.session.add(preset)
                    db.session.flush()  # Get preset.id
                    imported += 1
                    
                    # Import custom fields
                    for field_data in record.get('custom_fields', []):
                        field = PresetCustomField(
                            preset_id=preset.id,
                            name=field_data.get('name', ''),
                            label_de=field_data.get('label_de', ''),
                            label_en=field_data.get('label_en'),
                            field_type=field_data.get('field_type', 'text'),
                            is_required=field_data.get('is_required', False),
                            placeholder_de=field_data.get('placeholder_de'),
                            placeholder_en=field_data.get('placeholder_en'),
                            default_value=field_data.get('default_value'),
                            options=field_data.get('options'),
                            help_text_de=field_data.get('help_text_de'),
                            help_text_en=field_data.get('help_text_en'),
                            condition_field=field_data.get('condition_field'),
                            condition_operator=field_data.get('condition_operator'),
                            condition_value=field_data.get('condition_value'),
                            sort_order=field_data.get('sort_order', 0)
                        )
                        db.session.add(field)
                        fields_imported += 1
            
            # Handle Antraege.json format
            elif 'sheets' in data:
                for sheet_name, records in data['sheets'].items():
                    for record in records:
                        title = record.get('Zweck des Antrags') or record.get('title')
                        if not title:
                            skipped += 1
                            continue
                        existing = TaskPreset.query.filter_by(title=title, category='antrag').first()
                        if existing:
                            skipped += 1
                            continue
                        preset = TaskPreset(
                            category='antrag',
                            title=title,
                            law_reference=record.get('§ Paragraph') or record.get('law_reference'),
                            tax_type=record.get('Gesetz') or record.get('tax_type'),
                            description=record.get('Erläuterung') or record.get('description'),
                            source='json',
                            is_active=True
                        )
                        db.session.add(preset)
                        imported += 1
            
            # Handle steuerarten_aufgaben.json format
            elif 'records' in data:
                for record in data['records']:
                    title = record.get('aufgabe') or record.get('title')
                    if not title:
                        skipped += 1
                        continue
                    existing = TaskPreset.query.filter_by(title=title, category='aufgabe').first()
                    if existing:
                        skipped += 1
                        continue
                    preset = TaskPreset(
                        category='aufgabe',
                        title=title,
                        tax_type=record.get('steuerart') or record.get('tax_type'),
                        description=record.get('description'),
                        source='json',
                        is_active=True
                    )
                    db.session.add(preset)
                    imported += 1
        
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Find header row
            header_row = None
            headers = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                for col_idx, cell in enumerate(row):
                    if cell:
                        cell_lower = str(cell).lower()
                        if 'titel' in cell_lower or 'aufgabe' in cell_lower or 'zweck' in cell_lower:
                            header_row = row_idx
                            for i, h in enumerate(row):
                                if h:
                                    headers[str(h).lower().strip()] = i
                            break
                if header_row:
                    break
            
            if not header_row:
                flash('Keine gültige Kopfzeile gefunden.', 'danger')
                return redirect(url_for('admin_presets'))
            
            # Find title column
            title_col = None
            for key in ['titel', 'aufgabe', 'zweck des antrags', 'zweck']:
                if key in headers:
                    title_col = headers[key]
                    break
            
            if title_col is None:
                flash('Spalte "Titel" oder "Aufgabe" nicht gefunden.', 'danger')
                return redirect(url_for('admin_presets'))
            
            category_col = headers.get('kategorie') or headers.get('category')
            tax_type_col = headers.get('steuerart') or headers.get('tax_type') or headers.get('gesetz')
            law_ref_col = headers.get('§ paragraph') or headers.get('paragraph') or headers.get('law_reference')
            desc_col = headers.get('beschreibung') or headers.get('description') or headers.get('erläuterung')
            
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                title = row[title_col] if title_col < len(row) else None
                if not title or not str(title).strip():
                    skipped += 1
                    continue
                
                title = str(title).strip()
                category = 'aufgabe'
                if category_col is not None and category_col < len(row) and row[category_col]:
                    if 'antrag' in str(row[category_col]).lower():
                        category = 'antrag'
                
                existing = TaskPreset.query.filter_by(title=title, category=category).first()
                if existing:
                    skipped += 1
                    continue
                
                preset = TaskPreset(
                    category=category,
                    title=title,
                    tax_type=str(row[tax_type_col]).strip() if tax_type_col is not None and tax_type_col < len(row) and row[tax_type_col] else None,
                    law_reference=str(row[law_ref_col]).strip() if law_ref_col is not None and law_ref_col < len(row) and row[law_ref_col] else None,
                    description=str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else None,
                    source='excel',
                    is_active=True
                )
                db.session.add(preset)
                imported += 1
        else:
            flash('Ungültiges Dateiformat. Erlaubt: .json, .xlsx', 'danger')
            return redirect(url_for('admin_presets'))
        
        db.session.commit()
        log_action('IMPORT', 'TaskPreset', None, f'{imported} imported, {skipped} skipped, {fields_imported} fields')
        if fields_imported > 0:
            flash(f'{imported} Vorlagen importiert (inkl. {fields_imported} Felder), {skipped} übersprungen.', 'success')
        else:
            flash(f'{imported} Vorlagen importiert, {skipped} übersprungen.', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Import: {str(e)}', 'danger')
    
    return redirect(url_for('admin_presets'))


@app.route('/admin/presets/template')
@admin_required
def admin_preset_template():
    """Download Excel template for import"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aufgabenvorlagen"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['Kategorie', 'Titel', 'Steuerart', '§ Paragraph', 'Beschreibung']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    examples = [
        ['aufgabe', 'USt-Voranmeldung einreichen', 'Umsatzsteuer', '', 'Monatliche USt-Voranmeldung'],
        ['aufgabe', 'Körperschaftsteuererklärung erstellen', 'Körperschaftsteuer', '', 'Jährliche KSt-Erklärung'],
        ['antrag', 'Unbeschränkte Einkommensteuerpflicht', 'EStG', '§1 (3) EStG', 'Antrag für ausländische Mitarbeiter'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Aufgabenvorlagen_Vorlage.xlsx')


@app.route('/admin/presets/seed', methods=['POST'])
@admin_required
def admin_preset_seed():
    """Seed presets from JSON files in data folder"""
    import json as json_module
    import os as os_module
    
    data_dir = os_module.path.join(os_module.path.dirname(__file__), 'data')
    imported = 0
    skipped = 0
    
    antraege_path = os_module.path.join(data_dir, 'Antraege.json')
    if os_module.path.exists(antraege_path):
        with open(antraege_path, 'r', encoding='utf-8') as f:
            data = json_module.load(f)
            if 'sheets' in data:
                for sheet_name, records in data['sheets'].items():
                    for record in records:
                        title = record.get('Zweck des Antrags')
                        if not title:
                            continue
                        existing = TaskPreset.query.filter_by(title=title, category='antrag').first()
                        if existing:
                            skipped += 1
                            continue
                        preset = TaskPreset(category='antrag', title=title, law_reference=record.get('§ Paragraph'),
                                          tax_type=record.get('Gesetz'), description=record.get('Erläuterung'),
                                          source='json', is_active=True)
                        db.session.add(preset)
                        imported += 1
    
    aufgaben_path = os_module.path.join(data_dir, 'steuerarten_aufgaben.json')
    if os_module.path.exists(aufgaben_path):
        with open(aufgaben_path, 'r', encoding='utf-8') as f:
            data = json_module.load(f)
            if 'records' in data:
                for record in data['records']:
                    title = record.get('aufgabe')
                    if not title:
                        continue
                    existing = TaskPreset.query.filter_by(title=title, category='aufgabe').first()
                    if existing:
                        skipped += 1
                        continue
                    preset = TaskPreset(category='aufgabe', title=title, tax_type=record.get('steuerart'),
                                      source='json', is_active=True)
                    db.session.add(preset)
                    imported += 1
    
    db.session.commit()
    log_action('SEED', 'TaskPreset', None, f'{imported} seeded, {skipped} skipped')
    flash(f'{imported} Vorlagen aus JSON-Dateien importiert, {skipped} übersprungen.', 'success')
    return redirect(url_for('admin_presets'))


@app.route('/api/presets')
@login_required
def api_presets():
    """API endpoint for fetching presets"""
    category = request.args.get('category', '')
    search = request.args.get('search', '').strip()
    
    query = TaskPreset.query.filter(TaskPreset.is_active == True)
    if category:
        query = query.filter(TaskPreset.category == category)
    if search:
        query = query.filter(TaskPreset.title.ilike(f'%{search}%'))
    
    presets = query.order_by(TaskPreset.title).limit(50).all()
    return jsonify([p.to_dict() for p in presets])


@app.route('/api/tasks/<int:task_id>/approval-status')
@login_required
def api_task_approval_status(task_id):
    """API endpoint for task approval status"""
    task = Task.query.get_or_404(task_id)
    status = ApprovalService.get_approval_status(task)
    
    return jsonify({
        'total_reviewers': status.total_reviewers,
        'approved_count': status.approved_count,
        'rejected_count': status.rejected_count,
        'pending_count': status.pending_count,
        'is_complete': status.is_complete,
        'is_rejected': status.is_rejected,
        'progress_percent': status.progress_percent,
        'pending_reviewers': [{'id': u.id, 'name': u.name, 'email': u.email} for u in status.pending_reviewers],
        'approved_reviewers': [{'id': u.id, 'name': u.name, 'email': u.email} for u in status.approved_reviewers],
        'rejected_reviewers': [{'id': u.id, 'name': u.name, 'email': u.email} for u in status.rejected_reviewers],
        'summary': ApprovalService.format_approval_summary(task, session.get('lang', 'de'))
    })


@app.route('/api/tasks/<int:task_id>/workflow-timeline')
@login_required
def api_task_workflow_timeline(task_id):
    """API endpoint for task workflow timeline"""
    task = Task.query.get_or_404(task_id)
    timeline = WorkflowService.get_workflow_timeline(task)
    lang = session.get('lang', 'de')
    
    # Format for JSON response
    formatted = []
    for event in timeline:
        formatted.append({
            'timestamp': event['timestamp'].isoformat() if event['timestamp'] else None,
            'action': event['action'],
            'label': event[f'action_label_{lang}'],
            'user': event['user'].name if event.get('user') else None,
            'note': event.get('note'),
            'icon': event['icon'],
            'color': event['color']
        })
    
    return jsonify(formatted)


# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(404)
def not_found_error(error):
    """404 error handler"""
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    """500 error handler"""
    db.session.rollback()
    return render_template('errors/500.html'), 500


# ============================================================================
# CLI COMMANDS
# ============================================================================

@app.cli.command()
def initdb():
    """Initialize the database"""
    db.create_all()
    print('Database initialized.')


@app.cli.command('sync-modules')
def sync_modules():
    """Sync module definitions to database"""
    from modules import ModuleRegistry
    
    count = 0
    for module_class in ModuleRegistry.all():
        module_class.sync_to_db()
        print(f'✓ Module synced: {module_class.code}')
        count += 1
    
    print(f'\n{count} modules synchronized to database.')


@app.cli.command()
def createadmin():
    """Create admin user"""
    admin = User(
        email='admin@example.com',
        name='Administrator',
        role='admin',
        is_active=True
    )
    admin.set_password('admin123')
    db.session.add(admin)
    db.session.commit()
    print('Admin user created: admin@example.com / admin123')


@app.cli.command()
@click.option('--days', default=7, help='Send reminders for tasks due within this many days')
@click.option('--include-overdue', is_flag=True, default=True, help='Also send reminders for overdue tasks')
@click.option('--dry-run', is_flag=True, help='Show what would be sent without actually sending')
def send_due_reminders(days, include_overdue, dry_run):
    """Send email reminders for tasks due soon or overdue"""
    from datetime import date, timedelta
    
    today = date.today()
    soon = today + timedelta(days=days)
    
    # Find tasks that are due soon and not completed
    query = Task.query.filter(
        Task.status.notin_(['completed', 'rejected']),
        Task.due_date <= soon
    )
    
    if not include_overdue:
        query = query.filter(Task.due_date >= today)
    
    tasks = query.all()
    
    if dry_run:
        print(f"[DRY RUN] Would send reminders for {len(tasks)} tasks:")
    else:
        print(f"Sending reminders for {len(tasks)} tasks...")
    
    sent_count = 0
    for task in tasks:
        days_until_due = (task.due_date - today).days
        
        # Determine recipients (owner and reviewers)
        recipients = []
        if task.owner:
            recipients.append(task.owner)
        for tr in task.reviewers:
            if tr.user not in recipients:
                recipients.append(tr.user)
        
        for user in recipients:
            if dry_run:
                status = "overdue" if days_until_due < 0 else f"due in {days_until_due} days"
                print(f"  → {task.title} ({status}) → {user.email}")
            else:
                if email_service.send_due_reminder(task, user, days_until_due, 'de'):
                    sent_count += 1
    
    if not dry_run:
        print(f"Sent {sent_count} reminder emails.")


@app.cli.command('generate-recurring-tasks')
@click.option('--year', default=None, type=int, help='Year to generate tasks for (default: current year)')
@click.option('--preset-id', default=None, type=int, help='Generate tasks only for a specific preset ID')
@click.option('--entity-id', default=None, type=int, help='Generate tasks only for a specific entity ID')
@click.option('--dry-run', is_flag=True, help='Show what would be created without actually creating')
@click.option('--force', is_flag=True, help='Force creation even if tasks already exist')
def generate_recurring_tasks(year, preset_id, entity_id, dry_run, force):
    """Generate task instances from recurring presets"""
    from datetime import date
    
    # Default to current year
    if not year:
        year = date.today().year
    
    print(f"Generating recurring tasks for year {year}...")
    
    if preset_id:
        # Generate for specific preset
        preset = TaskPreset.query.get(preset_id)
        if not preset:
            print(f"Error: Preset with ID {preset_id} not found.")
            return
        
        if not preset.is_recurring:
            print(f"Error: Preset '{preset.title}' is not configured as recurring.")
            return
        
        entities = None
        if entity_id:
            entity = Entity.query.get(entity_id)
            if not entity:
                print(f"Error: Entity with ID {entity_id} not found.")
                return
            entities = [entity]
        
        if dry_run:
            periods = RecurrenceService.get_period_dates(
                preset.recurrence_frequency,
                year,
                preset.recurrence_day_offset or 10
            )
            print(f"\n[DRY RUN] Would create tasks for preset '{preset.title}':")
            print(f"  Frequency: {preset.recurrence_frequency}")
            print(f"  Periods: {len(periods)}")
            for period, due_date in periods:
                print(f"    - {period or 'Annual'}: due {due_date.strftime('%d.%m.%Y')}")
            return
        
        tasks = RecurrenceService.generate_tasks_from_preset(preset, year, entities, force=force)
        db.session.commit()
        print(f"Created {len(tasks)} tasks from preset '{preset.title}'.")
    else:
        # Generate for all recurring presets
        if dry_run:
            presets = TaskPreset.query.filter_by(is_recurring=True, is_active=True).all()
            print(f"\n[DRY RUN] Would process {len(presets)} recurring presets:")
            for preset in presets:
                periods = RecurrenceService.get_period_dates(
                    preset.recurrence_frequency,
                    year,
                    preset.recurrence_day_offset or 10
                )
                entity_count = Entity.query.filter_by(is_active=True).count() if not preset.default_entity_id else 1
                print(f"  - {preset.title}: {len(periods)} periods × {entity_count} entities = {len(periods) * entity_count} potential tasks")
            return
        
        stats = RecurrenceService.generate_all_recurring_tasks(year, dry_run=False)
        print(f"\nGeneration complete:")
        print(f"  Presets processed: {stats['presets_processed']}")
        print(f"  Tasks created: {stats['tasks_created']}")
        if stats['errors']:
            print(f"  Errors: {len(stats['errors'])}")
            for error in stats['errors']:
                print(f"    - {error}")


@app.cli.command('purge-archive')
@click.option('--days', default=365, help='Delete archived tasks older than N days')
@click.option('--dry-run', is_flag=True, help='Show what would be deleted without actually deleting')
@click.option('--force', is_flag=True, help='Skip confirmation prompt')
def purge_archive(days, dry_run, force):
    """Permanently delete archived tasks older than N days.
    
    Examples:
        flask purge-archive --days=365 --dry-run
        flask purge-archive --days=90 --force
    """
    from datetime import datetime, timedelta
    
    cutoff_date = datetime.utcnow() - timedelta(days=days)
    
    # Find archived tasks older than cutoff
    old_tasks = Task.query.filter(
        Task.is_archived == True,
        Task.archived_at < cutoff_date
    ).all()
    
    if not old_tasks:
        print(f"No archived tasks older than {days} days found.")
        return
    
    print(f"\nFound {len(old_tasks)} archived task(s) older than {days} days:")
    print("-" * 60)
    
    for task in old_tasks:
        age_days = (datetime.utcnow() - task.archived_at).days if task.archived_at else 0
        print(f"  [{task.id}] {task.title}")
        print(f"      Archived: {task.archived_at.strftime('%Y-%m-%d') if task.archived_at else 'Unknown'} ({age_days} days ago)")
        print(f"      Entity: {task.entity.name if task.entity else '-'}")
    
    print("-" * 60)
    
    if dry_run:
        print("\n[DRY RUN] No tasks were deleted.")
        return
    
    if not force:
        confirm = input(f"\nPermanently delete {len(old_tasks)} task(s)? This cannot be undone! [y/N]: ")
        if confirm.lower() != 'y':
            print("Aborted.")
            return
    
    deleted_count = 0
    for task in old_tasks:
        task_id = task.id
        task_title = task.title
        
        # Delete related records first
        TaskEvidence.query.filter_by(task_id=task_id).delete()
        Comment.query.filter_by(task_id=task_id).delete()
        TaskReviewer.query.filter_by(task_id=task_id).delete()
        Notification.query.filter(Notification.entity_type == 'task', Notification.entity_id == task_id).delete()
        
        db.session.delete(task)
        log_action('PURGE', 'Task', task_id, task_title, 'archived', 'deleted')
        deleted_count += 1
    
    db.session.commit()
    print(f"\n✓ Permanently deleted {deleted_count} task(s).")


@app.cli.command()
def seed():
    """Seed database with sample data for development"""
    from datetime import date, timedelta
    
    # Create sample users
    users_data = [
        {'email': 'manager@example.com', 'name': 'Max Manager', 'role': 'manager'},
        {'email': 'reviewer@example.com', 'name': 'Rita Reviewer', 'role': 'reviewer'},
        {'email': 'preparer@example.com', 'name': 'Peter Preparer', 'role': 'preparer'},
        {'email': 'readonly@example.com', 'name': 'Rolf Readonly', 'role': 'readonly'},
    ]
    
    for u in users_data:
        if not User.query.filter_by(email=u['email']).first():
            user = User(email=u['email'], name=u['name'], role=u['role'], is_active=True)
            user.set_password('password123')
            db.session.add(user)
    
    db.session.commit()
    print(f'Created {len(users_data)} sample users')
    
    # Create sample tax types
    tax_types_data = [
        {'code': 'KSt', 'name': 'Körperschaftsteuer', 'description': 'Corporate income tax'},
        {'code': 'USt', 'name': 'Umsatzsteuer', 'description': 'Value added tax (VAT)'},
        {'code': 'GewSt', 'name': 'Gewerbesteuer', 'description': 'Trade tax'},
        {'code': 'LSt', 'name': 'Lohnsteuer', 'description': 'Wage tax'},
        {'code': 'ESt', 'name': 'Einkommensteuer', 'description': 'Income tax'},
        {'code': 'ErbSt', 'name': 'Erbschaftsteuer', 'description': 'Inheritance tax'},
        {'code': 'GrSt', 'name': 'Grundsteuer', 'description': 'Property tax'},
    ]
    
    for tt in tax_types_data:
        if not TaxType.query.filter_by(code=tt['code']).first():
            tax_type = TaxType(**tt, is_active=True)
            db.session.add(tax_type)
    
    db.session.commit()
    print(f'Created {len(tax_types_data)} tax types')
    
    # Create sample entities
    entities_data = [
        {'name': 'Deloitte Germany GmbH', 'short_name': 'DEU', 'country': 'DE'},
        {'name': 'Deloitte Austria GmbH', 'short_name': 'AUT', 'country': 'AT'},
        {'name': 'Deloitte Switzerland AG', 'short_name': 'CHE', 'country': 'CH'},
        {'name': 'Tech Subsidiary GmbH', 'short_name': 'TECH', 'country': 'DE'},
        {'name': 'Finance Holding SE', 'short_name': 'FIN', 'country': 'DE'},
    ]
    
    for e in entities_data:
        if not Entity.query.filter_by(name=e['name']).first():
            entity = Entity(**e, is_active=True)
            db.session.add(entity)
    
    db.session.commit()
    print(f'Created {len(entities_data)} entities')
    
    # Create sample task templates
    kst = TaxType.query.filter_by(code='KSt').first()
    ust = TaxType.query.filter_by(code='USt').first()
    gewst = TaxType.query.filter_by(code='GewSt').first()
    
    templates_data = [
        {'tax_type_id': kst.id, 'keyword': 'KSt-Voranmeldung', 'description': 'Körperschaftsteuer Voranmeldung', 'default_recurrence': 'quarterly'},
        {'tax_type_id': kst.id, 'keyword': 'KSt-Erklärung', 'description': 'Jährliche Körperschaftsteuererklärung', 'default_recurrence': 'annual'},
        {'tax_type_id': ust.id, 'keyword': 'USt-Voranmeldung', 'description': 'Monatliche Umsatzsteuer-Voranmeldung', 'default_recurrence': 'monthly'},
        {'tax_type_id': ust.id, 'keyword': 'USt-Erklärung', 'description': 'Jährliche Umsatzsteuererklärung', 'default_recurrence': 'annual'},
        {'tax_type_id': gewst.id, 'keyword': 'GewSt-Erklärung', 'description': 'Jährliche Gewerbesteuererklärung', 'default_recurrence': 'annual'},
    ]
    
    for t in templates_data:
        if not TaskTemplate.query.filter_by(keyword=t['keyword']).first():
            template = TaskTemplate(**t, is_active=True)
            db.session.add(template)
    
    db.session.commit()
    print(f'Created {len(templates_data)} task templates')
    
    # Create sample tasks for current year
    year = date.today().year
    entities = Entity.query.filter_by(is_active=True).all()
    owner = User.query.filter_by(role='preparer').first()
    reviewer = User.query.filter_by(role='reviewer').first()
    
    tasks_created = 0
    for entity in entities[:3]:  # First 3 entities
        # Monthly USt tasks
        ust_template = TaskTemplate.query.filter_by(keyword='USt-Voranmeldung').first()
        for month in range(1, 13):
            due = date(year, month, 10) if month < 12 else date(year, 12, 10)
            if due < date.today():
                due = date.today() + timedelta(days=30)
            
            existing = Task.query.filter_by(
                entity_id=entity.id, 
                template_id=ust_template.id,
                year=year,
                period=f'M{month:02d}'
            ).first()
            
            if not existing:
                task = Task(
                    template_id=ust_template.id,
                    entity_id=entity.id,
                    year=year,
                    period=f'M{month:02d}',
                    title=f'{ust_template.keyword} {month:02d}/{year}',
                    description=ust_template.description,
                    due_date=due,
                    status='draft',
                    owner_id=owner.id if owner else None,
                    reviewer_id=reviewer.id if reviewer else None
                )
                db.session.add(task)
                tasks_created += 1
    
    db.session.commit()
    print(f'Created {tasks_created} sample tasks')
    print('Seed complete!')


# ============================================================================
# EXPORT ROUTES
# ============================================================================

@app.route('/tasks/export/excel')
@login_required
def export_tasks_excel():
    """Export filtered task list to Excel"""
    from datetime import date
    from flask import Response
    
    lang = session.get('lang', 'de')
    
    # Get filter parameters (same as task_list)
    status_filter = request.args.get('status', '')
    entity_filter = request.args.get('entity', type=int)
    tax_type_filter = request.args.get('tax_type', type=int)
    year_filter = request.args.get('year', type=int, default=date.today().year)
    
    # Base query
    query = Task.query
    
    # Apply role-based filtering
    if not (current_user.is_admin() or current_user.is_manager()):
        query = query.filter(
            (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
        )
    
    # Apply filters
    if status_filter:
        if status_filter == 'overdue':
            query = query.filter(Task.due_date < date.today(), Task.status != 'completed')
        elif status_filter == 'due_soon':
            from datetime import timedelta
            soon = date.today() + timedelta(days=7)
            query = query.filter(Task.due_date >= date.today(), Task.due_date <= soon, Task.status != 'completed')
        else:
            query = query.filter_by(status=status_filter)
    
    if entity_filter:
        query = query.filter_by(entity_id=entity_filter)
    
    if tax_type_filter:
        query = query.join(TaskTemplate).filter(TaskTemplate.tax_type_id == tax_type_filter)
    
    if year_filter:
        query = query.filter_by(year=year_filter)
    
    tasks = query.order_by(Task.due_date).all()
    
    # Generate Excel
    excel_bytes = ExportService.export_tasks_to_excel(tasks, lang)
    
    filename = f"aufgaben_{date.today().strftime('%Y%m%d')}.xlsx" if lang == 'de' else f"tasks_{date.today().strftime('%Y%m%d')}.xlsx"
    
    return Response(
        excel_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/tasks/export/summary')
@login_required
def export_summary_report():
    """Export summary report to Excel"""
    from datetime import date
    from flask import Response
    
    lang = session.get('lang', 'de')
    
    # Get all tasks (with role-based filtering)
    query = Task.query
    
    if not (current_user.is_admin() or current_user.is_manager()):
        query = query.filter(
            (Task.owner_id == current_user.id) | (Task.reviewer_id == current_user.id)
        )
    
    # Apply year filter if provided
    year_filter = request.args.get('year', type=int, default=date.today().year)
    if year_filter:
        query = query.filter_by(year=year_filter)
    
    tasks = query.all()
    
    # Generate report
    excel_bytes = ExportService.export_summary_report(tasks, lang)
    
    filename = f"bericht_{year_filter}.xlsx" if lang == 'de' else f"report_{year_filter}.xlsx"
    
    return Response(
        excel_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/tasks/<int:task_id>/export/pdf')
@login_required
def export_task_pdf(task_id):
    """Export single task to PDF"""
    from flask import Response
    from datetime import date
    
    task = Task.query.get_or_404(task_id)
    lang = session.get('lang', 'de')
    
    # Check access
    if not (current_user.is_admin() or current_user.is_manager() or 
            task.owner_id == current_user.id or task.is_reviewer(current_user)):
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('task_list'))
    
    # Generate PDF
    pdf_bytes = ExportService.export_task_to_pdf(task, lang)
    
    # Sanitize filename
    safe_title = "".join(c for c in task.title if c.isalnum() or c in (' ', '-', '_')).strip()[:50]
    filename = f"{safe_title}_{date.today().strftime('%Y%m%d')}.pdf"
    
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ============================================================================
# CALENDAR SYNC (iCal) ROUTES
# ============================================================================

@app.route('/calendar/feed/<token>.ics')
def calendar_feed(token):
    """
    Public iCal feed endpoint for calendar subscription.
    No login required - authenticated via unique token.
    """
    from flask import Response
    
    # Find user by calendar token
    user = User.query.filter_by(calendar_token=token).first()
    if not user:
        abort(404)
    
    if not user.is_active:
        abort(403)
    
    # Get tasks visible to this user
    lang = request.args.get('lang', 'de')
    
    # Build task query based on user role
    if user.is_admin() or user.is_manager():
        # Admin/Manager sees all tasks
        tasks = Task.query.filter(
            Task.status != 'completed'  # Optionally exclude completed
        ).all()
    else:
        # Regular users see tasks they own or review
        user_teams = [team.id for team in user.get_teams()]
        tasks = Task.query.filter(
            db.or_(
                Task.owner_id == user.id,
                Task.owner_team_id.in_(user_teams) if user_teams else False,
                Task.reviewers.any(TaskReviewer.user_id == user.id)
            )
        ).all()
    
    # Generate iCal feed
    ical_bytes = CalendarService.generate_ical_feed(tasks, user.name, lang)
    
    return Response(
        ical_bytes,
        mimetype='text/calendar',
        headers={
            'Content-Disposition': f'attachment; filename="taxops-calendar.ics"',
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0'
        }
    )


@app.route('/calendar/subscription')
@login_required
def calendar_subscription():
    """Show user's calendar subscription URL and settings"""
    lang = session.get('lang', 'de')
    
    # Get or create user's calendar token
    token = current_user.get_or_create_calendar_token()
    
    # Build full subscription URL
    subscription_url = url_for('calendar_feed', token=token, _external=True)
    
    return render_template('calendar_subscription.html',
                           subscription_url=subscription_url,
                           lang=lang,
                           t=t)


@app.route('/calendar/regenerate-token', methods=['POST'])
@login_required
def regenerate_calendar_token():
    """Regenerate the user's calendar token (invalidates old subscription URL)"""
    new_token = current_user.regenerate_calendar_token()
    
    lang = session.get('lang', 'de')
    if lang == 'de':
        flash('Kalender-Token wurde erneuert. Bitte aktualisieren Sie Ihre Kalender-Abonnements.', 'success')
    else:
        flash('Calendar token regenerated. Please update your calendar subscriptions.', 'success')
    
    return redirect(url_for('calendar_subscription'))


@app.route('/profile/notifications', methods=['GET', 'POST'])
@login_required
def profile_notifications():
    """Manage user's email notification preferences"""
    lang = session.get('lang', 'de')
    
    if request.method == 'POST':
        # Update preferences
        current_user.email_notifications = 'email_notifications' in request.form
        current_user.email_on_assignment = 'email_on_assignment' in request.form
        current_user.email_on_status_change = 'email_on_status_change' in request.form
        current_user.email_on_due_reminder = 'email_on_due_reminder' in request.form
        current_user.email_on_comment = 'email_on_comment' in request.form
        
        db.session.commit()
        
        if lang == 'de':
            flash('E-Mail-Einstellungen gespeichert.', 'success')
        else:
            flash('Email preferences saved.', 'success')
        
        return redirect(url_for('profile_notifications'))
    
    return render_template('profile_notifications.html', 
                           user=current_user,
                           lang=lang, 
                           t=t)


# ============================================================================
# BULK OPERATIONS API ROUTES
# ============================================================================

@app.route('/api/tasks/bulk-status', methods=['POST'])
@login_required
def api_bulk_status():
    """Bulk change status for multiple tasks"""
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    new_status = data.get('status')
    
    if not task_ids or not new_status:
        return jsonify({'success': False, 'error': 'Missing task_ids or status'}), 400
    
    valid_statuses = ['draft', 'submitted', 'in_review', 'approved', 'completed', 'rejected']
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'error': 'Invalid status'}), 400
    
    success_count = 0
    error_count = 0
    
    for task_id in task_ids:
        task = Task.query.get(task_id)
        if not task:
            error_count += 1
            continue
        
        # Check permissions
        if not (current_user.is_admin() or current_user.is_manager() or 
                task.owner_id == current_user.id or task.is_reviewer(current_user)):
            error_count += 1
            continue
        
        # Try to transition
        if task.can_transition_to(new_status, current_user):
            try:
                old_status = task.transition_to(new_status, current_user)
                log_action('STATUS_CHANGE', 'Task', task.id, task.title, old_status, new_status)
                success_count += 1
            except ValueError:
                error_count += 1
        else:
            error_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'changed': success_count,
        'skipped': error_count,
        'message': f'{success_count} Aufgaben geändert, {error_count} übersprungen'
    })


@app.route('/api/tasks/bulk-assign-owner', methods=['POST'])
@login_required
def api_bulk_assign_owner():
    """Bulk assign owner to multiple tasks"""
    # Only admin/manager can bulk assign
    if not (current_user.is_admin() or current_user.is_manager()):
        return jsonify({'success': False, 'error': 'Keine Berechtigung'}), 403
    
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    owner_id = data.get('owner_id')
    
    if not task_ids or not owner_id:
        return jsonify({'success': False, 'error': 'Missing task_ids or owner_id'}), 400
    
    # Verify owner exists
    owner = User.query.get(owner_id)
    if not owner or not owner.is_active:
        return jsonify({'success': False, 'error': 'Invalid owner'}), 400
    
    success_count = 0
    lang = session.get('lang', 'de')
    notifications = []
    
    for task_id in task_ids:
        task = Task.query.get(task_id)
        if task:
            old_owner_id = task.owner_id
            task.owner_id = owner_id
            log_action('UPDATE', 'Task', task.id, task.title, 
                      f'Owner: {old_owner_id}', f'Owner: {owner_id}')
            
            # Notify new owner
            if owner_id != current_user.id:
                notification = NotificationService.notify_task_assigned(task, owner_id, current_user.id)
                notifications.append(notification)
            
            success_count += 1
    
    db.session.commit()
    
    if notifications:
        emit_notifications_to_users(notifications, lang)
    
    return jsonify({
        'success': True,
        'assigned': success_count,
        'message': f'{success_count} Aufgaben zugewiesen'
    })


@app.route('/api/tasks/bulk-delete', methods=['POST'])
@login_required
def api_bulk_delete():
    """Bulk delete multiple tasks (admin/manager only)"""
    if not (current_user.is_admin() or current_user.is_manager()):
        return jsonify({'success': False, 'error': 'Keine Berechtigung'}), 403
    
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    
    if not task_ids:
        return jsonify({'success': False, 'error': 'Missing task_ids'}), 400
    
    success_count = 0
    
    for task_id in task_ids:
        task = Task.query.get(task_id)
        if task:
            task_title = task.title
            
            # Delete related records
            TaskReviewer.query.filter_by(task_id=task_id).delete()
            TaskEvidence.query.filter_by(task_id=task_id).delete()
            Comment.query.filter_by(task_id=task_id).delete()
            AuditLog.query.filter_by(entity_type='Task', entity_id=task_id).delete()
            Notification.query.filter_by(entity_type='task', entity_id=task_id).delete()
            
            db.session.delete(task)
            log_action('DELETE', 'Task', task_id, task_title, '-', '-')
            success_count += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'deleted': success_count,
        'message': f'{success_count} Aufgaben gelöscht'
    })


# ============================================================================
# DASHBOARD CHART API ROUTES
# ============================================================================

@app.route('/api/dashboard/status-chart')
@login_required
def api_status_chart():
    """Get task distribution by status for pie chart"""
    from collections import Counter
    
    # Get base query based on user role
    if current_user.is_admin() or current_user.is_manager():
        tasks = Task.query.all()
    else:
        user_teams = [team.id for team in current_user.get_teams()]
        tasks = Task.query.filter(
            db.or_(
                Task.owner_id == current_user.id,
                Task.owner_team_id.in_(user_teams) if user_teams else False,
                Task.reviewers.any(TaskReviewer.user_id == current_user.id)
            )
        ).all()
    
    lang = session.get('lang', 'de')
    
    # Count by status
    status_counts = Counter(task.status for task in tasks)
    
    # Status labels and colors (Deloitte palette)
    status_config = {
        'draft': {'label_de': 'Entwurf', 'label_en': 'Draft', 'color': '#6C757D'},
        'submitted': {'label_de': 'Eingereicht', 'label_en': 'Submitted', 'color': '#0D6EFD'},
        'in_review': {'label_de': 'In Prüfung', 'label_en': 'In Review', 'color': '#0DCAF0'},
        'approved': {'label_de': 'Genehmigt', 'label_en': 'Approved', 'color': '#198754'},
        'completed': {'label_de': 'Abgeschlossen', 'label_en': 'Completed', 'color': '#86BC25'},
        'rejected': {'label_de': 'Abgelehnt', 'label_en': 'Rejected', 'color': '#DC3545'},
    }
    
    labels = []
    data = []
    colors = []
    
    for status, config in status_config.items():
        if status_counts.get(status, 0) > 0:
            label = config['label_de'] if lang == 'de' else config['label_en']
            labels.append(label)
            data.append(status_counts[status])
            colors.append(config['color'])
    
    return jsonify({
        'labels': labels,
        'data': data,
        'colors': colors
    })


@app.route('/api/dashboard/monthly-chart')
@login_required
def api_monthly_chart():
    """Get tasks by month for bar chart"""
    from collections import defaultdict
    from datetime import date
    
    year = request.args.get('year', date.today().year, type=int)
    
    # Get base query based on user role
    if current_user.is_admin() or current_user.is_manager():
        tasks = Task.query.filter(Task.year == year).all()
    else:
        user_teams = [team.id for team in current_user.get_teams()]
        tasks = Task.query.filter(
            Task.year == year,
            db.or_(
                Task.owner_id == current_user.id,
                Task.owner_team_id.in_(user_teams) if user_teams else False,
                Task.reviewers.any(TaskReviewer.user_id == current_user.id)
            )
        ).all()
    
    lang = session.get('lang', 'de')
    
    # Month names
    month_names_de = ['Jan', 'Feb', 'Mär', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']
    month_names_en = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    month_names = month_names_de if lang == 'de' else month_names_en
    
    # Count by month and status
    monthly_data = defaultdict(lambda: {'completed': 0, 'pending': 0, 'overdue': 0})
    today = date.today()
    
    for task in tasks:
        if task.due_date:
            month = task.due_date.month - 1  # 0-indexed
            # Handle both date and datetime objects
            task_date = task.due_date if isinstance(task.due_date, date) else task.due_date.date()
            if task.status == 'completed':
                monthly_data[month]['completed'] += 1
            elif task_date < today and task.status != 'completed':
                monthly_data[month]['overdue'] += 1
            else:
                monthly_data[month]['pending'] += 1
    
    # Build datasets
    completed = [monthly_data[m]['completed'] for m in range(12)]
    pending = [monthly_data[m]['pending'] for m in range(12)]
    overdue = [monthly_data[m]['overdue'] for m in range(12)]
    
    return jsonify({
        'labels': month_names,
        'datasets': [
            {
                'label': 'Abgeschlossen' if lang == 'de' else 'Completed',
                'data': completed,
                'backgroundColor': '#86BC25'
            },
            {
                'label': 'Ausstehend' if lang == 'de' else 'Pending',
                'data': pending,
                'backgroundColor': '#0DCAF0'
            },
            {
                'label': 'Überfällig' if lang == 'de' else 'Overdue',
                'data': overdue,
                'backgroundColor': '#DC3545'
            }
        ]
    })


@app.route('/api/dashboard/team-chart')
@login_required
def api_team_chart():
    """Get workload by team/owner for bar chart"""
    from collections import Counter
    
    # Only admins and managers can see team workload
    if not (current_user.is_admin() or current_user.is_manager()):
        return jsonify({'labels': [], 'data': [], 'colors': []})
    
    lang = session.get('lang', 'de')
    
    # Get active tasks (not completed)
    tasks = Task.query.filter(Task.status != 'completed').all()
    
    # Count by team
    team_counts = Counter()
    for task in tasks:
        if task.owner_team:
            team_name = task.owner_team.get_name(lang)
        elif task.owner:
            team_name = task.owner.name
        else:
            team_name = 'Nicht zugewiesen' if lang == 'de' else 'Unassigned'
        team_counts[team_name] += 1
    
    # Sort by count (top 10)
    sorted_teams = team_counts.most_common(10)
    
    labels = [t[0] for t in sorted_teams]
    data = [t[1] for t in sorted_teams]
    
    # Generate colors (Deloitte palette variations)
    base_colors = ['#86BC25', '#0D6EFD', '#0DCAF0', '#198754', '#FFC107', '#6C757D', '#DC3545', '#6610F2', '#D63384', '#20C997']
    colors = [base_colors[i % len(base_colors)] for i in range(len(labels))]
    
    return jsonify({
        'labels': labels,
        'data': data,
        'colors': colors
    })


# ============================================================================
# NOTIFICATION API ROUTES
# ============================================================================

@app.route('/api/notifications')
@login_required
def api_notifications():
    """Get recent notifications for current user"""
    lang = session.get('lang', 'de')
    limit = request.args.get('limit', 10, type=int)
    include_read = request.args.get('include_read', 'true').lower() == 'true'
    
    notifications = NotificationService.get_recent(
        current_user.id, 
        limit=min(limit, 50),  # Cap at 50
        include_read=include_read
    )
    
    return jsonify({
        'notifications': [n.to_dict(lang) for n in notifications],
        'unread_count': NotificationService.get_unread_count(current_user.id)
    })


@app.route('/api/notifications/unread-count')
@login_required
def api_notifications_unread_count():
    """Get unread notification count for badge"""
    return jsonify({
        'count': NotificationService.get_unread_count(current_user.id)
    })


@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def api_notification_mark_read(notification_id):
    """Mark a single notification as read"""
    success = NotificationService.mark_as_read(notification_id, current_user.id)
    if success:
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Notification not found'}), 404


@app.route('/api/notifications/mark-all-read', methods=['POST'])
@login_required
def api_notifications_mark_all_read():
    """Mark all notifications as read"""
    count = NotificationService.mark_all_as_read(current_user.id)
    db.session.commit()
    return jsonify({'success': True, 'count': count})


@app.route('/notifications')
@login_required
def notifications_page():
    """Full notifications page with pagination"""
    lang = session.get('lang', 'de')
    page = request.args.get('page', 1, type=int)
    
    pagination = NotificationService.get_all_paginated(current_user.id, page=page, per_page=20)
    
    return render_template('notifications.html',
        notifications=pagination.items,
        pagination=pagination,
        lang=lang
    )


# ============================================================================
# RUN
# ============================================================================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    # Use socketio.run() for WebSocket support with threading mode
    # Hot reload works with use_reloader=True in threading mode
    socketio.run(app, debug=True, host='0.0.0.0', port=5000, use_reloader=True, allow_unsafe_werkzeug=True)
