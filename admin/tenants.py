"""
Super-Admin Routes for Tenant Management

This blueprint provides routes for managing tenants (clients) in the multi-tenant system.
Only accessible by users with is_superadmin=True.
"""
import base64
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, session
from flask_login import login_required, current_user
from extensions import db
from models import Tenant, TenantMembership, User, TenantApiKey
from middleware import superadmin_required

admin_tenants = Blueprint('admin_tenants', __name__, url_prefix='/admin/tenants')


# ============================================================================
# TENANT LIST & DASHBOARD
# ============================================================================

@admin_tenants.route('/')
@login_required
@superadmin_required
def tenant_list():
    """List all tenants with statistics"""
    show_archived = request.args.get('show_archived', 'false') == 'true'
    
    query = Tenant.query
    if not show_archived:
        query = query.filter_by(is_archived=False)
    
    tenants = query.order_by(Tenant.name).all()
    
    # Calculate stats for each tenant
    tenant_stats = []
    for tenant in tenants:
        stats = {
            'tenant': tenant,
            'user_count': len(tenant.members),
            'admin_count': TenantMembership.query.filter_by(tenant_id=tenant.id, role='admin').count(),
        }
        tenant_stats.append(stats)
    
    return render_template('admin/tenants/list.html', 
                          tenant_stats=tenant_stats,
                          show_archived=show_archived)


@admin_tenants.route('/new', methods=['GET', 'POST'])
@login_required
@superadmin_required
def tenant_create():
    """Create a new tenant"""
    if request.method == 'POST':
        slug = request.form.get('slug', '').strip().lower()
        name = request.form.get('name', '').strip()
        
        # Validation
        if not slug:
            flash('Mandanten-Kennung ist erforderlich.', 'error')
            return render_template('admin/tenants/form.html', tenant=None)
        
        # Check unique slug
        if Tenant.query.filter_by(slug=slug).first():
            flash('Diese Mandanten-Kennung existiert bereits.', 'error')
            return render_template('admin/tenants/form.html', tenant=None)
        
        # Handle logo upload
        logo_data = None
        logo_mime_type = None
        
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename:
                # Read and encode as base64
                logo_data = base64.b64encode(file.read()).decode('utf-8')
                logo_mime_type = file.content_type
        
        tenant = Tenant(
            slug=slug,
            name=name or slug,
            logo_data=logo_data,
            logo_mime_type=logo_mime_type,
            is_active=True,
            created_by_id=current_user.id
        )
        
        db.session.add(tenant)
        db.session.commit()
        
        flash(f'Mandant "{tenant.name}" wurde erstellt.', 'success')
        return redirect(url_for('admin_tenants.tenant_detail', tenant_id=tenant.id))
    
    return render_template('admin/tenants/form.html', tenant=None)


@admin_tenants.route('/<int:tenant_id>')
@login_required
@superadmin_required
def tenant_detail(tenant_id):
    """View tenant details"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    # Get members with their roles
    memberships = TenantMembership.query.filter_by(tenant_id=tenant_id)\
        .join(User, TenantMembership.user_id == User.id)\
        .order_by(User.name).all()
    
    # Get API keys
    api_keys = TenantApiKey.query.filter_by(tenant_id=tenant_id).all()
    
    return render_template('admin/tenants/detail.html',
                          tenant=tenant,
                          memberships=memberships,
                          api_keys=api_keys)


@admin_tenants.route('/<int:tenant_id>/edit', methods=['GET', 'POST'])
@login_required
@superadmin_required
def tenant_edit(tenant_id):
    """Edit tenant settings"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        # Validation
        if not name:
            flash('Mandanten-Name ist erforderlich.', 'error')
            return render_template('admin/tenants/form.html', tenant=tenant)
        
        tenant.name = name
        tenant.is_active = is_active
        
        # Handle logo upload
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename:
                tenant.logo_data = base64.b64encode(file.read()).decode('utf-8')
                tenant.logo_mime_type = file.content_type
        
        # Handle logo removal
        if request.form.get('remove_logo') == 'true':
            tenant.logo_data = None
            tenant.logo_mime_type = None
        
        db.session.commit()
        
        flash(f'Mandant "{tenant.name}" wurde aktualisiert.', 'success')
        return redirect(url_for('admin_tenants.tenant_detail', tenant_id=tenant.id))
    
    return render_template('admin/tenants/form.html', tenant=tenant)


@admin_tenants.route('/<int:tenant_id>/archive', methods=['POST'])
@login_required
@superadmin_required
def tenant_archive(tenant_id):
    """Archive a tenant (soft delete)"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    tenant.is_archived = True
    tenant.archived_at = datetime.utcnow()
    tenant.archived_by_id = current_user.id
    tenant.is_active = False
    
    db.session.commit()
    
    flash(f'Mandant "{tenant.name}" wurde archiviert.', 'success')
    return redirect(url_for('admin_tenants.tenant_list'))


@admin_tenants.route('/<int:tenant_id>/restore', methods=['POST'])
@login_required
@superadmin_required
def tenant_restore(tenant_id):
    """Restore an archived tenant"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    tenant.is_archived = False
    tenant.archived_at = None
    tenant.archived_by_id = None
    tenant.is_active = True
    
    db.session.commit()
    
    flash(f'Mandant "{tenant.name}" wurde wiederhergestellt.', 'success')
    return redirect(url_for('admin_tenants.tenant_detail', tenant_id=tenant.id))


@admin_tenants.route('/<int:tenant_id>/delete', methods=['POST'])
@login_required
@superadmin_required
def tenant_delete(tenant_id):
    """Permanently delete an archived tenant"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    if not tenant.is_archived:
        flash('Nur archivierte Mandanten können gelöscht werden.', 'error')
        return redirect(url_for('admin_tenants.tenant_detail', tenant_id=tenant.id))
    
    tenant_name = tenant.name
    
    # Delete all memberships first
    TenantMembership.query.filter_by(tenant_id=tenant_id).delete()
    TenantApiKey.query.filter_by(tenant_id=tenant_id).delete()
    
    db.session.delete(tenant)
    db.session.commit()
    
    flash(f'Mandant "{tenant_name}" wurde endgültig gelöscht.', 'success')
    return redirect(url_for('admin_tenants.tenant_list'))


# ============================================================================
# TENANT MEMBER MANAGEMENT
# ============================================================================

@admin_tenants.route('/<int:tenant_id>/members')
@login_required
@superadmin_required
def tenant_members(tenant_id):
    """Manage tenant members"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    memberships = TenantMembership.query.filter_by(tenant_id=tenant_id)\
        .join(User, TenantMembership.user_id == User.id)\
        .order_by(User.name).all()
    
    # Get users not in this tenant
    member_ids = [m.user_id for m in memberships]
    available_users = User.query.filter(
        User.id.notin_(member_ids) if member_ids else True,
        User.is_active == True
    ).order_by(User.name).all()
    
    return render_template('admin/tenants/members.html',
                          tenant=tenant,
                          memberships=memberships,
                          available_users=available_users)


@admin_tenants.route('/<int:tenant_id>/members/add', methods=['POST'])
@login_required
@superadmin_required
def tenant_add_member(tenant_id):
    """Add a user to tenant"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    user_id = request.form.get('user_id', type=int)
    role = request.form.get('role', 'member')
    
    if not user_id:
        flash('Benutzer ist erforderlich.', 'error')
        return redirect(url_for('admin_tenants.tenant_members', tenant_id=tenant_id))
    
    user = User.query.get_or_404(user_id)
    
    # Check if already member
    existing = TenantMembership.query.filter_by(
        tenant_id=tenant_id, user_id=user_id
    ).first()
    
    if existing:
        flash(f'{user.name} ist bereits Mitglied.', 'warning')
        return redirect(url_for('admin_tenants.tenant_members', tenant_id=tenant_id))
    
    membership = TenantMembership(
        tenant_id=tenant_id,
        user_id=user_id,
        role=role
    )
    db.session.add(membership)
    db.session.commit()
    
    flash(f'{user.name} wurde als {role} hinzugefügt.', 'success')
    return redirect(url_for('admin_tenants.tenant_members', tenant_id=tenant_id))


@admin_tenants.route('/<int:tenant_id>/members/<int:user_id>/update', methods=['POST'])
@login_required
@superadmin_required
def tenant_update_member(tenant_id, user_id):
    """Update member role"""
    membership = TenantMembership.query.filter_by(
        tenant_id=tenant_id, user_id=user_id
    ).first_or_404()
    
    role = request.form.get('role', 'member')
    membership.role = role
    db.session.commit()
    
    flash(f'Rolle wurde auf {role} geändert.', 'success')
    return redirect(url_for('admin_tenants.tenant_members', tenant_id=tenant_id))


@admin_tenants.route('/<int:tenant_id>/members/<int:user_id>/remove', methods=['POST'])
@login_required
@superadmin_required
def tenant_remove_member(tenant_id, user_id):
    """Remove user from tenant"""
    membership = TenantMembership.query.filter_by(
        tenant_id=tenant_id, user_id=user_id
    ).first_or_404()
    
    user_name = membership.user.name
    db.session.delete(membership)
    db.session.commit()
    
    flash(f'{user_name} wurde entfernt.', 'success')
    return redirect(url_for('admin_tenants.tenant_members', tenant_id=tenant_id))


# ============================================================================
# TENANT IMPERSONATION (Super-Admin views as tenant)
# ============================================================================

@admin_tenants.route('/<int:tenant_id>/enter', methods=['POST'])
@login_required
@superadmin_required
def enter_tenant(tenant_id):
    """Enter a tenant context (Super-Admin only)"""
    tenant = Tenant.query.get_or_404(tenant_id)
    
    session['current_tenant_id'] = tenant_id
    current_user.current_tenant_id = tenant_id
    db.session.commit()
    
    flash(f'Sie arbeiten jetzt im Mandanten: {tenant.name}', 'info')
    return redirect(url_for('dashboard'))


# ============================================================================
# TENANT EXPORT
# ============================================================================

@admin_tenants.route('/<int:tenant_id>/export', methods=['POST'])
@login_required
@superadmin_required
def tenant_export(tenant_id):
    """Export tenant data as JSON"""
    import json
    from flask import Response
    
    tenant = Tenant.query.get_or_404(tenant_id)
    
    # Build export data
    export_data = {
        'tenant': {
            'slug': tenant.slug,
            'name': tenant.name,
            'is_active': tenant.is_active,
            'is_archived': tenant.is_archived,
            'created_at': tenant.created_at.isoformat() if tenant.created_at else None,
        },
        'members': [
            {
                'user_email': m.user.email,
                'user_name': m.user.name,
                'role': m.role,
                'joined_at': m.joined_at.isoformat() if m.joined_at else None
            }
            for m in tenant.memberships
        ],
        'exported_at': datetime.utcnow().isoformat(),
        'exported_by': current_user.email
    }
    
    # TODO: Add export of tenant-specific data (tasks, entities, etc.)
    
    response = Response(
        json.dumps(export_data, indent=2, ensure_ascii=False),
        mimetype='application/json',
        headers={
            'Content-Disposition': f'attachment; filename={tenant.slug}_export.json'
        }
    )
    return response


@admin_tenants.route('/<int:tenant_id>/export-excel', methods=['GET', 'POST'])
@login_required
@superadmin_required
def tenant_export_excel(tenant_id):
    """Export tenant data as Excel"""
    from io import BytesIO
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    tenant = Tenant.query.get_or_404(tenant_id)
    
    # Create workbook
    wb = Workbook()
    
    # --- Sheet 1: Tenant Info ---
    ws_info = wb.active
    ws_info.title = "Mandant Info"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0076A8", end_color="0076A8", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tenant info
    info_data = [
        ("Eigenschaft", "Wert"),
        ("Slug", tenant.slug),
        ("Name", tenant.name),
        ("Status", "Aktiv" if tenant.is_active else "Inaktiv"),
        ("Archiviert", "Ja" if tenant.is_archived else "Nein"),
        ("Erstellt am", tenant.created_at.strftime("%d.%m.%Y %H:%M") if tenant.created_at else "-"),
        ("Mitglieder", str(len(tenant.memberships))),
        ("Exportiert am", datetime.utcnow().strftime("%d.%m.%Y %H:%M")),
        ("Exportiert von", current_user.email),
    ]
    
    for row_idx, (key, value) in enumerate(info_data, 1):
        ws_info.cell(row=row_idx, column=1, value=key).border = border
        ws_info.cell(row=row_idx, column=2, value=value).border = border
        if row_idx == 1:
            ws_info.cell(row=row_idx, column=1).font = header_font
            ws_info.cell(row=row_idx, column=1).fill = header_fill
            ws_info.cell(row=row_idx, column=2).font = header_font
            ws_info.cell(row=row_idx, column=2).fill = header_fill
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 40
    
    # --- Sheet 2: Members ---
    ws_members = wb.create_sheet("Mitglieder")
    
    member_headers = ["E-Mail", "Name", "Rolle", "Beigetreten am"]
    for col_idx, header in enumerate(member_headers, 1):
        cell = ws_members.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, m in enumerate(tenant.memberships, 2):
        ws_members.cell(row=row_idx, column=1, value=m.user.email).border = border
        ws_members.cell(row=row_idx, column=2, value=m.user.name).border = border
        ws_members.cell(row=row_idx, column=3, value=m.role.title()).border = border
        ws_members.cell(row=row_idx, column=4, value=m.joined_at.strftime("%d.%m.%Y") if m.joined_at else "-").border = border
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(member_headers, 1):
        ws_members.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # --- Sheet 3: Entities (if any) ---
    from models import Entity
    entities = Entity.query.filter(
        (Entity.tenant_id == tenant.id) | (Entity.tenant_id.is_(None))
    ).all()
    
    if entities:
        ws_entities = wb.create_sheet("Gesellschaften")
        entity_headers = ["ID", "Name"]
        for col_idx, header in enumerate(entity_headers, 1):
            cell = ws_entities.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, entity in enumerate(entities, 2):
            ws_entities.cell(row=row_idx, column=1, value=entity.id).border = border
            ws_entities.cell(row=row_idx, column=2, value=entity.name).border = border
        
        ws_entities.column_dimensions['A'].width = 10
        ws_entities.column_dimensions['B'].width = 40
    
    # --- Sheet 4: Projects ---
    from modules.projects.models import Project, Issue, IssueComment, IssueActivity, Sprint
    
    projects = Project.query.filter(
        (Project.tenant_id == tenant.id) | (Project.tenant_id.is_(None))
    ).order_by(Project.key).all()
    
    if projects:
        ws_projects = wb.create_sheet("Projekte")
        project_headers = ["Key", "Name", "Kategorie", "Methodologie", "Lead", "Status", "Erstellt am"]
        for col_idx, header in enumerate(project_headers, 1):
            cell = ws_projects.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, project in enumerate(projects, 2):
            ws_projects.cell(row=row_idx, column=1, value=project.key).border = border
            ws_projects.cell(row=row_idx, column=2, value=project.name).border = border
            ws_projects.cell(row=row_idx, column=3, value=project.category or "-").border = border
            ws_projects.cell(row=row_idx, column=4, value=project.methodology or "scrum").border = border
            ws_projects.cell(row=row_idx, column=5, value=project.lead.name if project.lead else "-").border = border
            ws_projects.cell(row=row_idx, column=6, value="Archiviert" if project.is_archived else "Aktiv").border = border
            ws_projects.cell(row=row_idx, column=7, value=project.created_at.strftime("%d.%m.%Y") if project.created_at else "-").border = border
        
        for col_idx in range(1, len(project_headers) + 1):
            ws_projects.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # --- Sheet 5: Items (Issues) ---
    issues = Issue.query.filter(
        (Issue.tenant_id == tenant.id) | (Issue.tenant_id.is_(None))
    ).order_by(Issue.created_at.desc()).all()
    
    if issues:
        ws_issues = wb.create_sheet("Items")
        issue_headers = ["Projekt", "Key", "Typ", "Status", "Zusammenfassung", "Priorität", "Story Points",
                         "Bearbeiter", "Reporter", "Fällig am", "Erstellt am", "Aktualisiert am", "Labels"]
        for col_idx, header in enumerate(issue_headers, 1):
            cell = ws_issues.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        priority_names = {1: "Höchste", 2: "Hoch", 3: "Mittel", 4: "Niedrig", 5: "Niedrigste"}
        
        for row_idx, issue in enumerate(issues, 2):
            ws_issues.cell(row=row_idx, column=1, value=issue.project.key if issue.project else "-").border = border
            ws_issues.cell(row=row_idx, column=2, value=issue.key or f"#{issue.id}").border = border
            ws_issues.cell(row=row_idx, column=3, value=issue.issue_type.name if issue.issue_type else "-").border = border
            ws_issues.cell(row=row_idx, column=4, value=issue.status.name if issue.status else "-").border = border
            ws_issues.cell(row=row_idx, column=5, value=issue.summary or "").border = border
            ws_issues.cell(row=row_idx, column=6, value=priority_names.get(issue.priority, str(issue.priority))).border = border
            ws_issues.cell(row=row_idx, column=7, value=issue.story_points or "-").border = border
            ws_issues.cell(row=row_idx, column=8, value=issue.assignee.name if issue.assignee else "-").border = border
            ws_issues.cell(row=row_idx, column=9, value=issue.reporter.name if issue.reporter else "-").border = border
            ws_issues.cell(row=row_idx, column=10, value=issue.due_date.strftime("%d.%m.%Y") if issue.due_date else "-").border = border
            ws_issues.cell(row=row_idx, column=11, value=issue.created_at.strftime("%d.%m.%Y %H:%M") if issue.created_at else "-").border = border
            ws_issues.cell(row=row_idx, column=12, value=issue.updated_at.strftime("%d.%m.%Y %H:%M") if issue.updated_at else "-").border = border
            # Convert labels list to comma-separated string
            labels_str = ", ".join(issue.labels) if issue.labels and isinstance(issue.labels, list) else (issue.labels or "-")
            ws_issues.cell(row=row_idx, column=13, value=labels_str).border = border
        
        col_widths = [10, 12, 12, 15, 50, 12, 12, 20, 20, 12, 18, 18, 30]
        for col_idx, width in enumerate(col_widths, 1):
            ws_issues.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Sheet 6: Iterations (Sprints) ---
    sprints = Sprint.query.join(Project).filter(
        (Project.tenant_id == tenant.id) | (Project.tenant_id.is_(None))
    ).order_by(Sprint.start_date.desc()).all()
    
    if sprints:
        ws_sprints = wb.create_sheet("Iterationen")
        sprint_headers = ["Projekt", "Name", "Ziel", "Status", "Startdatum", "Enddatum", "Items geplant", "Items abgeschlossen"]
        for col_idx, header in enumerate(sprint_headers, 1):
            cell = ws_sprints.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        status_names = {'planning': 'Planung', 'active': 'Aktiv', 'completed': 'Abgeschlossen', 'cancelled': 'Abgebrochen'}
        
        for row_idx, sprint in enumerate(sprints, 2):
            planned_count = Issue.query.filter_by(sprint_id=sprint.id).count()
            done_count = Issue.query.join(Issue.status).filter(
                Issue.sprint_id == sprint.id,
                Issue.status.has(category='done')
            ).count() if hasattr(Issue, 'status') else 0
            
            ws_sprints.cell(row=row_idx, column=1, value=sprint.project.key if sprint.project else "-").border = border
            ws_sprints.cell(row=row_idx, column=2, value=sprint.name or "").border = border
            ws_sprints.cell(row=row_idx, column=3, value=sprint.goal or "-").border = border
            ws_sprints.cell(row=row_idx, column=4, value=status_names.get(sprint.state, sprint.state or "-")).border = border
            ws_sprints.cell(row=row_idx, column=5, value=sprint.start_date.strftime("%d.%m.%Y") if sprint.start_date else "-").border = border
            ws_sprints.cell(row=row_idx, column=6, value=sprint.end_date.strftime("%d.%m.%Y") if sprint.end_date else "-").border = border
            ws_sprints.cell(row=row_idx, column=7, value=planned_count).border = border
            ws_sprints.cell(row=row_idx, column=8, value=done_count).border = border
        
        col_widths = [10, 30, 40, 15, 12, 12, 15, 18]
        for col_idx, width in enumerate(col_widths, 1):
            ws_sprints.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Sheet 7: Comments ---
    comments = IssueComment.query.join(Issue).filter(
        (Issue.tenant_id == tenant.id) | (Issue.tenant_id.is_(None))
    ).order_by(IssueComment.created_at.desc()).all()
    
    if comments:
        ws_comments = wb.create_sheet("Kommentare")
        comment_headers = ["Item Key", "Projekt", "Autor", "Kommentar", "Erstellt am"]
        for col_idx, header in enumerate(comment_headers, 1):
            cell = ws_comments.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, comment in enumerate(comments, 2):
            ws_comments.cell(row=row_idx, column=1, value=comment.issue.key if comment.issue else "-").border = border
            ws_comments.cell(row=row_idx, column=2, value=comment.issue.project.key if comment.issue and comment.issue.project else "-").border = border
            ws_comments.cell(row=row_idx, column=3, value=comment.author.name if comment.author else "-").border = border
            # Truncate long comments
            content = comment.content or ""
            if len(content) > 500:
                content = content[:500] + "..."
            ws_comments.cell(row=row_idx, column=4, value=content).border = border
            ws_comments.cell(row=row_idx, column=5, value=comment.created_at.strftime("%d.%m.%Y %H:%M") if comment.created_at else "-").border = border
        
        col_widths = [12, 10, 25, 80, 18]
        for col_idx, width in enumerate(col_widths, 1):
            ws_comments.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Sheet 8: Activity Log (Vorgänge) ---
    activities = IssueActivity.query.join(Issue).filter(
        (Issue.tenant_id == tenant.id) | (Issue.tenant_id.is_(None))
    ).order_by(IssueActivity.created_at.desc()).limit(5000).all()  # Limit for performance
    
    if activities:
        ws_activities = wb.create_sheet("Aktivitätsprotokoll")
        activity_headers = ["Datum/Uhrzeit", "Item Key", "Projekt", "Benutzer", "Aktion", "Feld", "Alter Wert", "Neuer Wert"]
        for col_idx, header in enumerate(activity_headers, 1):
            cell = ws_activities.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        action_names = {
            'created': 'Erstellt',
            'field_update': 'Aktualisiert',
            'status_change': 'Status geändert',
            'assignee_change': 'Zugewiesen',
            'comment': 'Kommentiert',
            'attachment': 'Anhang',
            'link': 'Verknüpft',
            'worklog': 'Arbeitszeit',
            'reviewer_added': 'Prüfer hinzugefügt',
            'approved': 'Genehmigt',
            'rejected': 'Abgelehnt',
        }
        
        for row_idx, activity in enumerate(activities, 2):
            ws_activities.cell(row=row_idx, column=1, value=activity.created_at.strftime("%d.%m.%Y %H:%M:%S") if activity.created_at else "-").border = border
            ws_activities.cell(row=row_idx, column=2, value=activity.issue.key if activity.issue else "-").border = border
            ws_activities.cell(row=row_idx, column=3, value=activity.issue.project.key if activity.issue and activity.issue.project else "-").border = border
            ws_activities.cell(row=row_idx, column=4, value=activity.user.name if activity.user else "System").border = border
            ws_activities.cell(row=row_idx, column=5, value=action_names.get(activity.activity_type, activity.activity_type or "-")).border = border
            ws_activities.cell(row=row_idx, column=6, value=activity.field_name or "-").border = border
            old_val = str(activity.old_value or "-")[:100]
            new_val = str(activity.new_value or "-")[:100]
            ws_activities.cell(row=row_idx, column=7, value=old_val).border = border
            ws_activities.cell(row=row_idx, column=8, value=new_val).border = border
        
        col_widths = [18, 12, 10, 25, 18, 15, 25, 25]
        for col_idx, width in enumerate(col_widths, 1):
            ws_activities.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Sheet 9: Tasks (ProjectOps Tasks) ---
    from models import Task
    tasks = Task.query.filter(
        (Task.tenant_id == tenant.id) | (Task.tenant_id.is_(None))
    ).order_by(Task.due_date.desc()).all()
    
    if tasks:
        ws_tasks = wb.create_sheet("Aufgaben (Kalender)")
        task_headers = ["ID", "Titel", "Gesellschaft", "Steuerart", "Status", "Fällig am", 
                        "Bearbeiter", "Verantwortlicher", "Team", "Erstellt am"]
        for col_idx, header in enumerate(task_headers, 1):
            cell = ws_tasks.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        status_names = {'open': 'Offen', 'in_progress': 'In Bearbeitung', 'in_review': 'In Prüfung', 
                        'completed': 'Abgeschlossen', 'archived': 'Archiviert'}
        
        for row_idx, task in enumerate(tasks, 2):
            ws_tasks.cell(row=row_idx, column=1, value=task.id).border = border
            ws_tasks.cell(row=row_idx, column=2, value=task.title or "").border = border
            ws_tasks.cell(row=row_idx, column=3, value=task.entity.name if task.entity else "-").border = border
            ws_tasks.cell(row=row_idx, column=4, value=task.tax_type.code if task.tax_type else "-").border = border
            ws_tasks.cell(row=row_idx, column=5, value=status_names.get(task.status, task.status or "-")).border = border
            ws_tasks.cell(row=row_idx, column=6, value=task.due_date.strftime("%d.%m.%Y") if task.due_date else "-").border = border
            ws_tasks.cell(row=row_idx, column=7, value=task.assignee.name if task.assignee else "-").border = border
            ws_tasks.cell(row=row_idx, column=8, value=task.responsible.name if task.responsible else "-").border = border
            ws_tasks.cell(row=row_idx, column=9, value=task.team.name if task.team else "-").border = border
            ws_tasks.cell(row=row_idx, column=10, value=task.created_at.strftime("%d.%m.%Y") if task.created_at else "-").border = border
        
        col_widths = [8, 40, 25, 12, 15, 12, 20, 20, 20, 12]
        for col_idx, width in enumerate(col_widths, 1):
            ws_tasks.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Sheet 10: Teams ---
    from models import Team
    teams = Team.query.filter(
        (Team.tenant_id == tenant.id) | (Team.tenant_id.is_(None))
    ).all()
    
    if teams:
        ws_teams = wb.create_sheet("Teams")
        team_headers = ["ID", "Name", "Name (EN)", "Mitglieder"]
        for col_idx, header in enumerate(team_headers, 1):
            cell = ws_teams.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, team in enumerate(teams, 2):
            # Handle both dynamic queries and lists
            if hasattr(team, 'members'):
                member_count = team.members.count() if hasattr(team.members, 'count') else len(list(team.members))
            else:
                member_count = 0
            ws_teams.cell(row=row_idx, column=1, value=team.id).border = border
            ws_teams.cell(row=row_idx, column=2, value=team.name or "").border = border
            ws_teams.cell(row=row_idx, column=3, value=team.name_en or "-").border = border
            ws_teams.cell(row=row_idx, column=4, value=member_count).border = border
        
        col_widths = [8, 30, 30, 12]
        for col_idx, width in enumerate(col_widths, 1):
            ws_teams.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Filename with timestamp for compliance
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{tenant.slug}_compliance_export_{timestamp}.xlsx"
    
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}'
        }
    )
    return response
